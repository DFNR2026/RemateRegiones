"""
Fix: Revertir 10 causas contaminadas por F3 buggy (PDF incorrecto).
Restaura de PENDIENTE_REVISION_MANUAL a PENDIENTE_LIQUIDACION.

Uso: python fix_revert_f3.py
"""

import openpyxl
from config import EXCEL_MADRE

print(f"Leyendo {EXCEL_MADRE}...")
wb = openpyxl.load_workbook(EXCEL_MADRE)

ws = wb["_datos_internos"]
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

col_rol = headers.index("ROL") + 1
col_anio = headers.index("AÑO") + 1
col_estado = headers.index("estado") + 1
col_log = headers.index("log_decision") + 1
col_ruta = headers.index("ruta_liquidacion") + 1

revertidas = 0
for row in range(2, ws.max_row + 1):
    estado = str(ws.cell(row=row, column=col_estado).value or "").strip()
    log_dec = str(ws.cell(row=row, column=col_log).value or "")

    if estado != "PENDIENTE_REVISION_MANUAL":
        continue
    if "F3: Liquidacion descargada pero saldo" not in log_dec:
        continue

    rol = str(ws.cell(row=row, column=col_rol).value or "").strip()
    anio = str(ws.cell(row=row, column=col_anio).value or "").strip()

    ws.cell(row=row, column=col_estado).value = "PENDIENTE_LIQUIDACION"
    ws.cell(row=row, column=col_log).value = (
        "REVERT: Restaurada a PENDIENTE_LIQUIDACION (F3 bug: PDF incorrecto). "
        "Senal 4: keyword liquidacion detectada en tabla OJV"
    )
    ws.cell(row=row, column=col_ruta).value = ""

    print(f"  REVERT: C-{rol}-{anio} -> PENDIENTE_LIQUIDACION")
    revertidas += 1

print(f"\nTotal revertidas: {revertidas}")

if revertidas > 0:
    for intento in range(3):
        try:
            wb.save(EXCEL_MADRE)
            print(f"Excel guardado: {EXCEL_MADRE}")
            break
        except PermissionError:
            if intento < 2:
                print("  PermissionError. Cierra el Excel y presiona Enter...")
                input()
            else:
                backup = EXCEL_MADRE.replace(".xlsx", "_backup.xlsx")
                wb.save(backup)
                print(f"  No se pudo guardar. Backup en: {backup}")
else:
    print("No se encontraron causas para revertir.")

wb.close()
