"""
procesar_audit4.py
==================
Aplica las decisiones del CUARTO_INFORME (Audit4) al Excel madre
'Causas_posible_saldo.xlsx'.

USO:
    python procesar_audit4.py            # DRY-RUN (no guarda)
    python procesar_audit4.py --apply    # EJECUCION REAL (escribe + backup)

QUE HACE:
1. Backup automatico del Excel con timestamp (solo en --apply).
2. Asegura que existen las columnas 'observacion_abogado' y
   'fecha_revision_abogado' en la hoja '_datos_internos' (las crea si faltan).
   (Nota: estas columnas se perdieron en un run del 28-abr; los datos del
    Audit3 sobreviven en log_decision. Este script las recrea.)
3. Para cada una de las 48 causas del Audit4:
   - 8 SI    -> mantiene estado, pobla observacion_abogado +
                fecha_revision_abogado, pre-pendea log_decision con
                "Audit4 [2026-04-29]: SI - <obs>"
   - 40 NO   -> estado = ELIMINAR (decision humana, NO se borra fisicamente),
                pre-pendea log_decision con "Audit4 [2026-04-29]: NO[ - <obs>]"
                (NO toca observacion_abogado ni fecha_revision_abogado, segun
                 instruccion del prompt)

IDEMPOTENCIA:
- Si ya existe una entrada "Audit4 [2026-04-29]" en el log de esa causa,
  NO la duplica.

ELIMINAR (sin D) = decision humana, permanece en el Excel.
NO confundir con ELIMINADA (decision automatica del filtrador, se borra fisico).

REQUIERE:
- openpyxl (ya instalado).
- Excel CERRADO al ejecutar (reintenta 3 veces si esta abierto).

NO MODIFICA pestanas derivadas (se regeneran en el proximo run del filtrador;
la fuente de verdad es _datos_internos).
"""

import sys
import shutil
import time
import argparse
from datetime import datetime
from pathlib import Path

# ----------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------
EXCEL_PATH = Path(r"D:\Remates\Causas con liq\Causas_posible_saldo.xlsx")
SHEET = "_datos_internos"
FECHA_AUDIT = "2026-04-29"            # fecha del informe (hardcoded, deterministico)
PREFIJO_LOG = f"Audit4 [{FECHA_AUDIT}]"

COLS_NUEVAS = ["observacion_abogado", "fecha_revision_abogado"]

# Decisiones del CUARTO_INFORME devuelto por el abogado.
# Formato: (rol, decision, observacion)   decision: "SI" o "NO"
# Observaciones ya vienen en ASCII (sin tildes) desde el prompt.
DECISIONES = [
    # --- 8 SI: mantener estado + persistir observacion --------------------
    ("C-881-2023",  "SI", "No se visualiza acta de remate el dia 9 de abril ni en fechas posteriores. Esperar liquidacion"),
    ("C-2100-2024", "SI", "Habra excedente pero hay un tercerista (Banco Estado) y el demandado tiene abogado de calle Ahumada (Camilo Cordova). Esperar liquidacion y giros de dineros antes de actuar"),
    ("C-1993-2024", "SI", "Esperar liquidacion ya que no se encuentra acta de remate"),
    ("C-1133-2024", "SI", "Esperar liquidacion por posible saldo pero hay dos deudas: un pagare de aprox 45 millones mas una deuda hipotecaria aun desconocida. Demandada con patrocinio de abogados"),
    ("C-48-2022",   "SI", "Esperar liquidacion ya que hay un credito hipotecario de por medio. Ojo que demandado tiene abogado en abril 2026"),
    ("C-1914-2025", "SI", "Valor comercial 200 millones, no hubo postores, esperar que baje el minimo hasta que se remate"),
    ("C-4055-2023", "SI", "Remate suspendido pero podria ser interesante"),
    ("C-4306-2024", "SI", "Con abogado el demandado en abril 2026, esperar liquidacion. Ojo que el banco aun persigue una deuda de aprox 15 millones"),

    # --- 40 NO: estado = ELIMINAR ----------------------------------------
    # 7 con observacion manuscrita:
    ("C-3103-2025", "NO", "Da cuenta del pago total de la deuda"),
    ("C-102-2025",  "NO", "Demandado cancelo integramente"),
    ("C-2255-2025", "NO", "Deuda pagada en su totalidad"),
    ("C-3308-2025", "NO", "Avenimiento/transaccion folio 19 apremio"),
    ("C-2331-2024", "NO", "Cancelado todo lo adeudado, suspension del remate"),
    ("C-2885-2023", "NO", "Se adjudica con cargo al credito"),
    ("C-1408-2024", "NO", "Deuda 5.063.107, remate se adjudico en 4.800.000, no habra excedente"),
    # 33 sin observacion:
    ("C-130-2025",  "NO", ""),
    ("C-273-2021",  "NO", ""),
    ("C-2758-2025", "NO", ""),
    ("C-3733-2023", "NO", ""),
    ("C-3755-2022", "NO", ""),
    ("C-3269-2025", "NO", ""),
    ("C-70-2025",   "NO", ""),
    ("C-2281-2025", "NO", ""),
    ("C-3698-2024", "NO", ""),
    ("C-1526-2025", "NO", ""),
    ("C-589-2025",  "NO", ""),
    ("C-2120-2025", "NO", ""),
    ("C-3819-2025", "NO", ""),
    ("C-305-2025",  "NO", ""),
    ("C-858-2022",  "NO", ""),
    ("C-2729-2023", "NO", ""),
    ("C-2333-2024", "NO", ""),
    ("C-1813-2025", "NO", ""),
    ("C-1993-2025", "NO", ""),
    ("C-3268-2025", "NO", ""),
    ("C-2145-2025", "NO", ""),
    ("C-2748-2025", "NO", ""),
    ("C-2304-2025", "NO", ""),
    ("C-3155-2025", "NO", ""),
    ("C-2038-2024", "NO", ""),
    ("C-1716-2021", "NO", ""),
    ("C-1070-2025", "NO", ""),
    ("C-2441-2025", "NO", ""),
    ("C-2038-2025", "NO", ""),
    ("C-2829-2018", "NO", ""),
    ("C-6592-2020", "NO", ""),
    ("C-15-2024",   "NO", ""),
    ("C-1485-2025", "NO", ""),
]


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def log(msg):
    """Print solo ASCII (Windows cp1252)."""
    safe = msg.encode("ascii", errors="replace").decode("ascii")
    print(safe)


def backup(path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = path.with_name(f"{path.stem}_backup_audit4_{ts}{path.suffix}")
    shutil.copy(path, dst)
    return dst


def cargar_excel_con_retry(path: Path, intentos: int = 3, espera: float = 2.0):
    from openpyxl import load_workbook
    for i in range(intentos):
        try:
            return load_workbook(path)
        except PermissionError:
            log(f"  Excel abierto, reintento {i+1}/{intentos} en {espera}s...")
            time.sleep(espera)
    raise PermissionError(f"No se pudo abrir {path}. Cierre Excel y reintente.")


def guardar_excel_con_retry(wb, path: Path, intentos: int = 3, espera: float = 2.0):
    for i in range(intentos):
        try:
            wb.save(path)
            return
        except PermissionError:
            log(f"  Excel abierto al guardar, reintento {i+1}/{intentos} en {espera}s...")
            time.sleep(espera)
    raise PermissionError(f"No se pudo guardar {path}. Cierre Excel y reintente.")


def construir_clave(rol_raw, ano):
    """Normaliza ROL+AAO a 'C-XXXX-YYYY'."""
    if rol_raw is None or ano is None:
        return None
    rol_str = str(rol_raw).strip()
    if rol_str.upper().startswith("C-"):
        num = rol_str[2:].split("-")[0]
    else:
        num = rol_str
    try:
        return f"C-{num}-{int(float(ano))}"
    except (ValueError, TypeError):
        return f"C-{num}-{str(ano).strip()}"


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main(dry_run: bool = True):
    n_si = sum(1 for _, d, _ in DECISIONES if d == "SI")
    n_no = sum(1 for _, d, _ in DECISIONES if d == "NO")

    log("=" * 70)
    log(f"procesar_audit4.py  (fecha informe: {FECHA_AUDIT})")
    log(f"Modo: {'DRY-RUN (no guarda)' if dry_run else 'EJECUCION REAL'}")
    log(f"Decisiones cargadas: {len(DECISIONES)}  ({n_si} SI / {n_no} NO)")
    log("=" * 70)

    if not EXCEL_PATH.exists():
        log(f"ERROR: no existe {EXCEL_PATH}")
        sys.exit(1)

    # 1) Backup
    if not dry_run:
        bak = backup(EXCEL_PATH)
        log(f"\n[1] Backup creado: {bak.name}")
    else:
        log("\n[1] Backup omitido (dry-run)")

    # 2) Cargar
    log(f"\n[2] Cargando {EXCEL_PATH.name}...")
    wb = cargar_excel_con_retry(EXCEL_PATH)
    if SHEET not in wb.sheetnames:
        log(f"ERROR: no existe la hoja '{SHEET}' en {EXCEL_PATH}")
        sys.exit(1)
    ws = wb[SHEET]

    # 3) Indexar columnas
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # Buscar AAO independientemente del encoding
    ano_col = None
    for h in headers:
        if h and str(h).strip().upper() in ("AÑO", "ANO", "AAO"):
            ano_col = h
            break
    if not ano_col:
        log("ERROR: no encuentro columna ANO en headers")
        sys.exit(1)

    for c in ["ROL", "estado", "log_decision"]:
        if c not in col_idx:
            log(f"ERROR: falta columna '{c}'")
            sys.exit(1)

    # 4) Crear columnas nuevas si faltan
    log("\n[3] Verificando columnas de auditoria...")
    for col in COLS_NUEVAS:
        if col not in col_idx:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col)
            col_idx[col] = new_col
            log(f"  Columna agregada: '{col}' (col {new_col})")
        else:
            log(f"  Columna ya existia: '{col}'")

    # 5) Mapear filas por ROL-ANO
    log("\n[4] Indexando filas por ROL-ANO...")
    mapa = {}
    for r in range(2, ws.max_row + 1):
        rol = ws.cell(row=r, column=col_idx["ROL"]).value
        ano = ws.cell(row=r, column=col_idx[ano_col]).value
        clave = construir_clave(rol, ano)
        if clave:
            mapa[clave] = r
    log(f"  {len(mapa)} filas indexadas")

    # 6) Aplicar decisiones
    log("\n[5] Aplicando decisiones del Audit4:")
    log("-" * 70)
    aplicadas_si = 0
    aplicadas_no = 0
    saltadas = 0
    no_encontradas = []

    for rol, decision, obs in DECISIONES:
        fila = mapa.get(rol)
        if not fila:
            log(f"  [SKIP] {rol}: NO ENCONTRADA en _datos_internos")
            no_encontradas.append(rol)
            continue

        log_actual = ws.cell(row=fila, column=col_idx["log_decision"]).value or ""

        # Idempotencia: si ya tiene la entrada Audit4 de esta fecha, saltar
        if PREFIJO_LOG in str(log_actual):
            log(f"  [SKIP] {rol}: ya procesada en {PREFIJO_LOG}")
            saltadas += 1
            continue

        estado_antes = ws.cell(row=fila, column=col_idx["estado"]).value

        if decision == "SI":
            nuevo_estado = estado_antes  # mantener
            nueva_linea_log = f"{PREFIJO_LOG}: SI - {obs}"
        else:  # NO
            nuevo_estado = "ELIMINAR"
            if obs:
                nueva_linea_log = f"{PREFIJO_LOG}: NO - {obs}"
            else:
                nueva_linea_log = f"{PREFIJO_LOG}: NO"

        nuevo_log = f"{nueva_linea_log}\n{log_actual}".strip()

        ws.cell(row=fila, column=col_idx["estado"]).value = nuevo_estado
        ws.cell(row=fila, column=col_idx["log_decision"]).value = nuevo_log

        # Solo SI pobla observacion_abogado + fecha_revision_abogado (per prompt)
        if decision == "SI":
            ws.cell(row=fila, column=col_idx["observacion_abogado"]).value = obs
            ws.cell(row=fila, column=col_idx["fecha_revision_abogado"]).value = FECHA_AUDIT
            aplicadas_si += 1
        else:
            aplicadas_no += 1

        log(f"  [{decision}] {rol:<14} fila {fila:<4} {estado_antes} -> {nuevo_estado}")

    # 7) Resumen
    log("-" * 70)
    log("\n[6] Resumen:")
    log(f"  SI (mantener + observacion):  {aplicadas_si}")
    log(f"  NO (ELIMINAR):                {aplicadas_no}")
    log(f"  Saltadas (ya Audit4):         {saltadas}")
    log(f"  No encontradas en Excel:      {len(no_encontradas)}")
    if no_encontradas:
        log("  -> ROLs no encontrados:")
        for rol in no_encontradas:
            log(f"       {rol}")

    # 8) Guardar (o no)
    if dry_run:
        log("\n>> DRY-RUN: cambios NO guardados. Re-correr con --apply para escribir.")
    else:
        log("\n[7] Guardando Excel...")
        guardar_excel_con_retry(wb, EXCEL_PATH)
        log(f"  Guardado: {EXCEL_PATH}")

    log("\n" + "=" * 70)
    log("LISTO.")
    log("=" * 70)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Aplica decisiones del Audit4 al Excel madre (dry-run por defecto).")
    ap.add_argument("--apply", action="store_true",
                    help="Escribe los cambios al Excel (sin esto corre en dry-run).")
    args = ap.parse_args()
    main(dry_run=not args.apply)
