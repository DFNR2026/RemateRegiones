"""
limpiar_run_2026_04_16.py — One-shot: limpiar estado residual del run 2026-04-16

Contexto:
  El regex de fecha del DOCX no capturaba "27 DE ABRIL" -> 59 causas procesadas
  con fecha_remate vacia. Este script revierte ese run para re-procesarlo con el
  regex arreglado.

Acciones:
  1. causas_ojv.xlsx    -> hoja CAUSAS: borra filas con FECHA_PROCESADO = 2026-04-16
  2. Causas_posible_saldo.xlsx -> hoja _datos_internos: borra filas con
     origen_reporte = Reporte_2026-04-16.xlsx
  3. Reportes/Reporte_2026-04-16*.xlsx -> borra el/los reportes

Cada Excel se respalda antes de modificarse.

Uso: python limpiar_run_2026_04_16.py
"""

import os
import glob
import shutil
import datetime as _dt
from openpyxl import load_workbook

from config import BASE_DIR, CAUSAS_LIQ_DIR, EXCEL_MADRE

# Archivos y parametros
CAUSAS_OJV_PATH   = os.path.join(BASE_DIR, "causas_ojv.xlsx")
REPORTES_DIR      = os.path.join(BASE_DIR, "Reportes")

TARGET_FECHA_STR  = "2026-04-16"  # string a matchear en FECHA_PROCESADO
TARGET_DATE       = _dt.date(2026, 4, 16)
TARGET_ORIGEN     = "Reporte_2026-04-16.xlsx"
REPORTE_GLOB      = os.path.join(REPORTES_DIR, "Reporte_2026-04-16*.xlsx")


def _fecha_matches(val, fecha_str: str, fecha_obj: _dt.date) -> bool:
    """Comparar celda vs fecha objetivo en varios formatos."""
    if val is None:
        return False
    if isinstance(val, _dt.datetime):
        return val.date() == fecha_obj
    if isinstance(val, _dt.date):
        return val == fecha_obj
    s = str(val).strip()
    if not s:
        return False
    if s == fecha_str:
        return True
    # Formatos comunes tipo "2026-04-16 00:00:00"
    if s.startswith(fecha_str):
        return True
    return False


def _backup(path: str) -> str:
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    backup_path = f"{base}_backup_{ts}{ext}"
    shutil.copy2(path, backup_path)
    return backup_path


def _headers(ws):
    h = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            h[str(cell.value).strip()] = col_idx
    return h


# ---------------------------------------------------------------------------
# Paso 1: Contar filas a borrar en causas_ojv.xlsx
# ---------------------------------------------------------------------------
def contar_ojv():
    if not os.path.exists(CAUSAS_OJV_PATH):
        print(f"[WARN] No existe: {CAUSAS_OJV_PATH}")
        return None, []

    wb = load_workbook(CAUSAS_OJV_PATH)
    if "CAUSAS" not in wb.sheetnames:
        print(f"[WARN] Hoja CAUSAS no existe en {CAUSAS_OJV_PATH}")
        wb.close()
        return wb, []

    ws = wb["CAUSAS"]
    headers = _headers(ws)
    col_fecha = headers.get("FECHA_PROCESADO")
    if not col_fecha:
        print(f"[ERROR] Columna FECHA_PROCESADO no encontrada. Headers: {list(headers.keys())}")
        wb.close()
        return None, []

    filas_a_borrar = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=col_fecha).value
        if _fecha_matches(val, TARGET_FECHA_STR, TARGET_DATE):
            filas_a_borrar.append(row_idx)

    return wb, filas_a_borrar


# ---------------------------------------------------------------------------
# Paso 2: Contar filas a borrar en Causas_posible_saldo.xlsx
# ---------------------------------------------------------------------------
def contar_madre():
    if not os.path.exists(EXCEL_MADRE):
        print(f"[WARN] No existe: {EXCEL_MADRE}")
        return None, []

    wb = load_workbook(EXCEL_MADRE)
    if "_datos_internos" not in wb.sheetnames:
        print(f"[WARN] Hoja _datos_internos no existe en {EXCEL_MADRE}")
        wb.close()
        return wb, []

    ws = wb["_datos_internos"]
    headers = _headers(ws)
    col_origen = headers.get("origen_reporte")
    if not col_origen:
        print(f"[ERROR] Columna origen_reporte no encontrada. Headers: {list(headers.keys())}")
        wb.close()
        return None, []

    filas_a_borrar = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=col_origen).value
        if val is not None and str(val).strip() == TARGET_ORIGEN:
            filas_a_borrar.append(row_idx)

    return wb, filas_a_borrar


# ---------------------------------------------------------------------------
# Paso 3: Reportes a borrar
# ---------------------------------------------------------------------------
def contar_reportes():
    return sorted(glob.glob(REPORTE_GLOB))


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print("=" * 70)
    print("Limpieza del run 2026-04-16 (pre re-procesamiento con regex arreglado)")
    print("=" * 70)

    wb_ojv, filas_ojv = contar_ojv()
    wb_madre, filas_madre = contar_madre()
    reportes = contar_reportes()

    print(f"\n[1] {CAUSAS_OJV_PATH}")
    print(f"     filas a borrar en hoja CAUSAS: {len(filas_ojv)}  (esperadas ~59)")

    print(f"\n[2] {EXCEL_MADRE}")
    print(f"     filas a borrar en hoja _datos_internos: {len(filas_madre)}  (esperadas ~55)")

    print(f"\n[3] Reportes a eliminar: {len(reportes)}")
    for r in reportes:
        print(f"     - {r}")

    total = len(filas_ojv) + len(filas_madre) + len(reportes)
    if total == 0:
        print("\nNada que limpiar. Saliendo.")
        if wb_ojv: wb_ojv.close()
        if wb_madre: wb_madre.close()
        return

    resp = input("\n¿Continuar con la limpieza? (s/n): ").strip().lower()
    if resp != "s":
        print("Cancelado por el usuario. Nada modificado.")
        if wb_ojv: wb_ojv.close()
        if wb_madre: wb_madre.close()
        return

    # --- Limpieza causas_ojv ---
    if wb_ojv is not None and filas_ojv:
        bkp = _backup(CAUSAS_OJV_PATH)
        print(f"\n[1] Backup: {bkp}")
        ws = wb_ojv["CAUSAS"]
        # Borrar de abajo hacia arriba
        for row_idx in sorted(filas_ojv, reverse=True):
            ws.delete_rows(row_idx, 1)
        wb_ojv.save(CAUSAS_OJV_PATH)
        wb_ojv.close()
        print(f"[1] OK: {len(filas_ojv)} filas borradas en CAUSAS")
    else:
        if wb_ojv: wb_ojv.close()

    # --- Limpieza Excel madre ---
    if wb_madre is not None and filas_madre:
        bkp = _backup(EXCEL_MADRE)
        print(f"\n[2] Backup: {bkp}")
        ws = wb_madre["_datos_internos"]
        for row_idx in sorted(filas_madre, reverse=True):
            ws.delete_rows(row_idx, 1)
        wb_madre.save(EXCEL_MADRE)
        wb_madre.close()
        print(f"[2] OK: {len(filas_madre)} filas borradas en _datos_internos")
    else:
        if wb_madre: wb_madre.close()

    # --- Borrado reportes ---
    if reportes:
        print()
        for r in reportes:
            try:
                os.remove(r)
                print(f"[3] Borrado: {r}")
            except OSError as e:
                print(f"[3] ERROR borrando {r}: {e}")

    print("\nLimpieza completada.")


if __name__ == "__main__":
    main()
