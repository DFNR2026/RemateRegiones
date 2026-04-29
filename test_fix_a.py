"""
test_fix_a.py
=============
Verifica automaticamente que el Fix A funciona: las columnas custom
del Excel madre deben sobrevivir cuando el filtrador regenera _datos_internos.

QUE HACE:
1. Backup del Excel madre
2. Agrega una columna 'test_persistencia_columna' al Excel con un valor en fila 2
3. Corre `python filtrador_saldos.py --primera-run --solo-merge`
   (este modo NO toca OJV, solo regenera el Excel desde el df cargado)
4. Verifica si la columna y su valor sobrevivieron
5. Limpia: borra la columna de test del Excel para dejarlo igual que antes

TIEMPO: ~30 segundos.
NO REQUIERE conexion a OJV.
NO MODIFICA datos reales (solo agrega/quita una columna sintetica).

USO:
    cd D:\\Remates
    python test_fix_a.py
"""

import shutil
import subprocess
import sys
import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

EXCEL = Path(r"D:\Remates\Causas con liq\Causas_posible_saldo.xlsx")
TEST_COL = 'test_persistencia_columna'
TEST_VAL = 'VALOR_TEST_42_FIX_A'
TEST_ROW = 2  # primera fila de datos


def log(msg):
    safe = msg.encode("ascii", errors="replace").decode("ascii")
    print(safe)


def main():
    log("=" * 64)
    log("  TEST FIX A: persistencia de columnas custom")
    log("=" * 64)

    if not EXCEL.exists():
        log(f"\nERROR: no existe {EXCEL}")
        sys.exit(1)

    # ----- 1) Backup -----
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = EXCEL.with_name(f"{EXCEL.stem}_pretest_{ts}{EXCEL.suffix}")
    shutil.copy(EXCEL, backup)
    log(f"\n[1] Backup creado: {backup.name}")

    # ----- 2) Agregar columna test -----
    log(f"\n[2] Agregando columna '{TEST_COL}' al Excel...")
    try:
        wb = load_workbook(EXCEL)
    except PermissionError:
        log("ERROR: Excel abierto. Cierralo y reintenta.")
        sys.exit(1)

    ws = wb['_datos_internos']
    headers_pre = [c.value for c in ws[1]]
    n_cols_pre = len([h for h in headers_pre if h])
    log(f"    Columnas antes del test: {n_cols_pre}")
    log(f"    Headers: {[h for h in headers_pre if h]}")

    if TEST_COL in headers_pre:
        log(f"    WARN: '{TEST_COL}' ya existia (test anterior). Sobreescribiendo.")
        col_idx = headers_pre.index(TEST_COL) + 1
    else:
        col_idx = ws.max_column + 1
        ws.cell(row=1, column=col_idx, value=TEST_COL)

    ws.cell(row=TEST_ROW, column=col_idx, value=TEST_VAL)
    wb.save(EXCEL)
    wb.close()
    log(f"    OK. Columna '{TEST_COL}' = '{TEST_VAL}' en fila {TEST_ROW}")

    # ----- 3) Forzar regeneracion via --solo-merge -----
    log(f"\n[3] Ejecutando: python filtrador_saldos.py --primera-run --solo-merge")
    log(f"    (esto regenera _datos_internos sin tocar OJV)")
    log(f"    ...")

    try:
        result = subprocess.run(
            [sys.executable, 'filtrador_saldos.py', '--primera-run', '--solo-merge'],
            cwd=r'D:\Remates',
            capture_output=True,
            text=True,
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        log("    ERROR: timeout >120s. El --solo-merge no deberia tardar tanto.")
        log("    Restaurando backup...")
        shutil.copy(backup, EXCEL)
        sys.exit(1)

    log(f"    Exit code: {result.returncode}")
    if result.returncode != 0:
        log(f"    stderr: {result.stderr[:500]}")
        log("    Restaurando backup...")
        shutil.copy(backup, EXCEL)
        sys.exit(1)

    # Mostrar ultimas lineas del stdout para confirmar lo que paso
    if result.stdout:
        last_lines = result.stdout.strip().split('\n')[-8:]
        log(f"    Ultimas lineas del filtrador:")
        for ln in last_lines:
            log(f"      {ln[:120]}")

    # ----- 4) Verificar -----
    log(f"\n[4] Verificando que la columna sobrevivio...")
    wb2 = load_workbook(EXCEL)
    ws2 = wb2['_datos_internos']
    headers_post = [c.value for c in ws2[1]]
    n_cols_post = len([h for h in headers_post if h])
    log(f"    Columnas despues del run: {n_cols_post}")

    test_pasado = False
    if TEST_COL in headers_post:
        col_idx_post = headers_post.index(TEST_COL) + 1
        val_post = ws2.cell(row=TEST_ROW, column=col_idx_post).value
        if val_post == TEST_VAL:
            test_pasado = True
            log(f"    OK: '{TEST_COL}' presente. Valor: '{val_post}'")
        else:
            log(f"    PARCIAL: columna '{TEST_COL}' presente pero valor diferente.")
            log(f"      esperado: '{TEST_VAL}'")
            log(f"      actual:   '{val_post}'")
    else:
        log(f"    FALLO: columna '{TEST_COL}' NO esta en el Excel post-run")
        log(f"    Headers post-run: {[h for h in headers_post if h]}")

    wb2.close()

    # ----- 5) Limpiar columna de test -----
    log(f"\n[5] Limpiando: removiendo columna de test...")
    wb3 = load_workbook(EXCEL)
    ws3 = wb3['_datos_internos']
    headers_clean = [c.value for c in ws3[1]]
    if TEST_COL in headers_clean:
        col_to_delete = headers_clean.index(TEST_COL) + 1
        ws3.delete_cols(col_to_delete, 1)
        wb3.save(EXCEL)
        log(f"    Columna '{TEST_COL}' removida del Excel.")
    else:
        log(f"    (no hay nada que limpiar; columna ya no estaba)")
    wb3.close()

    # ----- 6) Resultado final -----
    log("\n" + "=" * 64)
    if test_pasado:
        log("  RESULTADO: TEST PASADO  - Fix A funciona correctamente.")
        log("  Las columnas custom (observacion_abogado, etc.) ahora persisten.")
        log("  Backup conservado en: " + backup.name)
        log("  (puedes borrarlo cuando quieras)")
    else:
        log("  RESULTADO: TEST FALLIDO - Fix A no esta funcionando.")
        log("  Revisar el cambio aplicado por CC.")
        log("  Backup en: " + backup.name)
    log("=" * 64)
    sys.exit(0 if test_pasado else 1)


if __name__ == "__main__":
    main()
