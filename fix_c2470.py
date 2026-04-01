"""
Fix puntual: C-2470-2025 tiene delta negativo -> ELIMINAR.
monto_acta_remate ($51,651,141) < MONTO_DEUDA_CLP ($68,421,146).

Uso: python fix_c2470.py
"""

import openpyxl
import time
from config import EXCEL_MADRE

ROL_TARGET = "2470"
ANIO_TARGET = "2025"

print(f"Leyendo {EXCEL_MADRE}...")
wb = openpyxl.load_workbook(EXCEL_MADRE)

ws = wb["_datos_internos"]
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

col_rol = headers.index("ROL") + 1
col_anio = headers.index("AÑO") + 1
col_estado = headers.index("estado") + 1
col_log = headers.index("log_decision") + 1
col_deuda = headers.index("MONTO_DEUDA_CLP") + 1
col_monto_acta = headers.index("monto_acta_remate") + 1

encontrada = False
for row in range(2, ws.max_row + 1):
    rol = str(ws.cell(row=row, column=col_rol).value or "").strip()
    anio = str(ws.cell(row=row, column=col_anio).value or "").strip()
    if rol == ROL_TARGET and anio == ANIO_TARGET:
        estado_actual = ws.cell(row=row, column=col_estado).value
        deuda = ws.cell(row=row, column=col_deuda).value
        monto_acta = ws.cell(row=row, column=col_monto_acta).value
        print(f"  Encontrada fila {row}: estado={estado_actual}, deuda={deuda}, monto_acta={monto_acta}")

        # Verificar delta negativo
        try:
            d = int(float(str(deuda))) if deuda else 0
            a = int(float(str(monto_acta))) if monto_acta else 0
        except (ValueError, TypeError):
            d, a = 0, 0

        if a > 0 and d > 0 and a < d:
            ws.cell(row=row, column=col_estado).value = "ELIMINAR"
            log_anterior = ws.cell(row=row, column=col_log).value or ""
            ws.cell(row=row, column=col_log).value = (
                f"FIX: Delta negativo (${a:,} < ${d:,}). "
                f"Monto acta menor que deuda -> ELIMINAR. "
                f"[Anterior: {log_anterior}]"
            )
            print(f"  FIX: C-{ROL_TARGET}-{ANIO_TARGET} cambiado a ELIMINAR (delta negativo: ${a:,} < ${d:,})")
            encontrada = True
        else:
            print(f"  SKIP: delta no es negativo (acta=${a:,}, deuda=${d:,})")
        break

if not encontrada:
    print(f"  No se encontro C-{ROL_TARGET}-{ANIO_TARGET} con delta negativo")
else:
    # Guardar con retry
    for intento in range(3):
        try:
            wb.save(EXCEL_MADRE)
            print(f"Excel guardado: {EXCEL_MADRE}")
            break
        except PermissionError:
            if intento < 2:
                print(f"  PermissionError. Cierra el Excel y presiona Enter...")
                input()
            else:
                backup = EXCEL_MADRE.replace(".xlsx", "_backup.xlsx")
                wb.save(backup)
                print(f"  No se pudo guardar. Backup en: {backup}")

wb.close()
