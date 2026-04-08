"""
fix_auditoria_abogados.py -- Eliminar 32 causas segun auditoria de abogados

19 de Revision Manual + 13 de Pendientes Liquidacion.
Se mantienen: C-4837-2024, C-3276-2025.

Uso: python fix_auditoria_abogados.py
"""

import time
import openpyxl
from config import EXCEL_MADRE

# === SECCION 1: Revision Manual (19 causas) ===
eliminar_revision = [
    (1381, 2025, "Abogado: Demandado pago deuda, excedente de su propio dinero consignado"),
    (2543, 2025, "Abogado: Liquidacion con varios creditos de otros tribunales"),
    (1716, 2024, "Abogado: Error tribunal, 2 cuadernos apremio, sistema reviso el inactivo"),
    (392, 2024, "Abogado: Incidente del demandado folio 34 principal suspendio remate"),
    (1225, 2023, "Abogado: Juicio Ley de Arrendamiento, cuadernos distintos a ejecutivos"),
    (1809, 2025, "Abogado: No hubo remate, nueva fecha 6 mayo 2026"),
    (2967, 2024, "Abogado: Pagare + credito hipotecario, deuda total muy alta"),
    (1021, 2022, "Abogado: Causa archivada en OJV"),
    (2821, 2025, "Abogado: Causa paralizada por avenimiento/transaccion folio 18 apremio"),
    (2264, 2024, "Abogado: Adjudicado al ejecutante por $276M (art 499 CPC, similar a CCC)"),
    (2500, 2022, "Abogado: Deuda $723M en 2025, no hay postores interesados"),
    (2017, 2025, "Abogado: Liquidacion concursal Ley 20.720, eliminar todas estas"),
    (3746, 2024, "Abogado: Silencio del demandante indica posible avenimiento con demandado"),
    (1400, 2025, "Abogado: No hay postores (folio 30 apremio, certificado 31/03/2026)"),
    (463, 2025, "Abogado: Tribunal no acepta rebaja del minimo de remate"),
    (4028, 2025, "Abogado: ROL no asocia con ninguna caratula en OJV"),
    (3319, 2024, "Abogado: ROL no asocia con ninguna caratula en OJV"),
    (3338, 2024, "Abogado: CCC confirmado en acta folio 66 apremio"),
    (752, 2021, "Abogado: ROL no asocia con ninguna causa en tribunal"),
]

# === SECCION 2: Pendientes Liquidacion (13 causas) ===
eliminar_pendientes = [
    (1563, 2025, "Abogado: Toma la causa manualmente, retirar del sistema"),
    (2881, 2021, "Abogado: Demandado con abogado, recurrio a Corte de Apelaciones"),
    (857, 2025, "Abogado: CCC confirmado en acta folio 29 cuaderno apremio"),
    (862, 2021, "Abogado: Acta de remate como 'mero tramite' en folio 92 apremio"),
    (1467, 2025, "Abogado: Esperar liquidacion pero eliminar del seguimiento"),
    (8438, 2024, "Abogado: Demandado pago deuda (Da cuenta de pago folio 33 apremio)"),
    (2168, 2025, "Abogado: CCC en acta folio 29 apremio como 'mero tramite'"),
    (755, 2024, "Abogado: CCC en folio 31 apremio 'Se da por adjudicado al ejecutante'"),
    (817, 2023, "Abogado: Remate por $55M vs deuda $150M, sin excedente (Actuacion folio 58)"),
    (616, 2025, "Abogado: Liquidacion folio 42 muestra deuda persiste post-remate, sin excedente"),
    (1472, 2024, "Abogado: Demandado pago deuda capital (Da cuenta de pago folio 48), no hubo remate"),
    (749, 2025, "Abogado: Liquidacion muestra delta de solo $3M tras costas"),
    (3504, 2022, "Abogado: Acta en Actuacion folio 36 apremio, remate $60M vs deuda $60M, sin excedente"),
]

TODAS = eliminar_revision + eliminar_pendientes


def main():
    print(f"Leyendo Excel madre: {EXCEL_MADRE}")
    wb = openpyxl.load_workbook(EXCEL_MADRE)
    ws = wb["_datos_internos"]

    # Encontrar indices de columnas por header (fila 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx

    col_rol = headers.get("ROL") or headers.get("Rol")
    col_anio = headers.get("AÑO") or headers.get("Año")
    col_estado = headers.get("estado") or headers.get("Estado")
    col_log = headers.get("log_decision") or headers.get("Decision")
    col_saldo = headers.get("monto_liquidacion_saldo")

    if not all([col_rol, col_anio, col_estado, col_log]):
        print("ERROR: No se encontraron columnas requeridas")
        print(f"  Headers encontrados: {list(headers.keys())}")
        return

    cnt_eliminadas = 0
    cnt_skip = 0

    for rol, anio, razon in TODAS:
        found = False
        for row_idx in range(2, ws.max_row + 1):
            cell_rol = ws.cell(row=row_idx, column=col_rol).value
            cell_anio = ws.cell(row=row_idx, column=col_anio).value

            try:
                r = int(float(str(cell_rol)))
            except (ValueError, TypeError):
                continue
            try:
                a = int(float(str(cell_anio)))
            except (ValueError, TypeError):
                continue

            if r == rol and a == anio:
                found = True
                estado_actual = str(ws.cell(row=row_idx, column=col_estado).value or "").strip()

                if estado_actual == "ELIMINAR":
                    print(f"SKIP: C-{rol}-{anio} ya es ELIMINAR")
                    cnt_skip += 1
                else:
                    ws.cell(row=row_idx, column=col_estado).value = "ELIMINAR"
                    ws.cell(row=row_idx, column=col_log).value = (
                        f"ELIMINAR: {razon}. [Anterior: {estado_actual}]"
                    )

                    # Limpiar monto_liquidacion_saldo si tenia valor
                    if col_saldo:
                        val_saldo = ws.cell(row=row_idx, column=col_saldo).value
                        if val_saldo and str(val_saldo).strip():
                            print(f"  Limpiando monto_liquidacion_saldo: {val_saldo}")
                            ws.cell(row=row_idx, column=col_saldo).value = ""

                    print(f"ELIMINAR: C-{rol}-{anio} ({estado_actual} -> ELIMINAR): {razon}")
                    cnt_eliminadas += 1
                break

        if not found:
            print(f"WARNING: C-{rol}-{anio} no encontrada en Excel")

    # Guardar con retry
    for intento in range(1, 4):
        try:
            wb.save(EXCEL_MADRE)
            print(f"\nExcel guardado: {EXCEL_MADRE}")
            break
        except PermissionError:
            if intento < 3:
                print(f"PermissionError (intento {intento}/3). Cierra el Excel y espera...")
                time.sleep(3)
            else:
                print("ERROR: No se pudo guardar. Cierra el Excel e intenta de nuevo.")

    wb.close()

    print(f"\nResumen:")
    print(f"  Total eliminadas: {cnt_eliminadas}")
    print(f"  Total ya eliminadas: {cnt_skip}")
    print(f"  Causas que se mantienen: C-4837-2024, C-3276-2025")


if __name__ == "__main__":
    main()
