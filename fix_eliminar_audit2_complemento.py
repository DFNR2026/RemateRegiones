"""
fix_eliminar_audit2_complemento.py — Complemento Auditoria 2 (3 causas faltantes)

Abogado marco NO en estas 3 causas del Audit2 (2026-04-11) pero no se
eliminaron en el script anterior fix_eliminar_audit2.py.

Las 2 excedentes (C-1529, C-2540) ya fueron borradas manualmente del
Causas_con_liquidacion.xlsx por Diego. Este script actualiza el Excel madre.

Uso: python fix_eliminar_audit2_complemento.py
"""

import time
import openpyxl
from config import EXCEL_MADRE

# Estado objetivo. El codebase usa "ELIMINAR" (consistente con fix_eliminar_audit2.py
# y _ESTADOS_PROCESAR de filtrador_saldos.py).
_ESTADO_OBJETIVO = "ELIMINAR"

# 3 causas del complemento Audit2:
ELIMINAR_AUDIT2_COMPLEMENTO = [
    (1529, 2023, "Audit2: Abogado marco NO. Observacion: 'OK, hay dos abogados, ubicar al demandado'. Posibles tercerias o conflictos multiples."),
    (2540, 2025, "Audit2: Abogado marco NO. Observacion: 'Ya hable con ella, es la matrona castigada'. Ya contactada y descartada por abogado."),
    (53,   2022, "Audit2: Abogado marco NO. Observacion: 'OK, ubicare a los demandados'. Delta $123M pero abogado decidio no seguir."),
]


def main():
    print(f"Leyendo Excel madre: {EXCEL_MADRE}")
    wb = openpyxl.load_workbook(EXCEL_MADRE)
    ws = wb["_datos_internos"]

    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx

    col_rol = headers.get("ROL") or headers.get("Rol")
    col_anio = headers.get("AÑO") or headers.get("Año")
    col_estado = headers.get("estado") or headers.get("Estado")
    col_log = headers.get("log_decision") or headers.get("Decision")
    col_saldo = headers.get("monto_liquidacion_saldo")

    if not all([col_rol, col_anio, col_estado, col_log]):
        print("ERROR: No se encontraron columnas requeridas")
        print(f"  Headers encontrados: {list(headers.keys())}")
        return

    cnt_eliminadas = 0
    cnt_skip = 0
    cnt_not_found = 0

    for rol, anio, razon in ELIMINAR_AUDIT2_COMPLEMENTO:
        found = False
        for row_idx in range(2, ws.max_row + 1):
            cell_rol = ws.cell(row=row_idx, column=col_rol).value
            cell_anio = ws.cell(row=row_idx, column=col_anio).value

            try:
                r = int(float(str(cell_rol)))
            except (ValueError, TypeError):
                continue
            try:
                a = int(float(str(cell_anio)))
            except (ValueError, TypeError):
                continue

            if r == rol and a == anio:
                found = True
                estado_actual = str(ws.cell(row=row_idx, column=col_estado).value or "").strip()

                if estado_actual == _ESTADO_OBJETIVO:
                    print(f"SKIP: C-{rol}-{anio} ya es {_ESTADO_OBJETIVO}")
                    cnt_skip += 1
                else:
                    ws.cell(row=row_idx, column=col_estado).value = _ESTADO_OBJETIVO
                    ws.cell(row=row_idx, column=col_log).value = (
                        f"{_ESTADO_OBJETIVO}: {razon} [Anterior: {estado_actual}]"
                    )

                    if col_saldo:
                        val_saldo = ws.cell(row=row_idx, column=col_saldo).value
                        if val_saldo and str(val_saldo).strip():
                            print(f"  Limpiando monto_liquidacion_saldo: {val_saldo}")
                            ws.cell(row=row_idx, column=col_saldo).value = ""

                    print(f"{_ESTADO_OBJETIVO}: C-{rol}-{anio} ({estado_actual} -> {_ESTADO_OBJETIVO}): {razon}")
                    cnt_eliminadas += 1
                break

        if not found:
            print(f"WARNING: C-{rol}-{anio} no encontrada en Excel")
            cnt_not_found += 1

    # Guardar con retry (3 intentos)
    for intento in range(1, 4):
        try:
            wb.save(EXCEL_MADRE)
            print(f"\nExcel guardado: {EXCEL_MADRE}")
            break
        except PermissionError:
            if intento < 3:
                print(f"PermissionError (intento {intento}/3). Cierra el Excel y espera...")
                time.sleep(3)
            else:
                print("ERROR: No se pudo guardar. Cierra el Excel e intenta de nuevo.")

    wb.close()

    total_esperadas = len(ELIMINAR_AUDIT2_COMPLEMENTO)
    print(f"\nResumen: {cnt_eliminadas} eliminadas, {cnt_skip} ya estaban, {cnt_not_found} no encontradas "
          f"(esperadas: {total_esperadas})")


if __name__ == "__main__":
    main()
