"""Persistencia de Excel - extraído de filtrador_saldos.py (Frente B)."""
import os
import re
import shutil
import time
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _guardar_excel_con_retry(wb, filepath, max_intentos=3):
    """Guarda workbook con retry si el archivo esta abierto en Excel."""
    from filtrador_saldos import log
    for intento in range(max_intentos):
        try:
            wb.save(filepath)
            return True
        except PermissionError:
            if intento < max_intentos - 1:
                log.error("ERROR: No se puede guardar %s. Esta abierto en Excel?", filepath)
                log.error("  Cierra el archivo y presiona Enter (%d/%d)...",
                          intento + 1, max_intentos)
                try:
                    input(">>> Presiona Enter para reintentar: ")
                except EOFError:
                    time.sleep(5)
            else:
                log.error("ERROR CRITICO: No se pudo guardar despues de %d intentos.", max_intentos)
                backup = filepath.replace(
                    ".xlsx",
                    f"_backup_{datetime.now().strftime('%H%M%S')}.xlsx",
                )
                try:
                    wb.save(backup)
                    log.info("  Backup guardado en: %s", backup)
                except Exception:
                    log.error("  ERROR: Tampoco se pudo guardar backup.")
                return False
    return False


def _guardar_excel_formateado(df):
    """Guarda el Excel madre con formato visual completo (3 pestanas + revision manual)."""
    from filtrador_saldos import (
        log, EXCEL_MADRE, _COLS_MADRE,
        _DISPLAY_COLS_PRINCIPAL, _DISPLAY_COLS_ANTIG,
        _calcular_campos_derivados,
        _escribir_hoja, _escribir_hoja_revision_manual,
    )
    log.info("Generando Excel formateado...")

    df = _calcular_campos_derivados(df.copy())

    # Separar causas por estado para las pestañas visibles
    ESTADOS_ACTIVOS = ["PENDIENTE_FILTRO1", "PENDIENTE_ACTA", "PENDIENTE_LIQUIDACION",
                       "NECESITA_PDF_ACTA", "LIQUIDACION_ENCONTRADA", "EXCEDENTE_CONFIRMADO"]
    mask_revision = df["estado"] == "PENDIENTE_REVISION_MANUAL"
    df_revision = df[mask_revision].copy()
    df_activas = df[df["estado"].isin(ESTADOS_ACTIVOS)].copy()

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Causas con Saldo"

    df_principal = df_activas.copy()
    df_principal["_sort_delta"] = df_principal["_delta"].apply(
        lambda x: float("-inf") if x is None or (isinstance(x, float) and pd.isna(x)) else x
    )
    df_principal = df_principal.sort_values("_sort_delta", ascending=False).drop(columns=["_sort_delta"])

    _escribir_hoja(ws1, df_principal, _DISPLAY_COLS_PRINCIPAL)

    ws2 = wb.create_sheet("Por Antigüedad")

    df_antig = df_activas.copy()
    df_antig["_sort_fecha"] = df_antig["fecha_remate"].apply(
        lambda x: "9999-99-99" if not x or (isinstance(x, float) and pd.isna(x)) or str(x).strip() == "" else str(x)
    )
    df_antig = df_antig.sort_values("_sort_fecha", ascending=True).drop(columns=["_sort_fecha"])

    _escribir_hoja(ws2, df_antig, _DISPLAY_COLS_ANTIG, es_antiguedad=True)

    # --- Cambio 4: Pestana Revision Manual ---
    if len(df_revision) > 0:
        ws_rm = wb.create_sheet("Revisión Manual")
        _escribir_hoja_revision_manual(ws_rm, df_revision, _DISPLAY_COLS_PRINCIPAL)
        log.info("  Pestana 'Revision Manual': %d causas", len(df_revision))

    cols_extra = [c for c in df.columns if c and c not in _COLS_MADRE]
    cols_a_escribir = list(_COLS_MADRE) + cols_extra
    if cols_extra:
        log.info("  Preservando %d columna(s) extra en _datos_internos: %s",
                 len(cols_extra), ", ".join(cols_extra))

    ws_data = wb.create_sheet("_datos_internos")
    for col_idx, col_name in enumerate(cols_a_escribir, 1):
        ws_data.cell(row=1, column=col_idx, value=col_name)
    for row_idx, (_, data_row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(cols_a_escribir, 1):
            val = data_row.get(col_name, "")
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = ""
            ws_data.cell(row=row_idx, column=col_idx, value=val)
    ws_data.sheet_state = "hidden"

    _guardar_excel_con_retry(wb, EXCEL_MADRE)
    log.info("Excel madre guardado: %s", EXCEL_MADRE)

    # Copia con fecha
    ahora = datetime.now()
    meses = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre']
    nombre_fecha = f"{ahora.day}_{meses[ahora.month-1]}_{ahora.year}"
    copia_path = os.path.join(os.path.dirname(EXCEL_MADRE), f"Causas_posible_saldo_{nombre_fecha}.xlsx")
    shutil.copy2(EXCEL_MADRE, copia_path)
    log.info(f"Copia con fecha: {copia_path}")


def _generar_excel_liquidaciones(df):
    """Genera Causas_con_liquidacion.xlsx con formato profesional para abogados."""
    from filtrador_saldos import log, EXCEL_LIQUIDACIONES
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Estilos
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill_blue = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_fill_orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    data_font = Font(name="Calibri", size=11)
    verde_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    gris_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(vertical="center")

    # Columnas: (header_amigable, ancho, campo_df, es_monto)
    cols_excedentes = [
        ("Causa (ROL)", 15, None, False),       # A: combinado
        ("Tribunal", 35, "TRIBUNAL", False),     # B
        ("Demandado", 35, "DEMANDADO", False),   # C
        ("Direccion", 45, "DIRECCIÓN", False),   # D
        ("Comuna", 15, "COMUNA", False),         # E
        ("Deuda Original ($)", 18, "MONTO_DEUDA_CLP", True),  # F
        ("Precio Remate ($)", 18, "monto_acta_remate", True),  # G
        ("Deuda Liquidada ($)", 18, "monto_credito_liquidado", True),  # H
        ("Excedente Confirmado ($)", 22, "monto_liquidacion_saldo", True),  # I
        ("Fecha Remate", 14, "fecha_remate", False),  # J
        ("Estado", 12, "estado", False),          # K
    ]

    def _escribir_hoja_liq(ws, df_data, cols_spec, header_fill):
        # Headers
        for col_idx, (header, ancho, _, _) in enumerate(cols_spec, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = borde
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26
                                  else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)].width = ancho

        # Ajustar anchos con get_column_letter
        from openpyxl.utils import get_column_letter
        for col_idx, (_, ancho, _, _) in enumerate(cols_spec, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        # Data
        for row_idx, (_, row) in enumerate(df_data.iterrows(), 2):
            for col_idx, (_, _, campo, es_monto) in enumerate(cols_spec, 1):
                if campo is None:
                    # Campo combinado ROL-AÑO
                    rol = str(row.get("ROL", "")).strip()
                    anio = str(row.get("AÑO", "")).strip()
                    val = f"C-{rol}-{anio}"
                elif es_monto:
                    raw = row.get(campo, "")
                    try:
                        val = int(float(str(raw))) if raw and str(raw).strip() else None
                    except (ValueError, TypeError):
                        val = None
                else:
                    val = row.get(campo, "")
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        val = ""

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = borde

                if es_monto and val is not None:
                    cell.number_format = '#,##0'

                # Fila alterna gris
                if row_idx % 2 == 0:
                    cell.fill = gris_fill

            # Columna excedente (I) con fondo verde
            excedente_col = next(
                (i for i, (h, _, _, _) in enumerate(cols_spec, 1)
                 if "Excedente" in h), None
            )
            if excedente_col:
                cell_exc = ws.cell(row=row_idx, column=excedente_col)
                cell_exc.fill = verde_fill
                cell_exc.font = Font(name="Calibri", size=11, bold=True)

            ws.row_dimensions[row_idx].height = 20

        # Freeze + autofilter
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    # ------------------------------------------------------------------
    # Leer Excel anterior para detectar causas ya registradas
    # ------------------------------------------------------------------
    from openpyxl.utils import get_column_letter

    _SEPARADOR_PREFIX_EXC = "LIQUIDACIONES ENCONTRADAS EL "
    _SEPARADOR_PREFIX_REV = "REVISION MANUAL - "

    sep_font_blue = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    sep_fill_blue = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sep_border_blue = Border(bottom=Side(style="thick", color="1F4E79"))
    sep_align = Alignment(horizontal="center", vertical="center")

    sep_font_orange = Font(name="Calibri", size=12, bold=True, color="C65911")
    sep_fill_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    sep_border_orange = Border(bottom=Side(style="thick", color="C65911"))

    def _leer_causas_anteriores(sheet_name):
        """lee ROL-AÑO de causas ya en el Excel anterior (ignora filas separadoras)."""
        anteriores = set()
        if not os.path.exists(EXCEL_LIQUIDACIONES):
            return anteriores
        try:
            wb_prev = openpyxl.load_workbook(EXCEL_LIQUIDACIONES, read_only=True)
            if sheet_name not in wb_prev.sheetnames:
                wb_prev.close()
                return anteriores
            ws_prev = wb_prev[sheet_name]
            for row in ws_prev.iter_rows(min_row=1, values_only=True):
                if not row or row[0] is None:
                    continue
                val = str(row[0]).strip()
                # Ignorar filas separadoras y headers
                if val.startswith(_SEPARADOR_PREFIX_EXC) or val.startswith(_SEPARADOR_PREFIX_REV):
                    continue
                if val == "Causa (ROL)":
                    continue
                # Extraer ROL-AÑO del formato "C-ROL-AÑO"
                m = re.match(r"C-(\d+)-(\d+)", val)
                if m:
                    anteriores.add(f"{m.group(1)}-{m.group(2)}")
            wb_prev.close()
        except Exception as e:
            log.warning("[LIQ] Error leyendo Excel anterior: %s", e)
        return anteriores

    def _leer_secciones_anteriores(sheet_name, cols_spec, header_fill_color):
        """Lee secciones (separadores + datos) del Excel anterior."""
        secciones = []
        if not os.path.exists(EXCEL_LIQUIDACIONES):
            return secciones
        try:
            wb_prev = openpyxl.load_workbook(EXCEL_LIQUIDACIONES, read_only=True)
            if sheet_name not in wb_prev.sheetnames:
                wb_prev.close()
                return secciones
            ws_prev = wb_prev[sheet_name]
            n_cols = len(cols_spec)

            seccion_actual = None
            filas_seccion = []

            for row in ws_prev.iter_rows(min_row=1, values_only=True):
                if not row or row[0] is None:
                    continue
                val = str(row[0]).strip()
                # Header row — skip
                if val == "Causa (ROL)":
                    continue
                # Check if it's a separator
                is_sep = (val.startswith(_SEPARADOR_PREFIX_EXC) or
                          val.startswith(_SEPARADOR_PREFIX_REV))
                if is_sep:
                    # Save previous section
                    if seccion_actual is not None:
                        secciones.append((seccion_actual, filas_seccion))
                    seccion_actual = val
                    filas_seccion = []
                elif seccion_actual is not None:
                    # Data row belonging to current section
                    filas_seccion.append(list(row[:n_cols]))

            # Save last section
            if seccion_actual is not None:
                secciones.append((seccion_actual, filas_seccion))

            wb_prev.close()
        except Exception as e:
            log.warning("[LIQ] Error leyendo secciones anteriores: %s", e)
        return secciones

    def _escribir_separador(ws, row_num, texto, n_cols, sep_font, sep_fill, sep_bord):
        """Escribe una fila separadora con merge y formato."""
        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num, end_column=n_cols
        )
        cell = ws.cell(row=row_num, column=1, value=texto)
        cell.font = sep_font
        cell.fill = sep_fill
        cell.alignment = sep_align
        cell.border = sep_bord
        ws.row_dimensions[row_num].height = 30

    def _escribir_hoja_con_secciones(ws, secciones_prev, df_nuevas, cols_spec,
                                     header_fill, sep_font_s, sep_fill_s,
                                     sep_bord_s, sep_prefix, fecha_hoy):
        """Escribe hoja con headers + secciones anteriores + seccion nueva."""
        n_cols = len(cols_spec)

        # Headers
        for col_idx, (header, ancho, _, _) in enumerate(cols_spec, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = borde
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        current_row = 2

        # Escribir secciones anteriores
        for sep_texto, filas in secciones_prev:
            _escribir_separador(ws, current_row, sep_texto, n_cols,
                                sep_font_s, sep_fill_s, sep_bord_s)
            current_row += 1

            for fila_vals in filas:
                for col_idx in range(1, n_cols + 1):
                    val = fila_vals[col_idx - 1] if col_idx - 1 < len(fila_vals) else ""
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = borde
                    # Formato monto
                    _, _, _, es_monto = cols_spec[col_idx - 1]
                    if es_monto and val is not None and isinstance(val, (int, float)):
                        cell.number_format = '#,##0'
                current_row += 1

        # Escribir seccion nueva (si hay nuevas)
        if len(df_nuevas) > 0:
            n_nuevas = len(df_nuevas)
            sep_texto = f"{sep_prefix}{fecha_hoy} ({n_nuevas} nuevas)"
            _escribir_separador(ws, current_row, sep_texto, n_cols,
                                sep_font_s, sep_fill_s, sep_bord_s)
            current_row += 1

            for _, row in df_nuevas.iterrows():
                for col_idx, (_, _, campo, es_monto) in enumerate(cols_spec, 1):
                    if campo is None:
                        rol = str(row.get("ROL", "")).strip()
                        anio = str(row.get("AÑO", "")).strip()
                        val = f"C-{rol}-{anio}"
                    elif es_monto:
                        raw = row.get(campo, "")
                        try:
                            val = int(float(str(raw))) if raw and str(raw).strip() else None
                        except (ValueError, TypeError):
                            val = None
                    else:
                        val = row.get(campo, "")
                        if val is None or (isinstance(val, float) and pd.isna(val)):
                            val = ""

                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = borde

                    if es_monto and val is not None:
                        cell.number_format = '#,##0'

                    # Columna excedente con fondo verde
                    excedente_col = next(
                        (i for i, (h, _, _, _) in enumerate(cols_spec, 1)
                         if "Excedente" in h), None
                    )
                    if excedente_col:
                        cell_exc = ws.cell(row=current_row, column=excedente_col)
                        cell_exc.fill = verde_fill
                        cell_exc.font = Font(name="Calibri", size=11, bold=True)

                    ws.row_dimensions[current_row].height = 20
                    current_row += 1

            # Freeze header
            ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # Detectar nuevas causas
    # ------------------------------------------------------------------
    fecha_hoy = datetime.now().strftime("%d/%m/%Y")

    # Excedentes
    mask_exc = df["estado"] == "EXCEDENTE_CONFIRMADO"
    df_exc = df[mask_exc].copy()

    anteriores_exc = _leer_causas_anteriores("Excedentes Confirmados")
    if len(df_exc) > 0:
        df_exc["_clave"] = df_exc.apply(
            lambda r: f"{str(r['ROL']).strip()}-{str(r['AÑO']).strip()}", axis=1
        )
        df_exc_nuevas = df_exc[~df_exc["_clave"].isin(anteriores_exc)].copy()
        # Ordenar nuevas por excedente desc
        try:
            df_exc_nuevas["_sort"] = df_exc_nuevas["monto_liquidacion_saldo"].apply(
                lambda x: int(float(str(x))) if x and str(x).strip() else 0
            )
            df_exc_nuevas = df_exc_nuevas.sort_values("_sort", ascending=False).drop(columns=["_sort"])
        except Exception:
            pass
        df_exc_nuevas = df_exc_nuevas.drop(columns=["_clave"], errors="ignore")
        df_exc = df_exc.drop(columns=["_clave"], errors="ignore")
    else:
        df_exc_nuevas = pd.DataFrame()

    # Revision Manual
    mask_rev = (
        (df["estado"] == "PENDIENTE_REVISION_MANUAL") &
        (df["log_decision"].str.contains("Tipo B", na=False))
    )
    df_rev = df[mask_rev].copy()

    anteriores_rev = _leer_causas_anteriores("Revision Manual")
    if len(df_rev) > 0:
        df_rev["_clave"] = df_rev.apply(
            lambda r: f"{str(r['ROL']).strip()}-{str(r['AÑO']).strip()}", axis=1
        )
        df_rev_nuevas = df_rev[~df_rev["_clave"].isin(anteriores_rev)].copy()
        df_rev_nuevas = df_rev_nuevas.drop(columns=["_clave"], errors="ignore")
        df_rev = df_rev.drop(columns=["_clave"], errors="ignore")
    else:
        df_rev_nuevas = pd.DataFrame()

    # ------------------------------------------------------------------
    # Leer secciones anteriores del Excel
    # ------------------------------------------------------------------
    secciones_exc_prev = _leer_secciones_anteriores(
        "Excedentes Confirmados", cols_excedentes, "1F4E79"
    )
    cols_revision = [
        ("Causa (ROL)", 15, None, False),
        ("Tribunal", 35, "TRIBUNAL", False),
        ("Demandado", 35, "DEMANDADO", False),
        ("Direccion", 45, "DIRECCIÓN", False),
        ("Comuna", 15, "COMUNA", False),
        ("Deuda Original ($)", 18, "MONTO_DEUDA_CLP", True),
        ("Precio Remate ($)", 18, "monto_acta_remate", True),
        ("Credito Adeudado ($)", 22, "monto_credito_liquidado", True),
        ("Fecha Remate", 14, "fecha_remate", False),
        ("Estado", 12, "estado", False),
    ]
    secciones_rev_prev = _leer_secciones_anteriores(
        "Revision Manual", cols_revision, "ED7D31"
    )

    # ------------------------------------------------------------------
    # Construir nuevo Excel
    # ------------------------------------------------------------------
    wb = openpyxl.Workbook()

    # --- Pestana 1: Excedentes Confirmados ---
    ws1 = wb.active
    ws1.title = "Excedentes Confirmados"

    if len(df_exc) > 0 or secciones_exc_prev:
        _escribir_hoja_con_secciones(
            ws1, secciones_exc_prev, df_exc_nuevas, cols_excedentes,
            header_fill_blue, sep_font_blue, sep_fill_blue, sep_border_blue,
            _SEPARADOR_PREFIX_EXC, fecha_hoy,
        )
        n_nuevas_exc = len(df_exc_nuevas)
        log.info("[LIQ] Pestana 'Excedentes Confirmados': %d nuevas, %d secciones previas",
                 n_nuevas_exc, len(secciones_exc_prev))
    else:
        ws1.cell(row=1, column=1, value="No hay excedentes confirmados")

    # --- Pestana 2: Revision Manual ---
    if len(df_rev) > 0 or secciones_rev_prev:
        ws2 = wb.create_sheet("Revision Manual")
        _escribir_hoja_con_secciones(
            ws2, secciones_rev_prev, df_rev_nuevas, cols_revision,
            header_fill_orange, sep_font_orange, sep_fill_orange, sep_border_orange,
            _SEPARADOR_PREFIX_REV, fecha_hoy,
        )
        n_nuevas_rev = len(df_rev_nuevas)
        log.info("[LIQ] Pestana 'Revision Manual': %d nuevas, %d secciones previas",
                 n_nuevas_rev, len(secciones_rev_prev))

    _guardar_excel_con_retry(wb, EXCEL_LIQUIDACIONES)
    log.info("[LIQ] Excel de liquidaciones guardado: %s", EXCEL_LIQUIDACIONES)