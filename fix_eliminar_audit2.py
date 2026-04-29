"""
fix_eliminar_audit2.py — Procesar resultados Auditoria 2 (2026-04-14)

17 causas a marcar como ELIMINAR segun revision del abogado.
Las otras 5 (C-1529, C-2540, C-53, C-1485, C-4306) NO se tocan.

Uso: python fix_eliminar_audit2.py
"""

import time
import openpyxl
from config import EXCEL_MADRE

# 17 causas determinadas por Auditoria 2 (abogado):
ELIMINAR_AUDIT2 = [
    # SECCION 2: REVISION MANUAL -> ELIMINAR
    (3282, 2025, "Audit2: Avenimiento folio 24 Apremio. Demandada llego a acuerdo de pago con banco."),
    (54,   2024, "Audit2: Juzgado de Pemuco no aparece en Corte de Chillan. Tribunal inubicable."),
    (592,  2023, "Audit2: Cargo al credito. Banco demandante se adjudico con CCC folio 52 Apremio."),
    (992,  2022, "Audit2: Da cuenta de pago. Demandada pago capital + intereses + costas."),
    (1394, 2019, "Audit2: Demandado con 3 abogados. Posible avenimiento extrajudicial -> abandono."),
    (3276, 2025, "Audit2: Liquidacion descargada, saldo $7.2M pero faltan costas. Abogados ya en conocimiento, eliminada de plano."),

    # SECCION 3: PENDIENTE_LIQUIDACION -> ELIMINAR
    (1967, 2023, "Audit2: Liquidacion NEGATIVA. Demandado quedo debiendo despues de la subasta."),
    (1986, 2019, "Audit2: Dos tercerias. No quedara saldo positivo."),
    (5040, 2024, "Audit2: Delta $1.1M muy bajo. Intereses + costas absorberan saldo."),
    (953,  2022, "Audit2: Deuda de 2022. Delta $9.2M insuficiente post intereses 4 anios."),
    (2149, 2024, "Audit2: Deudas adicionales Banco Estado (pagare ~$19M + credito hipotecario)."),
    (4837, 2024, "Audit2: Pagare con intereses altisimos. Monto liquidado sera similar a subasta."),
    (3601, 2020, "Audit2: Avenimiento en cuaderno Principal folio 43 (debio estar en Apremio)."),
    (621,  2024, "Audit2: Suspension de remate solicitada por demandante folio 21 Apremio."),
    (3628, 2024, "Audit2: Renegociacion extrajudicial. Demandada renegocio con banco ejecutante."),
    (706,  2025, "Audit2: Mandamiento ~$20M vs remate $25M. Poco saldo post intereses."),
    (850,  2022, "Audit2: Remate suspendido folio 13 Apremio. Demandado tiene 3 abogados."),
]


def main():
    print(f"Leyendo Excel madre: {EXCEL_MADRE}")
    wb = openpyxl.load_workbook(EXCEL_MADRE)
    ws = wb["_datos_internos"]

    # Encontrar indices de columnas por header (fila 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx

    col_rol = headers.get("ROL") or headers.get("Rol")
    col_anio = headers.get("AÑO") or headers.get("Año")
    col_estado = headers.get("estado") or headers.get("Estado")
    col_log = headers.get("log_decision") or headers.get("Decision")
    col_saldo = headers.get("monto_liquidacion_saldo")

    if not all([col_rol, col_anio, col_estado, col_log]):
        print("ERROR: No se encontraron columnas requeridas")
        print(f"  Headers encontrados: {list(headers.keys())}")
        return

    cnt_eliminadas = 0
    cnt_skip = 0
    cnt_not_found = 0

    for rol, anio, razon in ELIMINAR_AUDIT2:
        found = False
        for row_idx in range(2, ws.max_row + 1):
            cell_rol = ws.cell(row=row_idx, column=col_rol).value
            cell_anio = ws.cell(row=row_idx, column=col_anio).value

            try:
                r = int(float(str(cell_rol)))
            except (ValueError, TypeError):
                continue
            try:
                a = int(float(str(cell_anio)))
            except (ValueError, TypeError):
                continue

            if r == rol and a == anio:
                found = True
                estado_actual = str(ws.cell(row=row_idx, column=col_estado).value or "").strip()

                if estado_actual == "ELIMINAR":
                    print(f"SKIP: C-{rol}-{anio} ya es ELIMINAR")
                    cnt_skip += 1
                else:
                    ws.cell(row=row_idx, column=col_estado).value = "ELIMINAR"
                    ws.cell(row=row_idx, column=col_log).value = (
                        f"ELIMINAR: {razon} [Anterior: {estado_actual}]"
                    )

                    # Limpiar monto_liquidacion_saldo si tenia valor
                    if col_saldo:
                        val_saldo = ws.cell(row=row_idx, column=col_saldo).value
                        if val_saldo and str(val_saldo).strip():
                            print(f"  Limpiando monto_liquidacion_saldo: {val_saldo}")
                            ws.cell(row=row_idx, column=col_saldo).value = ""

                    print(f"ELIMINAR: C-{rol}-{anio} ({estado_actual} -> ELIMINAR): {razon}")
                    cnt_eliminadas += 1
                break

        if not found:
            print(f"WARNING: C-{rol}-{anio} no encontrada en Excel")
            cnt_not_found += 1

    # Guardar con retry
    for intento in range(1, 4):
        try:
            wb.save(EXCEL_MADRE)
            print(f"\nExcel guardado: {EXCEL_MADRE}")
            break
        except PermissionError:
            if intento < 3:
                print(f"PermissionError (intento {intento}/3). Cierra el Excel y espera...")
                time.sleep(3)
            else:
                print("ERROR: No se pudo guardar. Cierra el Excel e intenta de nuevo.")

    wb.close()

    total_esperadas = len(ELIMINAR_AUDIT2)
    print(f"\nResumen: {cnt_eliminadas} eliminadas, {cnt_skip} ya estaban, {cnt_not_found} no encontradas "
          f"(esperadas: {total_esperadas})")


if __name__ == "__main__":
    main()
