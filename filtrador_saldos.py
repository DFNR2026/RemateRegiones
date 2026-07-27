"""
Filtrador de Saldos - Tracking Post-Remate

Sistema 100% independiente del pipeline M1-M5.
Lee reportes Excel generados por M5, consulta la OJV,
y filtra causas buscando excedentes de remates judiciales.

Uso:
    python filtrador_saldos.py                              # ejecucion normal
    python filtrador_saldos.py --primera-run                # modo auditoria (no elimina)
    python filtrador_saldos.py --solo-merge                 # solo importar, sin OJV
    python filtrador_saldos.py --solo-filtro1               # merge + solo Filtro 1
    python filtrador_saldos.py --primera-run --solo-filtro1  # merge + Filtro 1 en auditoria
    python filtrador_saldos.py --reaudit --workers 5        # reprocesar PENDIENTE_ACTA
"""

import os
import sys
import re
import time
import logging
import argparse
import tempfile
import json
import shutil
import subprocess
import math
from datetime import datetime, date, timedelta

import pandas as pd
import openpyxl
import fitz  # PyMuPDF
fitz.TOOLS.mupdf_display_errors(False)  # Suprimir warnings de MuPDF en consola
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

from ojv_remates import (
    navegar_a_consulta,
    buscar_causa,
    abrir_detalle,
    seleccionar_cuaderno,
    seleccionar_por_texto,
    filas_del_modal,
    descargar_pdf_de_fila,
    limpiar_formulario,
    cerrar_popups,
)
from config import (
    REPORTES_DIR,
    CAUSAS_LIQ_DIR,
    EXCEL_MADRE,
    LIQUIDACIONES_DIR,
    LIQUIDACIONES_RAW_DIR,
    CORTES_RM,
    LOGS_LIQUI_DIR,
    CAUSAS_ELIMINAR_MANUAL,
    ACTAS_DIR,
    EXCEL_LIQUIDACIONES,
    DESCARGAS_DIR,
)

# Ruta base del script (usada por config de directorios)
import config as _config
REMATES_DIR = _config.BASE_DIR

from persistencia_excel import (
    _guardar_excel_con_retry,
    _guardar_excel_formateado,
    _generar_excel_liquidaciones,
)

from analisis_pdf import (
    _extraer_texto_pdf,
    _ocr_pdf,
    _analizar_pdf_acta,
)

# ---------------------------------------------------------------------------
# Logging -- dual: consola + archivo en LOGS_LIQUI_DIR
# ---------------------------------------------------------------------------
os.makedirs(LOGS_LIQUI_DIR, exist_ok=True)

# Detectar si somos un worker subprocess (no crear log de archivo propio).
# Nota: el proceso principal recibe "--workers N" (plural); solo los workers reales
# reciben "--worker-mode" (flag store_true, ver argparse mas abajo).
_ES_WORKER = "--worker-mode" in sys.argv

_root = logging.getLogger()
for _h in _root.handlers[:]:
    _root.removeHandler(_h)
_root.setLevel(logging.INFO)

_FMT = logging.Formatter("%(asctime)s [FILTRADOR] %(message)s", datefmt="%H:%M:%S")

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(_FMT)
_root.addHandler(_console_handler)


class _TeeWriter:
    """Redirige print() a consola + archivo de log simultaneamente."""

    def __init__(self, original_stream, log_file_handle):
        self._orig = original_stream
        self._log = log_file_handle

    def write(self, text):
        self._orig.write(text)
        self._orig.flush()
        self._log.write(text)
        self._log.flush()

    def flush(self):
        self._orig.flush()
        self._log.flush()

    def fileno(self):
        return self._orig.fileno()


if not _ES_WORKER:
    _LOG_FILE = os.path.join(
        LOGS_LIQUI_DIR,
        f"filtrador_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log",
    )

    _file_handler = logging.FileHandler(_LOG_FILE, encoding="utf-8")
    _file_handler.setFormatter(_FMT)
    _root.addHandler(_file_handler)

    _log_fh = open(_LOG_FILE, "a", encoding="utf-8")
    sys.stdout = _TeeWriter(sys.__stdout__, _log_fh)
    sys.stderr = _TeeWriter(sys.__stderr__, _log_fh)
else:
    _LOG_FILE = None  # Los workers escriben a stdout, el orquestador lo captura

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Modulo 3 (extractor de monto) - opcional, usado por Paso 0.5
# Si M3 falla al importar, el filtrador sigue corriendo sin Paso 0.5.
# ---------------------------------------------------------------------------
try:
    from modulo3_extractor import (
        _extraer_de_mandamiento,
        _extraer_texto_pdf,
        obtener_uf_hoy,
    )
    _M3_DISPONIBLE = True
except Exception as _e_m3:
    log.warning("M3 no disponible, Paso 0.5 deshabilitado: %s", _e_m3)
    _M3_DISPONIBLE = False

# ---------------------------------------------------------------------------
# Regex patterns (Filtro 2 y 3 -- requieren descarga de PDFs)
# ---------------------------------------------------------------------------
PATRON_ACTA_REMATE = re.compile(
    r"acta\s+de\s+remate"
    r"|acta\s+de\s+subasta",
    re.IGNORECASE,
)

PATRON_MONTO_ACTA = re.compile(
    r"(?:suma\s+de|por\s+la\s+suma\s+de|adjudica.*?en\s+la\s+suma\s+de)"
    r"\s*\$?\s*([\d.,]+)",
    re.IGNORECASE,
)

PATRON_LIQUIDACION = re.compile(
    r"liquidaci[o\u00f3]n|liquidacion",
    re.IGNORECASE,
)

PATRON_SALDO = re.compile(
    r"[Ss]aldo\s+a\s+favor\s+de(?:l)?\s+ejecutad[oa].*?\$\s*([\d.,]+)",
    re.DOTALL,
)

# ---------------------------------------------------------------------------
# Meses en espanol -> numero
# ---------------------------------------------------------------------------
_MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# ---------------------------------------------------------------------------
# Columnas del Excel madre (internas)
# ---------------------------------------------------------------------------
_COLS_MADRE = [
    "ROL", "AÑO", "CORTE", "TRIBUNAL", "DEMANDANTE", "DEMANDADO",
    "DIRECCIÓN", "COMUNA", "MONTO_DEUDA_CLP", "CBR_MOTIVO",
    # Tracking
    "estado", "fecha_remate", "monto_acta_remate", "monto_liquidacion_saldo",
    "monto_credito_liquidado", "ruta_liquidacion",
    "fecha_ultimo_check", "notas", "origen_reporte", "log_decision",
    "detalle_auditoria",
]

# Historial de causas eliminadas fisicamente del Excel (para evitar
# re-importarlas desde reportes M5 en runs futuros).
HISTORIAL_ELIMINADAS = os.path.join(CAUSAS_LIQ_DIR, "causas_eliminadas_historial.csv")

# Mapeo de headers M5 -> columnas madre
_HEADER_MAP = {
    "ROL": "ROL",
    "Año": "AÑO", "AÑO": "AÑO",
    "Corte": "CORTE", "CORTE": "CORTE",
    "Tribunal": "TRIBUNAL", "TRIBUNAL": "TRIBUNAL",
    "Demandante": "DEMANDANTE", "DEMANDANTE": "DEMANDANTE",
    "Demandado": "DEMANDADO", "DEMANDADO": "DEMANDADO",
    "Dirección": "DIRECCIÓN", "DIRECCIÓN": "DIRECCIÓN",
    "Comuna": "COMUNA", "COMUNA": "COMUNA",
    "Deuda (CLP)": "MONTO_DEUDA_CLP", "MONTO_DEUDA_CLP": "MONTO_DEUDA_CLP",
    "Fechas Public.": "FECHA_PUBLICACION", "FECHA_PUBLICACION": "FECHA_PUBLICACION",
    "CBR Motivo": "CBR_MOTIVO", "CBR_MOTIVO": "CBR_MOTIVO",
}

# ---------------------------------------------------------------------------
# Columnas de visualizacion Excel -- pestana principal "Causas con Saldo"
# (internal_col, display_name, width)
# ---------------------------------------------------------------------------
_DISPLAY_COLS_PRINCIPAL = [
    ("ROL",                "Rol",                  10),
    ("AÑO",               "Año",                   8),
    ("CORTE",             "Corte",                 25),
    ("TRIBUNAL",          "Tribunal",              35),
    ("DEMANDANTE",        "Demandante",            30),
    ("DEMANDADO",         "Demandado (Ejecutado)", 30),
    ("DIRECCIÓN",         "Dirección",             40),
    ("COMUNA",            "Comuna",                18),
    ("monto_acta_remate", "Monto Remate (CLP)",    20),
    ("MONTO_DEUDA_CLP",   "Deuda (CLP)",           18),
    ("_delta",            "Delta (CLP)",           18),
    ("ruta_liquidacion",  "Liquidación",           15),
    ("fecha_remate",      "Fecha Remate",          15),
    ("estado",            "Estado",                22),
    ("log_decision",      "Decisión",              50),
    ("detalle_auditoria", "Detalle Auditoría",     60),
    ("fecha_ultimo_check","Último Check",          15),
    ("notas",             "Notas",                 30),
    ("origen_reporte",    "Origen Reporte",        25),
]

# Pestana "Por Antigüedad"
_DISPLAY_COLS_ANTIG = [
    ("_dias_desde_remate", "Días desde Remate",    18),
    ("fecha_remate",       "Fecha Remate",         15),
    ("ruta_liquidacion",   "Liquidación",          15),
    ("_delta",             "Delta (CLP)",          18),
    ("monto_acta_remate",  "Monto Remate (CLP)",   20),
    ("MONTO_DEUDA_CLP",    "Deuda (CLP)",          18),
    ("ROL",                "Rol",                  10),
    ("AÑO",               "Año",                    8),
    ("TRIBUNAL",           "Tribunal",             35),
    ("DEMANDANTE",         "Demandante",           30),
    ("DEMANDADO",          "Demandado (Ejecutado)", 30),
    ("estado",             "Estado",               22),
    ("log_decision",       "Decisión",             50),
    ("detalle_auditoria",  "Detalle Auditoría",    60),
    ("origen_reporte",     "Origen Reporte",       25),
]

# Estados validos para procesamiento OJV
_ESTADOS_PROCESAR = {
    "PENDIENTE_FILTRO1",
    "PENDIENTE_ACTA", "PENDIENTE_LIQUIDACION",
}

# Estados que requieren solo Filtro 1
_ESTADOS_FILTRO1 = {"PENDIENTE_FILTRO1"}


# =========================================================================
# HELPERS
# =========================================================================

def _parsear_fecha_publicacion(texto):
    """Parse fecha desde 'DD MMMM YYYY' o 'dd/mm/yyyy'. Devuelve date o None."""
    if not texto or not isinstance(texto, str):
        return None
    texto = texto.strip()

    # Si hay multiples fechas separadas por coma, tomar la ultima
    partes = texto.split(",")
    for parte in reversed(partes):
        parte = parte.strip()
        tokens = parte.upper().split()
        tokens = [t for t in tokens if t != "DE"]
        if len(tokens) >= 3:
            try:
                dia = int(tokens[0])
                mes = _MESES.get(tokens[1])
                anio = int(tokens[2])
                if mes:
                    return date(anio, mes, dia)
            except (ValueError, KeyError):
                pass

    # Fallback dd/mm/yyyy
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", texto)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


def _parsear_fecha_dd_mm_yyyy(texto):
    """Parse 'dd/mm/yyyy' -> date o None."""
    if not texto:
        return None
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(texto))
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


def _parsear_monto(texto_monto):
    """Parse monto chileno a int. Ej: '125.757.208' -> 125757208."""
    if not texto_monto:
        return 0
    limpio = re.sub(r"[\s$]", "", str(texto_monto))
    if "," in limpio:
        entero = limpio.split(",")[0].replace(".", "")
    else:
        entero = limpio.replace(".", "")
    try:
        return int(entero)
    except ValueError:
        return 0


def _deuda_a_int(valor):
    """Convierte MONTO_DEUDA_CLP (puede ser str, float, int) a int."""
    if not valor or str(valor).strip() in ("", "nan", "None"):
        return 0
    try:
        return int(float(str(valor)))
    except (ValueError, TypeError):
        return _parsear_monto(str(valor))


def _parsear_filas_tabla(page, filas_elements):
    """
    Extrae datos de texto de cada fila de la tabla de tramitacion (cuaderno).
    Retorna lista de dicts con datos + referencia al elemento Playwright.
    Usado por Filtro 2 y Filtro 3 (requieren element para descarga PDF).

    Columnas (0-indexed):
      0=Folio, 1=Doc, 2=Anexo, 3=Etapa, 4=Tramite,
      5=Desc.Tramite, 6=Fec.Tramite, 7=Foja, 8=Georref.
    """
    filas_data = []
    for fila in filas_elements:
        try:
            celdas = fila.query_selector_all("td")
            if len(celdas) < 7:
                continue

            folio_text = celdas[0].inner_text().strip()
            folio_match = re.search(r"\d+", folio_text)
            folio = int(folio_match.group()) if folio_match else 0

            fecha_str = celdas[6].inner_text().strip() if len(celdas) > 6 else ""

            filas_data.append({
                "folio": folio,
                "etapa": celdas[3].inner_text().strip() if len(celdas) > 3 else "",
                "tramite": celdas[4].inner_text().strip() if len(celdas) > 4 else "",
                "desc_tramite": celdas[5].inner_text().strip() if len(celdas) > 5 else "",
                "fecha_str": fecha_str,
                "fecha": _parsear_fecha_dd_mm_yyyy(fecha_str),
                "foja": celdas[7].inner_text().strip() if len(celdas) > 7 else "",
                "element": fila,
            })
        except Exception as e:
            log.warning("Error parseando fila de tabla: %s", e)
    return filas_data


def _descargar_pdf_temporal(page, context, fila_element, etiqueta="temp"):
    """Descarga PDF de una fila a directorio temporal. Retorna ruta o None."""
    tmp_dir = tempfile.mkdtemp()
    ruta = os.path.join(tmp_dir, f"{etiqueta}.pdf")
    try:
        ok = descargar_pdf_de_fila(page, context, fila_element, ruta)
        if ok and os.path.exists(ruta) and os.path.getsize(ruta) > 500:
            return ruta
    except Exception as e:
        log.warning("Error descargando PDF temporal: %s", e)
    try:
        if os.path.exists(ruta):
            os.unlink(ruta)
        os.rmdir(tmp_dir)
    except OSError:
        pass
    return None


def _limpiar_pdf_temporal(ruta):
    """Elimina PDF temporal y su directorio."""
    if not ruta:
        return
    try:
        os.unlink(ruta)
        os.rmdir(os.path.dirname(ruta))
    except OSError:
        pass


def _cerrar_modal_detalle(page):
    """Cierra modal #modalDetalleCivil con 4 niveles de fallback.
    Retorna True si cerro normalmente, "reloaded" si tuvo que recargar pagina,
    False si fallo completamente.
    """
    # Verificar si el modal esta visible
    try:
        is_visible = page.evaluate('''() => {
            const modal = document.getElementById("modalDetalleCivil");
            return modal && modal.classList.contains("in");
        }''')
        if not is_visible:
            return True
    except Exception:
        pass

    # Nivel 1: jQuery Bootstrap
    try:
        page.evaluate('$("#modalDetalleCivil").modal("hide")')
        page.wait_for_timeout(500)
        if not page.query_selector("#modalDetalleCivil.in"):
            return True
    except Exception:
        pass

    # Nivel 2: Click en boton X
    try:
        close_btn = page.query_selector(
            "#modalDetalleCivil .close, #modalDetalleCivil button[data-dismiss='modal']"
        )
        if close_btn:
            close_btn.click()
            page.wait_for_timeout(500)
            if not page.query_selector("#modalDetalleCivil.in"):
                return True
    except Exception:
        pass

    # Nivel 3: Fuerza bruta DOM
    try:
        page.evaluate('''() => {
            var modal = document.getElementById("modalDetalleCivil");
            if (modal) {
                modal.classList.remove("in");
                modal.style.display = "none";
                modal.setAttribute("aria-hidden", "true");
            }
            var backdrops = document.querySelectorAll(".modal-backdrop");
            backdrops.forEach(function(b) { b.remove(); });
            document.body.classList.remove("modal-open");
            document.body.style.removeProperty("padding-right");
        }''')
        page.wait_for_timeout(300)
        if not page.query_selector("#modalDetalleCivil.in"):
            return True
    except Exception:
        pass

    # Nivel 4 (nuclear): reload pagina completa
    try:
        log.warning("  [MODAL] Niveles 1-3 fallaron, recargando pagina...")
        page.goto("https://oficinajudicialvirtual.pjud.cl/consultaunificadacausas.php")
        page.wait_for_selector("#btnConCivil", timeout=15000)
        page.click("#btnConCivil")
        page.wait_for_selector("#conCorte", timeout=10000)
        return "reloaded"
    except Exception as e:
        log.error("  [MODAL] ERROR CRITICO: No se pudo recargar pagina: %s", e)
        return False


def _cerrar_modal(page):
    """Wrapper legacy — llama a _cerrar_modal_detalle."""
    _cerrar_modal_detalle(page)


# =========================================================================
# PASO 0: MERGE DE REPORTES -> EXCEL MADRE
# =========================================================================

def _leer_reporte_m5(ruta_xlsx):
    """Lee causas de un reporte M5 (pestanas Regiones y/o RM)."""
    causas = []
    try:
        wb = openpyxl.load_workbook(ruta_xlsx, read_only=True)
    except Exception as e:
        log.warning("No se pudo abrir %s: %s", ruta_xlsx, e)
        return causas

    sheets_to_read = [
        name for name in wb.sheetnames
        if name.upper() in ("REGIONES", "RM")
    ]

    for sheet_name in sheets_to_read:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers_raw = [str(h) if h else "" for h in rows[0]]
        headers = [_HEADER_MAP.get(h, h) for h in headers_raw]

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            d = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    d[headers[i]] = val
            if not d.get("ROL"):
                continue

            corte = str(d.get("CORTE", ""))
            if corte and not corte.startswith("C.A."):
                d["CORTE"] = f"C.A. de {corte}"

            causas.append(d)

    wb.close()
    return causas


def paso0_merge_reportes():
    """
    Escanea Reportes/, importa causas nuevas al Excel madre.
    Retorna DataFrame completo del Excel madre.
    """
    log.info("=== PASO 0: Merge de Reportes ===")

    os.makedirs(CAUSAS_LIQ_DIR, exist_ok=True)
    os.makedirs(LIQUIDACIONES_DIR, exist_ok=True)

    if os.path.exists(EXCEL_MADRE):
        try:
            df = pd.read_excel(
                EXCEL_MADRE, sheet_name="_datos_internos",
                dtype=str, engine="openpyxl",
            )
        except (ValueError, KeyError):
            df = pd.read_excel(EXCEL_MADRE, dtype=str, engine="openpyxl")
        for col in _COLS_MADRE:
            if col not in df.columns:
                df[col] = ""
        log.info("Excel madre cargado: %d causas existentes", len(df))
    else:
        df = pd.DataFrame(columns=_COLS_MADRE)
        log.info("Excel madre creado desde cero")

    reportes_importados = set(df["origen_reporte"].dropna().unique())

    causas_eliminadas_historial = _cargar_historial_eliminadas()
    if causas_eliminadas_historial:
        log.info("  Historial de eliminadas: %d causas no se re-importaran",
                 len(causas_eliminadas_historial))

    if not os.path.isdir(REPORTES_DIR):
        log.warning("Directorio no encontrado: %s", REPORTES_DIR)
        return df

    archivos = sorted([
        f for f in os.listdir(REPORTES_DIR)
        if f.startswith("Reporte_") and f.endswith(".xlsx")
    ])

    nuevas_total = 0
    for archivo in archivos:
        if archivo in reportes_importados:
            log.info("  Saltando %s (ya importado)", archivo)
            continue

        ruta = os.path.join(REPORTES_DIR, archivo)
        causas_reporte = _leer_reporte_m5(ruta)
        if not causas_reporte:
            log.info("  %s: sin causas", archivo)
            continue

        nuevas = 0
        for causa in causas_reporte:
            rol = str(causa.get("ROL", "")).strip()
            anio = str(causa.get("AÑO", "")).strip()
            if not rol or not anio:
                continue

            corte_causa = str(causa.get("CORTE", ""))
            if "DESCONOCIDA" in corte_causa.upper():
                continue

            if corte_causa in CORTES_RM:
                continue

            existe = (
                (df["ROL"].astype(str).str.strip() == rol)
                & (df["AÑO"].astype(str).str.strip() == anio)
            ).any()
            if existe:
                continue

            # Verificar historial de eliminadas (causas borradas fisicamente
            # en runs anteriores, no deben re-importarse).
            if (rol, anio) in causas_eliminadas_historial:
                continue

            row = {col: "" for col in _COLS_MADRE}
            for col in _COLS_MADRE:
                if col in causa and causa[col] is not None:
                    row[col] = str(causa[col])

            if row.get("CBR_MOTIVO", "").strip():
                row["estado"] = "PENDIENTE_REVISION_MANUAL"
                row["log_decision"] = f"Audit11 CBR: {row['CBR_MOTIVO']}"
            else:
                row["estado"] = "PENDIENTE_FILTRO1"
            row["origen_reporte"] = archivo

            fecha = _parsear_fecha_publicacion(
                str(causa.get("FECHA_PUBLICACION", ""))
            )
            if fecha:
                row["fecha_remate"] = fecha.isoformat()

            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
            nuevas += 1

        log.info("  %s: %d causas nuevas importadas", archivo, nuevas)
        nuevas_total += nuevas

    log.info("Total causas nuevas importadas: %d", nuevas_total)
    log.info("Total causas en Excel madre: %d", len(df))
    return df


def _paso_refresh_deuda_m3(df):
    """Paso 0.5 - Refresca MONTO_DEUDA_CLP via M3 sobre PDFs locales.

    Itera causas activas con deuda vacia/$0; si el PDF de mandamiento existe
    en Descargas/, llama a M3 para extraer el monto y actualiza la fila.
    Recalcula _delta via _calcular_campos_derivados (formula unica) y guarda
    el Excel antes del lanzamiento de workers (crash safety).
    """
    if not _M3_DISPONIBLE:
        log.info("[PASO 0.5] Saltado (M3 no disponible)")
        return df

    estados_excluidos = {"ELIMINADA", "ELIMINAR", "EXCEDENTE_CONFIRMADO"}
    candidatas_idx = []
    for idx, row in df.iterrows():
        if _deuda_a_int(row.get("MONTO_DEUDA_CLP", "")) > 0:
            continue
        estado = str(row.get("estado", "")).strip()
        if estado in estados_excluidos:
            continue
        rol = str(row.get("ROL", "") or "").strip()
        ano = str(row.get("AÑO", "") or "").strip()
        if not rol or not ano:
            continue
        candidatas_idx.append(idx)

    if not candidatas_idx:
        log.info("[PASO 0.5] Sin candidatas (todas las causas activas tienen deuda > 0)")
        return df

    log.info("[PASO 0.5] Refresh deuda M3 - %d candidatas con MONTO_DEUDA_CLP vacio",
             len(candidatas_idx))

    n_eval = len(candidatas_idx)
    n_sin_pdf = 0
    n_m3_fallo = 0
    n_actualizadas = 0
    fecha_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    for idx in candidatas_idx:
        rol = str(df.at[idx, "ROL"] or "").strip()
        ano = str(df.at[idx, "AÑO"] or "").strip()
        nombre_pdf = f"C-{rol}-{ano}_MANDAMIENTO.pdf"
        ruta_pdf = os.path.join(DESCARGAS_DIR, nombre_pdf)
        if not os.path.exists(ruta_pdf):
            n_sin_pdf += 1
            continue
        try:
            texto = _extraer_texto_pdf(ruta_pdf)
            orig, valor, moneda = _extraer_de_mandamiento(texto)
        except Exception as e:
            log.warning("[PASO 0.5] C-%s-%s: M3 fallo - %s", rol, ano, e)
            n_m3_fallo += 1
            continue
        if not orig or valor is None:
            n_m3_fallo += 1
            continue
        try:
            if moneda == "UF":
                monto_clp = int(round(float(valor) * float(obtener_uf_hoy())))
            else:
                monto_clp = int(float(valor))
        except Exception as e:
            log.warning("[PASO 0.5] C-%s-%s: conversion fallo - %s", rol, ano, e)
            n_m3_fallo += 1
            continue
        if monto_clp <= 0:
            n_m3_fallo += 1
            continue

        # Idempotencia: si ya hay marca con mismo monto, no re-loguear ni sobreescribir
        log_actual = str(df.at[idx, "log_decision"] or "")
        monto_fmt = f"${monto_clp:,}".replace(",", ".")
        if "RefreshDeudaM3" in log_actual and monto_fmt in log_actual:
            log.info("[PASO 0.5] C-%s-%s: ya marcada con mismo monto, salteo", rol, ano)
            continue

        # El df se carga con dtype=str (L555/L558), guardar como string
        # para consistencia con las filas existentes y con _deuda_a_int.
        df.at[idx, "MONTO_DEUDA_CLP"] = str(monto_clp)
        msg = (f"RefreshDeudaM3 {fecha_str}: deuda actualizada de $0 a "
               f"{monto_fmt} (M3 sobre {nombre_pdf})")
        nuevo_log = f"{log_actual} | {msg}".strip(" |") if log_actual else msg
        df.at[idx, "log_decision"] = nuevo_log
        log.info("[PASO 0.5] C-%s-%s: $0 -> %s (M3 sobre %s)",
                 rol, ano, monto_fmt, nombre_pdf)
        n_actualizadas += 1

    log.info("[PASO 0.5] Resumen: %d evaluadas / %d sin PDF / %d M3 fallo / %d actualizadas",
             n_eval, n_sin_pdf, n_m3_fallo, n_actualizadas)

    if n_actualizadas > 0:
        # Recalcular _delta via la formula unica del filtrador
        df = _calcular_campos_derivados(df)
        # Persistir antes del lanzamiento de workers (crash safety)
        _guardar_excel_formateado(df)
        log.info("[PASO 0.5] Excel guardado con %d actualizaciones", n_actualizadas)

    return df


# =========================================================================
# FORMATEO EXCEL
# =========================================================================

def _calcular_campos_derivados(df):
    """Agrega columnas _delta y _dias_desde_remate al DataFrame."""
    hoy = date.today()
    deltas = []
    dias = []
    for _, row in df.iterrows():
        monto_acta = _deuda_a_int(row.get("monto_acta_remate", ""))
        deuda = _deuda_a_int(row.get("MONTO_DEUDA_CLP", ""))
        if monto_acta > 0 and deuda > 0:
            deltas.append(monto_acta - deuda)
        else:
            deltas.append(None)

        fr_str = str(row.get("fecha_remate", ""))
        try:
            fecha_r = date.fromisoformat(fr_str)
            dias.append((hoy - fecha_r).days)
        except (ValueError, TypeError):
            dias.append(None)

    df["_delta"] = deltas
    df["_dias_desde_remate"] = dias
    return df


def _escribir_hoja(ws, df, cols_spec, es_antiguedad=False):
    """Escribe datos + formato en una hoja de openpyxl."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=11)
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    data_align = Alignment(vertical="center", wrap_text=False)
    center_align = Alignment(horizontal="center", vertical="center")

    _CENTER_INTERNALS = {"COMUNA", "MONTO_DEUDA_CLP", "_delta", "ruta_liquidacion"}

    verde_fill_dias = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    amarillo_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    naranja_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
    rojo_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    ws.row_dimensions[1].height = 30
    for col_idx, (_, display_name, width) in enumerate(cols_spec, 1):
        cell = ws.cell(row=1, column=col_idx, value=display_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    monto_cols = set()
    center_cols = set()
    delta_col = None
    dias_col = None
    for col_idx, (internal, _, _) in enumerate(cols_spec, 1):
        if internal in ("monto_acta_remate", "MONTO_DEUDA_CLP", "_delta"):
            monto_cols.add(col_idx)
        if internal == "_delta":
            delta_col = col_idx
        if internal == "_dias_desde_remate":
            dias_col = col_idx
        if internal in _CENTER_INTERNALS:
            center_cols.add(col_idx)

    for row_idx, (_, data_row) in enumerate(df.iterrows(), 2):
        fill = zebra_fill if row_idx % 2 == 0 else no_fill

        for col_idx, (internal, _, _) in enumerate(cols_spec, 1):
            valor = data_row.get(internal)

            if col_idx in monto_cols and valor is not None:
                valor_num = None
                if isinstance(valor, (int, float)):
                    valor_num = int(valor) if not (isinstance(valor, float) and pd.isna(valor)) else None
                elif isinstance(valor, str) and valor.strip():
                    try:
                        valor_num = int(float(valor))
                    except (ValueError, TypeError):
                        valor_num = _parsear_monto(valor)
                        if valor_num == 0:
                            valor_num = None

                cell = ws.cell(row=row_idx, column=col_idx, value=valor_num)
                cell.number_format = '#,##0'

                if col_idx == delta_col and valor_num is not None:
                    if valor_num > 0:
                        cell.font = Font(name="Calibri", size=11, color="27AE60", bold=True)
                    else:
                        cell.font = Font(name="Calibri", size=11, color="E74C3C", bold=True)
                    cell.fill = fill
                    cell.alignment = center_align if col_idx in center_cols else data_align
                    continue
            else:
                display_val = "" if valor is None or (isinstance(valor, float) and pd.isna(valor)) else valor
                cell = ws.cell(row=row_idx, column=col_idx, value=display_val)

            cell.font = data_font
            cell.fill = fill
            cell.alignment = center_align if col_idx in center_cols else data_align

            if col_idx == dias_col and valor is not None:
                try:
                    dias_val = int(valor) if not (isinstance(valor, float) and pd.isna(valor)) else None
                except (ValueError, TypeError):
                    dias_val = None
                if dias_val is not None:
                    cell.value = dias_val
                    if dias_val < 14:
                        cell.fill = verde_fill_dias
                    elif dias_val <= 30:
                        cell.fill = amarillo_fill
                    elif dias_val <= 60:
                        cell.fill = naranja_fill
                    else:
                        cell.fill = rojo_fill

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(cols_spec))
    last_row = len(df) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"



def _escribir_hoja_revision_manual(ws, df, cols_spec):
    """Escribe datos + formato en la hoja Revision Manual (header amarillo)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=11)
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    data_align = Alignment(vertical="center", wrap_text=False)
    wrap_align = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, (_, display_name, width) in enumerate(cols_spec, 1):
        cell = ws.cell(row=1, column=col_idx, value=display_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, (_, data_row) in enumerate(df.iterrows(), 2):
        fill = zebra_fill if row_idx % 2 == 0 else no_fill

        for col_idx, (internal, _, _) in enumerate(cols_spec, 1):
            valor = data_row.get(internal)
            display_val = "" if valor is None or (isinstance(valor, float) and pd.isna(valor)) else valor
            cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
            cell.font = data_font
            cell.fill = fill
            # wrap_text para columna Detalle Auditoria
            if internal == "detalle_auditoria":
                cell.alignment = wrap_align
            else:
                cell.alignment = data_align

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(cols_spec))
    last_row = len(df) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"


def _registrar_en_historial_eliminadas(df, mascara_eliminadas):
    """Anexa al historial CSV las causas que se van a eliminar fisicamente.

    Args:
        df: DataFrame antes de eliminar.
        mascara_eliminadas: boolean Series, True para filas a eliminar.
    """
    indices = df[mascara_eliminadas].index.tolist()
    if not indices:
        return

    os.makedirs(os.path.dirname(HISTORIAL_ELIMINADAS), exist_ok=True)
    header_needed = not os.path.exists(HISTORIAL_ELIMINADAS)
    fecha_hoy = date.today().isoformat()

    with open(HISTORIAL_ELIMINADAS, 'a', encoding='utf-8', newline='') as f:
        if header_needed:
            f.write("ROL,ANO,fecha_eliminacion,motivo\n")
        for idx in indices:
            rol = str(df.at[idx, "ROL"] or "").strip()
            ano = str(df.at[idx, "AÑO"] or "").strip()
            log_dec = str(df.at[idx, "log_decision"] or "")
            # Primer linea del log_decision, sin saltos de linea ni comas
            motivo = log_dec.split('\n')[0].replace(',', ';').replace('"', "'")[:200]
            f.write(f"{rol},{ano},{fecha_hoy},{motivo}\n")


def _cargar_historial_eliminadas():
    """Carga el set de (ROL, ANO) de causas previamente eliminadas.

    Returns:
        set de tuplas (rol_str, ano_str). Vacio si el archivo no existe.
    """
    if not os.path.exists(HISTORIAL_ELIMINADAS):
        return set()
    eliminadas = set()
    try:
        with open(HISTORIAL_ELIMINADAS, encoding='utf-8') as f:
            next(f, None)  # skip header
            for linea in f:
                partes = linea.strip().split(',', 3)
                if len(partes) >= 2:
                    eliminadas.add((partes[0].strip(), partes[1].strip()))
    except Exception as e:
        log.warning("No se pudo cargar historial de eliminadas: %s", e)
    return eliminadas



def _leer_tabla_historial(page, fecha_remate):
    """
    Lee la tabla de historial (#historiaCiv) de la causa abierta en el modal.

    Input:
        page: objeto Playwright ya navegado al modal de detalle de causa
        fecha_remate: date -- fecha del remate segun el Excel madre

    Output:
        Tupla (filas_filtradas, total_filas, fecha_limite):
        - filas_filtradas: lista de dicts ordenada por folio desc (solo >= fecha_remate - 5d)
        - total_filas: int con el total de filas en la tabla (antes del filtro fecha)
        - fecha_limite: date usado como corte

    Estructura de cada fila en #historiaCiv (0-indexed):
        td[0] = Folio (numero)
        td[1] = Doc. (icono PDF -- contiene <form> con input hidden name="dtaDoc" value="JWT")
        td[2] = Anexo
        td[3] = Etapa ("Apremio", "Excepciones", etc.)
        td[4] = Tramite ("Resolucion", "Escrito", "(COM)Comparendo", etc.)
        td[5] = Desc. Tramite (**CAMPO CLAVE**)
        td[6] = Fec. Tramite (dd/mm/yyyy)
        td[7] = Foja (numero)
        td[8] = Georref.
    """
    fecha_limite = fecha_remate - timedelta(days=5)

    # Seleccionar cuaderno Apremio ANTES de leer la tabla.
    # Sin esto, la tabla muestra el cuaderno por defecto (Principal/Excepciones)
    # y las filas relevantes (actas, suspensiones, liquidaciones) no aparecen.
    try:
        seleccionar_cuaderno(page, "Apremio")
        page.wait_for_timeout(1500)
    except Exception as e:
        log.warning("No se pudo seleccionar cuaderno Apremio: %s", e)

    # Click en tab #historiaCiv si no esta activo
    try:
        tab = page.query_selector('a[href="#historiaCiv"]')
        if tab:
            tab.click()
            page.wait_for_timeout(1000)
    except Exception as e:
        log.warning("No se pudo hacer click en tab #historiaCiv: %s", e)

    # Esperar que la tabla cargue
    selectores_tabla = [
        "#historiaCiv table tbody tr",
        "#historiaCiv > div > div > table tbody tr",
    ]

    filas_elements = []
    for sel in selectores_tabla:
        try:
            page.wait_for_selector(sel, timeout=10000)
            filas_elements = page.query_selector_all(sel)
            if filas_elements:
                break
        except Exception:
            continue

    if not filas_elements:
        log.warning("Tabla #historiaCiv no encontrada o vacia")
        return [], 0, fecha_limite

    # Parsear filas
    resultado = []
    for fila in filas_elements:
        try:
            celdas = fila.query_selector_all("td")
            if len(celdas) < 7:
                continue

            # td[0]: Folio
            folio_text = celdas[0].inner_text().strip()
            folio_match = re.search(r"\d+", folio_text)
            folio = int(folio_match.group()) if folio_match else 0

            # td[1]: Doc -- tiene_pdf y jwt_token
            tiene_pdf = False
            jwt_token = ""
            form_el = celdas[1].query_selector("form")
            if form_el:
                input_el = form_el.query_selector('input[name="dtaDoc"]')
                if input_el:
                    tiene_pdf = True
                    jwt_token = input_el.get_attribute("value") or ""

            # td[3]: Etapa
            etapa = celdas[3].inner_text().strip() if len(celdas) > 3 else ""

            # td[4]: Tramite
            tramite = celdas[4].inner_text().strip() if len(celdas) > 4 else ""

            # td[5]: Desc. Tramite
            desc_tramite = celdas[5].inner_text().strip() if len(celdas) > 5 else ""

            # td[6]: Fec. Tramite
            fecha_str = celdas[6].inner_text().strip() if len(celdas) > 6 else ""
            fecha_tramite = _parsear_fecha_dd_mm_yyyy(fecha_str)

            if not fecha_tramite:
                log.warning("  Fila folio %d: fecha no parseada '%s', saltando",
                            folio, fecha_str)
                continue

            # td[7]: Foja
            foja_text = celdas[7].inner_text().strip() if len(celdas) > 7 else "0"
            foja_match = re.search(r"\d+", foja_text)
            foja = int(foja_match.group()) if foja_match else 0

            # Filtrar por fecha
            if fecha_tramite < fecha_limite:
                continue

            resultado.append({
                "folio": folio,
                "tiene_pdf": tiene_pdf,
                "jwt_token": jwt_token,
                "etapa": etapa,
                "tramite": tramite,
                "desc_tramite": desc_tramite,
                "fecha_tramite": fecha_tramite,
                "foja": foja,
            })

        except Exception as e:
            log.warning("Error parseando fila de #historiaCiv: %s", e)

    # Ordenar por folio descendente (mas reciente primero)
    resultado.sort(key=lambda x: x["folio"], reverse=True)

    total_filas = len(filas_elements)
    log.info("  #historiaCiv: %d filas leidas (de %d totales, filtro fecha >= %s)",
             len(resultado), total_filas, fecha_limite.isoformat())

    return resultado, total_filas, fecha_limite


def _leer_tabla_historial_completa(page):
    """
    Lee TODAS las filas de #historiaCiv SIN filtro de fecha.
    Usado por --reaudit para reprocesar causas PENDIENTE_ACTA.
    Retorna lista de dicts (misma estructura que _leer_tabla_historial).
    """
    # Seleccionar cuaderno Apremio
    try:
        seleccionar_cuaderno(page, "Apremio")
        page.wait_for_timeout(1500)
    except Exception as e:
        log.warning("No se pudo seleccionar cuaderno Apremio: %s", e)

    # Click en tab #historiaCiv
    try:
        tab = page.query_selector('a[href="#historiaCiv"]')
        if tab:
            tab.click()
            page.wait_for_timeout(1000)
    except Exception as e:
        log.warning("No se pudo hacer click en tab #historiaCiv: %s", e)

    selectores_tabla = [
        "#historiaCiv table tbody tr",
        "#historiaCiv > div > div > table tbody tr",
    ]
    filas_elements = []
    for sel in selectores_tabla:
        try:
            page.wait_for_selector(sel, timeout=10000)
            filas_elements = page.query_selector_all(sel)
            if filas_elements:
                break
        except Exception:
            continue

    if not filas_elements:
        log.warning("Tabla #historiaCiv no encontrada o vacia (reaudit)")
        return []

    resultado = []
    for fila in filas_elements:
        try:
            celdas = fila.query_selector_all("td")
            if len(celdas) < 7:
                continue

            folio_text = celdas[0].inner_text().strip()
            folio_match = re.search(r"\d+", folio_text)
            folio = int(folio_match.group()) if folio_match else 0

            etapa = celdas[3].inner_text().strip() if len(celdas) > 3 else ""
            tramite = celdas[4].inner_text().strip() if len(celdas) > 4 else ""
            desc_tramite = celdas[5].inner_text().strip() if len(celdas) > 5 else ""
            fecha_str = celdas[6].inner_text().strip() if len(celdas) > 6 else ""
            fecha_tramite = _parsear_fecha_dd_mm_yyyy(fecha_str)
            foja_text = celdas[7].inner_text().strip() if len(celdas) > 7 else "0"
            foja_match = re.search(r"\d+", foja_text)
            foja = int(foja_match.group()) if foja_match else 0

            resultado.append({
                "folio": folio,
                "tiene_pdf": False,
                "jwt_token": "",
                "etapa": etapa,
                "tramite": tramite,
                "desc_tramite": desc_tramite,
                "fecha_tramite": fecha_tramite,
                "foja": foja,
            })
        except Exception as e:
            log.warning("Error parseando fila de #historiaCiv (reaudit): %s", e)

    resultado.sort(key=lambda x: x["folio"], reverse=True)
    log.info("  #historiaCiv (reaudit): %d filas leidas (sin filtro fecha)", len(resultado))
    return resultado


# =========================================================================
# KEYWORDS Y SEÑALES
# =========================================================================

KEYWORDS_LIQUIDACION = [
    "ordena liquidar el crédito",
    "ordena liquidar el credito",
    "liquidacion (credito)",
    "liquidación (crédito)",
    "pone en conocimiento liquidación de crédito",
    "pone en conocimiento liquidacion de credito",
    "liquidación de crédito",
    "liquidacion de credito",
    "solicita liquidación",
    "solicita liquidacion",
    "resolución de liquidación",
    "resolucion de liquidacion",
    "ordena liquidar",
]

KEYWORDS_ACTA_REMATE = [
    "acta de remate",
]

KEYWORDS_INFO = {
    "giro cheque": "Giro cheque",
    "gírese": "Girese",
    "girese": "Girese",
    "consigna precio de remate": "Precio consignado",
    "objeta liquidación": "Liquidacion objetada",
    "objeta liquidacion": "Liquidacion objetada",
    "nulidad de lo obrado": "Nulidad detectada",
    "abandono procedimiento": "Abandono procedimiento",
    "desistimiento": "Desistimiento",
    "da cuenta de pago": "Deudor pago (posible cierre sin remate)",
}


# =========================================================================
# FUNCION UNIFICADA: evalua F1+F2 sobre filas de tabla
# =========================================================================

def _evaluar_filtros_tabla(filas):
    """Evalua F1+F2 sobre las filas de la tabla.

    IMPORTANTE: La deteccion de acta de remate se hace en pasada separada
    porque "cargo al credito" en el acta ANULA cualquier otra senal.
    Si hay acta, SIEMPRE retornamos NECESITA_PDF_ACTA para que el caller
    descargue el PDF y verifique cargo al credito.

    Returns:
        tuple: (estado, folio, fecha, detalle) o (None, None, None, None) si no hay senales

    Estados posibles:
        'PENDIENTE_LIQUIDACION' — senal de liquidacion encontrada
        'NECESITA_PDF_ACTA' — acta de remate detectada, necesita descarga de PDF
        'ELIMINAR' — suspension, no postores, o reprogramado
        'PENDIENTE_REVISION_MANUAL' — liquidacion concursal
    """
    # === PASADA 1: Buscar acta de remate en CUALQUIER fila ===
    acta_encontrada = None
    for fila in filas:
        desc_lower = fila.get("desc_tramite", "").strip().lower()
        for kw in KEYWORDS_ACTA_REMATE:
            if kw in desc_lower:
                acta_encontrada = {
                    "folio": fila.get("folio", ""),
                    "fecha": fila.get("fecha_tramite", ""),
                    "desc": fila.get("desc_tramite", "").strip(),
                }
                break
        if acta_encontrada:
            break  # Tomar la primera (mas reciente)

    # Si hay acta, SIEMPRE retornar NECESITA_PDF_ACTA
    # El caller descargara el PDF y verificara cargo al credito
    if acta_encontrada:
        return ("NECESITA_PDF_ACTA",
                acta_encontrada["folio"],
                acta_encontrada["fecha"],
                f"Acta de remate detectada en tabla (folio {acta_encontrada['folio']}, {acta_encontrada['fecha']})")

    # === PASADA 2: Sin acta → evaluar senales F1 normales (mas reciente primero) ===
    for fila in filas:
        desc = fila.get("desc_tramite", "").strip()
        desc_lower = desc.lower()
        tramite = fila.get("tramite", "").strip()
        tramite_lower = tramite.lower()
        folio = fila.get("folio", "")
        fecha = fila.get("fecha_tramite", "")
        etapa = fila.get("etapa", "").strip()
        etapa_lower = etapa.lower()

        # --- Deteccion liquidacion concursal (Ley 20.720) ---
        if "tramitación liquidación" in etapa_lower or "tramitacion liquidacion" in etapa_lower:
            if "resolución de liquidación" in desc_lower or "resolucion de liquidacion" in desc_lower:
                return ("PENDIENTE_REVISION_MANUAL", folio, fecha,
                        f"Posible liquidacion concursal Ley 20.720 (folio {folio}, {fecha})")

        # --- Senal 4: Liquidacion de credito ---
        for kw in KEYWORDS_LIQUIDACION:
            if kw in desc_lower:
                return ("PENDIENTE_LIQUIDACION", folio, fecha,
                        f"Senal 4: \"{desc}\" (folio {folio}, {fecha}) -> directo a Filtro 3")

        # --- Senal 1: No postores ---
        if "no postores" in desc_lower or "certificado de no postores" in desc_lower:
            return ("ELIMINAR", folio, fecha,
                    f"Sin postores: \"{desc}\" (folio {folio}, {fecha})")

        # --- Senal 2: Suspension concedida (solo Resolucion, NO Escrito) ---
        if tramite_lower.startswith("resolución") or tramite_lower.startswith("resolucion"):
            if "suspende" in desc_lower or "reprograma" in desc_lower:
                return ("ELIMINAR", folio, fecha,
                        f"Suspension/Reprograma: \"{desc}\" (folio {folio}, {fecha})")

        # --- Senal 3: Nuevo dia y hora ---
        if "nuevo día y hora" in desc_lower or "nuevo dia y hora" in desc_lower:
            return ("ELIMINAR", folio, fecha,
                    f"Reprogramado: \"{desc}\" (folio {folio}, {fecha})")

    return (None, None, None, None)


def _recopilar_info_tabla(filas):
    """Segundo pase: recopilar senales informativas (no determinantes)."""
    info_items = []
    for fila in filas:
        desc_lower = fila.get("desc_tramite", "").lower()
        folio = fila.get("folio", "")
        for kw, label in KEYWORDS_INFO.items():
            if kw in desc_lower:
                info_items.append(f"INFO: {label} (folio {folio})")
                break  # Solo un match por fila
    return info_items


def _aplicar_filtro1(tabla_historial, causa, modo_auditoria):
    """Wrapper de compatibilidad: aplica _evaluar_filtros_tabla sobre filas filtradas por fecha.

    Input:
        tabla_historial: lista de dicts del _leer_tabla_historial
        causa: dict con datos de la causa del Excel madre
        modo_auditoria: bool (--primera-run)

    Output:
        Tupla (nuevo_estado, detalle_auditoria)
    """
    fecha_remate_str = str(causa.get("fecha_remate", ""))
    try:
        fecha_remate = date.fromisoformat(fecha_remate_str)
    except (ValueError, TypeError):
        return ("PENDIENTE_ACTA",
                "Filtro 1: fecha_remate no parseada, no se puede filtrar")

    # Solo filas con fecha >= fecha_remate - 2 dias
    fecha_corte = fecha_remate - timedelta(days=2)
    filas_relevantes = [f for f in tabla_historial
                        if f.get("fecha_tramite") and f["fecha_tramite"] >= fecha_corte]

    estado, folio, fecha, detalle = _evaluar_filtros_tabla(filas_relevantes)

    if estado is None:
        # Sin senales principales -> recopilar informativas
        info = _recopilar_info_tabla(filas_relevantes)
        detalle_extra = "; ".join(info) if info else ""
        return ("PENDIENTE_ACTA",
                f"Sin senales de suspension o no postores -> paso a Filtro 2"
                + (f". {detalle_extra}" if detalle_extra else ""))

    if estado == "ELIMINAR":
        return ("ELIMINAR", detalle)
    if estado == "NECESITA_PDF_ACTA":
        return ("NECESITA_PDF_ACTA", detalle, folio)
    if estado == "PENDIENTE_REVISION_MANUAL":
        return ("PENDIENTE_REVISION_MANUAL", detalle)

    return (estado, detalle)


# =========================================================================
def _encontrar_fila_por_folio(page, folio_target):
    """Encuentra el <tr> de #historiaCiv que tiene el folio indicado."""
    folio_str = str(folio_target)
    try:
        selectores = [
            "#historiaCiv table tbody tr",
            "#historiaCiv > div > div > table tbody tr",
        ]
        for sel in selectores:
            filas_tr = page.query_selector_all(sel)
            if not filas_tr:
                continue
            for tr in filas_tr:
                tds = tr.query_selector_all("td")
                if len(tds) < 2:
                    continue
                folio_texto = tds[0].inner_text().strip()
                folio_num = re.search(r"\d+", folio_texto)
                if folio_num and folio_num.group() == folio_str:
                    return tr
    except Exception as e:
        log.warning("  Error buscando fila folio %s: %s", folio_str, e)
    return None


def _descargar_pdf_desde_fila(page, fila_tr, download_dir, filename):
    """Descarga el primer PDF disponible de una fila de #historiaCiv.

    Busca form con docuS.php (resoluciones/actuaciones/comparendos) primero,
    luego docuN.php (escritos) como fallback.
    Evita docCertificadoEscrito.php (certificados, no relevantes).

    Returns:
        filepath del PDF guardado, o None si falla.
    """
    try:
        tds = fila_tr.query_selector_all("td")
        if len(tds) < 2:
            return None

        td_doc = tds[1]  # columna Doc

        # Buscar form con docuS.php primero (actas, resoluciones)
        form = td_doc.query_selector('form[action*="docuS.php"]')
        if not form:
            form = td_doc.query_selector('form[action*="docuN.php"]')
        if not form:
            # Ultimo fallback: cualquier form (excepto certificados)
            forms = td_doc.query_selector_all("form")
            for f in forms:
                action = f.get_attribute("action") or ""
                if "docCertificado" not in action:
                    form = f
                    break
        if not form:
            log.warning("  No se encontro form de descarga en fila")
            return None

        # Obtener link clickeable dentro del form
        link = form.query_selector('a[onclick*="submit"]')
        if not link:
            link = form.query_selector("a")
        if not link:
            log.warning("  No se encontro link clickeable en form de descarga")
            return None

        filepath = os.path.join(download_dir, filename)

        # Click abre PDF en nueva pestana (target="p1", "p2", etc.)
        try:
            with page.context.expect_page(timeout=15000) as new_page_info:
                link.click()

            new_page = new_page_info.value
            new_page.wait_for_load_state("load", timeout=15000)
            pdf_url = new_page.url

            # Intentar descargar via response del popup
            try:
                response = new_page.request.get(pdf_url, timeout=20000)
                if response.ok and len(response.body()) > 500:
                    with open(filepath, "wb") as f:
                        f.write(response.body())
                    log.info("  PDF descargado: %s (%d bytes)",
                             filename, len(response.body()))
                    new_page.close()
                    return filepath
            except Exception:
                pass

            new_page.close()
        except Exception as e:
            log.warning("  Popup PDF no capturado: %s. Intentando expect_download...", e)

        # Fallback: expect_download (algunos browsers descargan directo)
        try:
            with page.expect_download(timeout=15000) as dl:
                link.click()
            dl.value.save_as(filepath)
            if os.path.exists(filepath) and os.path.getsize(filepath) > 500:
                log.info("  PDF descargado (download): %s", filename)
                return filepath
        except Exception as e:
            log.warning("  expect_download fallo: %s", e)

        return None

    except Exception as e:
        log.warning("  Error descargando PDF desde fila: %s", e)
        return None


def _descargar_pdf_acta_por_folio(page, context, folio_target, causa, download_dir):
    """Descarga el PDF del acta de remate identificado por su folio.

    Busca la fila por folio en #historiaCiv y descarga el PDF usando
    la mecanica de form submit (docuS.php/docuN.php).

    Returns:
        filepath del PDF descargado, o None si falla.
    """
    os.makedirs(download_dir, exist_ok=True)

    fila_tr = _encontrar_fila_por_folio(page, folio_target)
    if not fila_tr:
        log.warning("  Folio %s no encontrado en tabla para descarga PDF", folio_target)
        return None

    rol = str(causa.get("ROL", "unknown"))
    anio = str(causa.get("AÑO", ""))
    filename = f"acta_{rol}_{anio}.pdf"

    return _descargar_pdf_desde_fila(page, fila_tr, download_dir, filename)


# =========================================================================
# FILTRO 2 LEGACY: ACTA DE REMATE (fecha_remate + 3 dias)
# =========================================================================

def filtro2_acta_remate(page, context, causa, filas_data):
    """
    Filtro 2: Busca Acta de Remate, compara montos.

    Retorna (nuevo_estado, log_decision).
    Si nuevo_estado es None -> causa debe ser ELIMINADA.
    """
    rol = causa.get("ROL", "")

    fecha_remate = None
    fr_str = causa.get("fecha_remate", "")
    if fr_str:
        try:
            fecha_remate = date.fromisoformat(fr_str)
        except ValueError:
            pass

    deuda = _deuda_a_int(causa.get("MONTO_DEUDA_CLP", "0"))

    candidatas = []
    for fila in filas_data:
        if fecha_remate and fila["fecha"]:
            delta = (fila["fecha"] - fecha_remate).days
            if 0 <= delta <= 3:
                candidatas.append(fila)
        elif not fecha_remate:
            candidatas.append(fila)

    if not candidatas:
        return (
            "PENDIENTE_ACTA",
            f"EN ESPERA: No hay filas en rango fecha_remate +3"
            f" (ultima revision {date.today().isoformat()})",
        )

    def _prioridad(f):
        desc = f["desc_tramite"].lower()
        if "acta de remate" in desc or "acta de subasta" in desc:
            return 0
        if "mero" in desc:
            return 1
        if "certificado" in desc:
            return 2
        return 3

    candidatas.sort(key=_prioridad)

    for fila in candidatas:
        ruta_pdf = _descargar_pdf_temporal(
            page, context, fila["element"],
            f"acta_{rol}_{fila['folio']}",
        )
        if not ruta_pdf:
            continue

        texto_pdf = _extraer_texto_pdf(ruta_pdf)
        _limpiar_pdf_temporal(ruta_pdf)

        if not texto_pdf:
            return (
                "PENDIENTE_ACTA",
                f"PENDIENTE MANUAL: Acta encontrada pero PDF es imagen"
                f" sin texto - folio {fila['folio']}",
            )

        if not PATRON_ACTA_REMATE.search(texto_pdf):
            continue

        log.info("  %s: Acta de remate encontrada en folio %d", rol, fila["folio"])

        m_monto = PATRON_MONTO_ACTA.search(texto_pdf)
        if not m_monto:
            return (
                "PENDIENTE_ACTA",
                f"PENDIENTE MANUAL: Acta encontrada pero monto"
                f" no extraido - folio {fila['folio']}",
            )

        monto_acta = _parsear_monto(m_monto.group(1))
        causa["monto_acta_remate"] = str(monto_acta)

        if monto_acta > deuda > 0:
            excedente = monto_acta - deuda
            return (
                "PENDIENTE_LIQUIDACION",
                f"Filtro 2 OK: Acta ${monto_acta:,} > Deuda ${deuda:,}"
                f" - excedente potencial ${excedente:,}",
            )
        else:
            return (
                None,
                f"ELIMINADA: Monto acta (${monto_acta:,})"
                f" <= deuda (${deuda:,}), sin excedente",
            )

    return (
        "PENDIENTE_ACTA",
        f"EN ESPERA: Acta de remate no encontrada aun"
        f" (ultima revision {date.today().isoformat()})",
    )


# =========================================================================
# FILTRO 3: LIQUIDACION — Descarga PDF + extraccion de saldo
# =========================================================================

# Keywords para buscar la fila de liquidacion REAL en la tabla OJV
# (solo el tramite con calculos, NO resoluciones/ordenes del juez)
_KEYWORDS_LIQUIDACION_PDF = [
    "liquidacion (credito)",
    "liquidación (crédito)",
    "liquidacion(credito)",
    "liquidación(crédito)",
]

# Regex principal: "Saldo a favor del ejecutado" + monto (re.DOTALL por saltos de linea PyMuPDF)
_REGEX_SALDO_FAVOR = re.compile(
    r'[Ss]aldo\s+a\s+favor\s+de(?:l)?\s+(?:[Ee]jecutado|[Dd]emandado).*?(\d{1,3}(?:\.\d{3})+)',
    re.DOTALL,
)

# Patron negativo: saldo en contra / adeudado (ejecutado sigue debiendo)
_REGEX_SALDO_CONTRA = re.compile(
    r'[Ss]aldo\s+(?:en\s+contra|adeudado|pendiente)\s+(?:del\s+)?[Ee]jecutado.*?(\d{1,3}(?:\.\d{3})+)',
    re.DOTALL,
)

# Patron: no existe saldo / sin excedente
_REGEX_SIN_SALDO = re.compile(
    r'no\s+(?:existe|hay|queda)\s+(?:saldo|excedente|remanente)',
    re.IGNORECASE,
)

# Patron Tipo B: "Credito/Capital adeudado al [fecha]... monto"
_REGEX_CREDITO_ADEUDADO_MONTO = re.compile(
    r'(?:[Cc]r[eé]dito|CAPITAL|[Cc]apital)\s+[Aa]deudado\s+al\s+[^$\d]*?(\d{1,3}(?:\.\d{3}){2,})',
    re.DOTALL,
)


def _seleccionar_causas_f3(df):
    """Filtra causas elegibles para Filtro 3: TODAS las PENDIENTE_LIQUIDACION.

    Returns:
        list of (df_index, causa_dict)
    """
    resultado = []
    for idx, row in df.iterrows():
        if row["estado"] == "PENDIENTE_LIQUIDACION":
            resultado.append((idx, row.to_dict()))
    return resultado


def _buscar_fila_liquidacion(page):
    """Busca la fila de 'Liquidacion (Credito)' mas reciente en #historiaCiv.

    Busca en columnas tramite (td[4]) y desc_tramite (td[5]) por las keywords
    especificas del documento de liquidacion real (no resoluciones del juez).

    Returns:
        dict con folio, desc, fecha, fila_tr (ElementHandle) o None si no encuentra.
    """
    try:
        selectores = [
            "#historiaCiv table tbody tr",
            "#historiaCiv > div > div > table tbody tr",
        ]
        for sel in selectores:
            filas_tr = page.query_selector_all(sel)
            if not filas_tr:
                continue

            # Parsear y ordenar por folio desc
            filas_parsed = []
            for tr in filas_tr:
                tds = tr.query_selector_all("td")
                if len(tds) < 6:
                    continue
                folio_texto = tds[0].inner_text().strip()
                folio_match = re.search(r"\d+", folio_texto)
                folio = int(folio_match.group()) if folio_match else 0
                tramite = tds[4].inner_text().strip() if len(tds) > 4 else ""
                desc = tds[5].inner_text().strip() if len(tds) > 5 else ""
                fecha = tds[6].inner_text().strip() if len(tds) > 6 else ""
                filas_parsed.append({
                    "folio": folio,
                    "tramite": tramite,
                    "desc": desc,
                    "fecha": fecha,
                    "fila_tr": tr,
                })

            filas_parsed.sort(key=lambda x: x["folio"], reverse=True)

            for fila in filas_parsed:
                # Buscar en tramite Y desc (puede aparecer en cualquiera)
                texto_buscar = (fila["tramite"] + " " + fila["desc"]).lower()
                for kw in _KEYWORDS_LIQUIDACION_PDF:
                    if kw in texto_buscar:
                        return fila
    except Exception as e:
        log.warning("  [F3] Error buscando fila de liquidacion: %s", e)
    return None


def _descargar_pdf_liquidacion(page, fila_info, causa, output_dir):
    """Descarga PDF de liquidacion desde la fila encontrada.

    Returns:
        filepath del PDF descargado, o None si falla.
    """
    os.makedirs(output_dir, exist_ok=True)
    rol = str(causa.get("ROL", "unknown"))
    anio = str(causa.get("AÑO", ""))
    filename = f"{rol}-{anio}_liquidacion.pdf"
    return _descargar_pdf_desde_fila(page, fila_info["fila_tr"], output_dir, filename)


def _analizar_pdf_liquidacion(ruta_pdf, causa):
    """Extrae texto del PDF de liquidacion y busca saldo/excedente.

    1. Extrae texto (PyMuPDF nativo + OCR fallback)
    2. Guarda raw en LIQUIDACIONES_RAW_DIR
    3. Aplica regex para saldo

    Returns:
        dict con texto_raw, ruta_raw, saldo_encontrado, monto_saldo,
             patron_usado, sin_excedente
    """
    resultado = {
        "texto_raw": "",
        "ruta_raw": "",
        "saldo_encontrado": False,
        "monto_saldo": None,
        "patron_usado": None,
        "sin_excedente": False,
        "tipo_liquidacion": None,  # "A" (con comparacion) o "B" (solo deuda)
        "credito_adeudado": None,  # monto deuda real (Tipo B)
    }

    rol = str(causa.get("ROL", "unknown"))
    anio = str(causa.get("AÑO", ""))

    # --- Extraer texto ---
    try:
        doc = fitz.open(ruta_pdf)
        texto = ""
        for pagina in doc:
            texto += pagina.get_text()

        # Detectar si necesita OCR
        texto_lower = texto.lower()
        tiene_contenido = (
            ('$' in texto and any(c.isdigit() for c in texto)) or
            ('liquidaci' in texto_lower) or
            ('capital' in texto_lower) or
            ('saldo' in texto_lower) or
            ('costas' in texto_lower) or
            ('excedente' in texto_lower)
        )

        if not tiene_contenido:
            log.info("    [F3] PDF sin contenido util en texto nativo. Intentando OCR...")
            texto_ocr = _ocr_pdf(doc)
            if texto_ocr and len(texto_ocr.strip()) > 50:
                log.info("    [F3] OCR exitoso: %d chars", len(texto_ocr))
                texto = texto_ocr
            else:
                log.info("    [F3] OCR fallo o sin texto util")

        doc.close()
    except Exception as e:
        log.warning("  [F3] Error leyendo PDF liquidacion: %s", e)
        return resultado

    resultado["texto_raw"] = texto

    # --- Guardar raw ---
    try:
        os.makedirs(LIQUIDACIONES_RAW_DIR, exist_ok=True)
        raw_filename = f"{rol}-{anio}_liquidacion_raw.txt"
        ruta_raw = os.path.join(LIQUIDACIONES_RAW_DIR, raw_filename)
        with open(ruta_raw, "w", encoding="utf-8") as f:
            f.write(texto)
        resultado["ruta_raw"] = ruta_raw
    except Exception as e:
        log.warning("  [F3] Error guardando raw: %s", e)

    if not texto.strip():
        return resultado

    # --- 1. Buscar "Saldo a favor del ejecutado" + monto ---
    match = _REGEX_SALDO_FAVOR.search(texto)
    if match:
        monto_str = match.group(1).replace(".", "")
        try:
            monto = int(monto_str)
            if 0 < monto <= 2_000_000_000:
                resultado["saldo_encontrado"] = True
                resultado["monto_saldo"] = monto
                resultado["patron_usado"] = "saldo a favor del ejecutado"
                resultado["tipo_liquidacion"] = "A"
                return resultado
        except ValueError:
            pass

    # --- 2. Buscar "Saldo en contra/adeudado" (ejecutado sigue debiendo) ---
    match = _REGEX_SALDO_CONTRA.search(texto)
    if match:
        resultado["saldo_contra"] = True
        resultado["saldo_encontrado"] = True
        resultado["monto_saldo"] = 0
        resultado["patron_usado"] = "saldo en contra del ejecutado"
        return resultado

    # --- 3. Buscar "no existe saldo/excedente" ---
    if _REGEX_SIN_SALDO.search(texto):
        resultado["sin_excedente"] = True
        resultado["saldo_encontrado"] = True
        resultado["monto_saldo"] = 0
        resultado["patron_usado"] = "sin saldo explicito"
        return resultado

    # --- 4. Buscar Tipo B: "Credito adeudado al [fecha]... monto" ---
    matches_credito = list(_REGEX_CREDITO_ADEUDADO_MONTO.finditer(texto))
    if matches_credito:
        # Tomar el match con monto MAS GRANDE (total > subtotales)
        mejor_credito = 0
        for m in matches_credito:
            try:
                val = int(m.group(1).replace(".", ""))
                if val > mejor_credito and val <= 2_000_000_000:
                    mejor_credito = val
            except ValueError:
                continue

        if mejor_credito > 0:
            # Regla de sanidad: credito debe ser >= 50% de deuda original
            deuda_original = _deuda_a_int(causa.get("MONTO_DEUDA_CLP", ""))
            if deuda_original > 0 and mejor_credito < deuda_original * 0.5:
                log.warning("    [F3] SANIDAD: credito $%s < 50%% deuda $%s. Posible error de extraccion.",
                            f"{mejor_credito:,}", f"{deuda_original:,}")
                return resultado  # Cae al fallback PENDIENTE_REVISION_MANUAL

            if mejor_credito < 100_000:
                log.warning("    [F3] SANIDAD: credito $%s sospechosamente bajo.",
                            f"{mejor_credito:,}")
                return resultado

            resultado["tipo_liquidacion"] = "B"
            resultado["credito_adeudado"] = mejor_credito
            resultado["patron_usado"] = "credito adeudado (Tipo B)"
            return resultado

    return resultado


def _procesar_causa_f3(page, context, causa, worker_id=0):
    """Orquesta Filtro 3 para una causa: navegar OJV, buscar fila, descargar PDF, analizar.

    Returns:
        dict con resultado: estado, log_decision, monto_saldo, ruta_liquidacion
    """
    rol = str(causa.get("ROL", ""))
    anio = str(causa.get("AÑO", ""))
    wtag = f"[W{worker_id}] " if worker_id > 0 else ""
    tag = f"{wtag}[F3] C-{rol}-{anio}"

    resultado = {
        "estado": causa.get("estado", "PENDIENTE_LIQUIDACION"),
        "log_decision": causa.get("log_decision", ""),
        "monto_liquidacion_saldo": "",
        "monto_credito_liquidado": "",
        "ruta_liquidacion": causa.get("ruta_liquidacion", ""),
    }

    # --- Buscar fila de liquidacion REAL en tabla ---
    log.info("  %s: Buscando Liquidacion (Credito) en tabla...", tag)
    fila_liq = _buscar_fila_liquidacion(page)

    if not fila_liq:
        log.info("  %s: Liquidacion (Credito) no encontrada en tabla. Pendiente.", tag)
        # NO cambiar estado — se reintentara en futuros runs
        return resultado

    folio = fila_liq["folio"]
    log.info("  %s: Liquidacion (Credito) encontrada en folio %d. Descargando...", tag, folio)

    # --- Descargar PDF ---
    pdf_path = _descargar_pdf_liquidacion(page, fila_liq, causa, LIQUIDACIONES_DIR)

    if not pdf_path:
        log.info("  %s: PDF no descargable (folio %d)", tag, folio)
        resultado["log_decision"] = (
            f"F3: Liquidacion en tabla (folio {folio}) pero PDF no descargable"
        )
        return resultado

    resultado["ruta_liquidacion"] = pdf_path
    log.info("  %s: PDF descargado. Analizando...", tag)

    # --- Analizar PDF ---
    analisis = _analizar_pdf_liquidacion(pdf_path, causa)

    n_chars = len(analisis["texto_raw"])
    usa_ocr = "OCR" if analisis["texto_raw"] and not any(
        c in analisis["texto_raw"][:200] for c in ['liquidaci', 'capital', '$']
    ) else "nativo"
    log.info("  %s: Texto extraido (%d chars, %s)", tag, n_chars, usa_ocr)

    # --- Decidir estado ---
    if analisis["saldo_encontrado"] and analisis["monto_saldo"] and analisis["monto_saldo"] > 0:
        monto = analisis["monto_saldo"]
        patron = analisis["patron_usado"]
        resultado["estado"] = "EXCEDENTE_CONFIRMADO"
        resultado["monto_liquidacion_saldo"] = str(monto)
        resultado["log_decision"] = (
            f"F3: Saldo ${monto:,} a favor del ejecutado "
            f"(folio {folio}, patron: '{patron}')"
        )
        log.info("  %s: -> EXCEDENTE_CONFIRMADO ($%s, patron: '%s')", tag, f"{monto:,}", patron)

    elif analisis.get("saldo_contra"):
        resultado["estado"] = "ELIMINAR"
        resultado["log_decision"] = (
            f"F3: Saldo en contra del ejecutado (folio {folio}). Sin excedente."
        )
        log.info("  %s: -> ELIMINAR (saldo en contra)", tag)

    elif analisis["sin_excedente"]:
        resultado["estado"] = "ELIMINAR"
        resultado["log_decision"] = (
            f"F3: Sin excedente en liquidacion (folio {folio})"
        )
        log.info("  %s: -> ELIMINAR (sin excedente)", tag)

    elif analisis.get("tipo_liquidacion") == "B" and analisis.get("credito_adeudado"):
        credito = analisis["credito_adeudado"]
        resultado["monto_credito_liquidado"] = str(credito)
        log.info("  %s: Liquidacion Tipo B (solo deuda). Credito adeudado: $%s",
                 tag, f"{credito:,}")

        # Si tiene monto_acta -> calcular delta
        monto_acta = _deuda_a_int(causa.get("monto_acta_remate", ""))
        if monto_acta > 0:
            delta = monto_acta - credito
            if delta > 0:
                resultado["estado"] = "EXCEDENTE_CONFIRMADO"
                resultado["monto_liquidacion_saldo"] = str(delta)
                resultado["log_decision"] = (
                    f"F3: Tipo B. Acta ${monto_acta:,} - Credito ${credito:,} = "
                    f"Excedente ${delta:,} (folio {folio})"
                )
                log.info("  %s: Tipo B + acta: $%s - $%s = $%s -> EXCEDENTE_CONFIRMADO",
                         tag, f"{monto_acta:,}", f"{credito:,}", f"{delta:,}")
            else:
                resultado["estado"] = "ELIMINAR"
                resultado["log_decision"] = (
                    f"F3: Tipo B. Acta ${monto_acta:,} - Credito ${credito:,} = "
                    f"${delta:,} (sin excedente, folio {folio})"
                )
                log.info("  %s: Tipo B + acta: delta $%s -> ELIMINAR",
                         tag, f"{delta:,}")
        else:
            resultado["estado"] = "PENDIENTE_REVISION_MANUAL"
            resultado["log_decision"] = (
                f"F3: Liquidacion Tipo B. Credito adeudado ${credito:,}. "
                f"Sin monto acta para comparar (folio {folio})"
            )
            log.info("  %s: Tipo B sin acta -> PENDIENTE_REVISION_MANUAL", tag)

    else:
        ruta_raw = analisis.get("ruta_raw", "")
        resultado["estado"] = "PENDIENTE_REVISION_MANUAL"
        resultado["log_decision"] = (
            f"F3: Liquidacion descargada pero saldo no extraido. "
            f"Ver raw: {ruta_raw}"
        )
        log.info("  %s: -> PENDIENTE_REVISION_MANUAL (saldo no extraido)", tag)

    return resultado



def filtro3_liquidacion(page, context, causa, filas_data):
    """Filtro 3 legacy: Busca documento de liquidacion (compatibilidad con flujo normal).

    Retorna (nuevo_estado, log_decision).
    """
    rol = causa.get("ROL", "")
    anio = causa.get("AÑO", "")

    for fila in filas_data:
        # Buscar solo "Liquidacion (Credito)" — el PDF con calculos reales
        texto_buscar = (fila.get("tramite", "") + " " + fila.get("desc_tramite", "")).lower()
        tiene_liq_pdf = any(kw in texto_buscar for kw in _KEYWORDS_LIQUIDACION_PDF)
        if not tiene_liq_pdf:
            continue

        log.info("  %s: Liquidacion encontrada en folio %d", rol, fila["folio"])

        nombre_pdf = f"{rol}-{anio}_liquidacion.pdf"
        ruta_destino = os.path.join(LIQUIDACIONES_DIR, nombre_pdf)

        try:
            ruta_pdf = _descargar_pdf_desde_fila(page, fila["element"], LIQUIDACIONES_DIR, nombre_pdf)
        except Exception as e:
            log.warning("Error descargando liquidacion: %s", e)
            ruta_pdf = None

        if ruta_pdf and os.path.exists(ruta_pdf):
            causa["ruta_liquidacion"] = ruta_pdf

            analisis = _analizar_pdf_liquidacion(ruta_pdf, causa)

            if analisis.get("saldo_contra"):
                return (
                    "ELIMINAR",
                    f"F3: Saldo en contra del ejecutado (folio {fila['folio']}). Sin excedente.",
                )
            elif analisis["sin_excedente"]:
                return (
                    "ELIMINAR",
                    f"F3: Sin excedente en liquidacion (folio {fila['folio']})",
                )
            elif analisis["saldo_encontrado"] and analisis["monto_saldo"]:
                monto = analisis["monto_saldo"]
                causa["monto_liquidacion_saldo"] = str(monto)
                return (
                    "EXCEDENTE_CONFIRMADO",
                    f"F3: Saldo ${monto:,} a favor del ejecutado "
                    f"(folio {fila['folio']}, patron: '{analisis['patron_usado']}')",
                )
            else:
                return (
                    "PENDIENTE_REVISION_MANUAL",
                    f"F3: Liquidacion descargada pero saldo no extraido. "
                    f"Ver raw: {analisis.get('ruta_raw', '')}",
                )
        else:
            return (
                "PENDIENTE_LIQUIDACION",
                f"EN ESPERA: Liquidacion detectada folio {fila['folio']}"
                " pero descarga fallo",
            )

    return (
        "PENDIENTE_LIQUIDACION",
        f"EN ESPERA: Liquidacion no encontrada aun"
        f" (ultima revision {date.today().isoformat()})",
    )


# =========================================================================
# FALLBACK: BUSQUEDA CON "TODOS" LOS TRIBUNALES
# =========================================================================

def _buscar_causa_fallback_todos(page, rol, anio, corte, causa):
    """
    Fallback cuando buscar_causa no encuentra la causa con el tribunal del Excel.
    Reintenta seleccionando 'Todos' (value='0') en #conTribunal.

    Si aparecen multiples resultados, busca por caratulado cruzando con
    demandante/demandado del Excel madre.

    Retorna True si encontro la causa, False si no.
    """
    try:
        # 1. Competencia = Civil
        page.select_option("#competencia", value="3")
        time.sleep(0.8)

        # 2. Corte (igual que buscar_causa)
        if not seleccionar_por_texto(page, "conCorte", corte, timeout_seg=10):
            log.warning("  Fallback: Corte no encontrada")
            return False
        time.sleep(1.0)

        # 3. Tribunal = "Todos" (value="0")
        page.select_option("#conTribunal", value="0")
        time.sleep(1.5)

        # 4. Libro = C
        page.select_option("#conTipoCausa", value="C")
        time.sleep(0.5)

        # 5. ROL + anio
        page.fill("#conRolCausa", "")
        page.fill("#conEraCausa", "")
        page.fill("#conRolCausa", rol)
        page.fill("#conEraCausa", anio)
        time.sleep(0.3)

        if page.input_value("#conRolCausa") != rol or page.input_value("#conEraCausa") != anio:
            page.fill("#conRolCausa", rol)
            page.fill("#conEraCausa", anio)
            time.sleep(0.5)

        # 6. Buscar
        page.click("#btnConConsulta")
        page.wait_for_load_state("domcontentloaded", timeout=20000)
        time.sleep(2.5)
        cerrar_popups(page)

        # Detectar "sin resultados"
        try:
            body_text = page.inner_text("body")
            if "No se han encontrado resultados" in body_text:
                log.info("  Fallback Todos: Sin resultados")
                return False
        except Exception:
            pass

        # Buscar filas de resultado
        filas = page.query_selector_all("table#veDetalle tbody tr")
        if not filas:
            filas = [
                f for f in page.query_selector_all("table tbody tr")
                if f.inner_text().strip()
                and "No se han encontrado" not in f.inner_text()
                and "VALOR RECUSACI" not in f.inner_text()
            ]

        if not filas:
            log.info("  Fallback Todos: Sin filas de resultado")
            return False

        if len(filas) == 1:
            log.info("  Fallback Todos: 1 resultado encontrado")
            return True

        # Multiples resultados: cruzar por caratulado
        demandante_excel = str(causa.get("DEMANDANTE", "")).upper().strip()
        demandado_excel = str(causa.get("DEMANDADO", "")).upper().strip()

        log.info("  Fallback Todos: %d resultados, buscando por caratulado"
                 " (dte='%s', ddo='%s')...",
                 len(filas), demandante_excel[:30], demandado_excel[:30])

        for fila in filas:
            try:
                texto_fila = fila.inner_text().upper().strip()

                # El caratulado OJV tiene formato "DEMANDANTE/DEMANDADO"
                # Verificar si alguna parte del demandante o demandado aparece
                match_dte = False
                match_ddo = False

                if demandante_excel:
                    # Tomar primera palabra significativa (>3 chars) del demandante
                    palabras_dte = [p for p in demandante_excel.split()
                                    if len(p) > 3 and p not in ("CHILE", "S.A.", "S.A", "LTDA")]
                    if palabras_dte:
                        match_dte = any(p in texto_fila for p in palabras_dte[:2])

                if demandado_excel:
                    palabras_ddo = [p for p in demandado_excel.split()
                                    if len(p) > 3 and p not in ("CHILE", "S.A.", "S.A", "LTDA")]
                    if palabras_ddo:
                        match_ddo = any(p in texto_fila for p in palabras_ddo[:2])

                if match_dte or match_ddo:
                    # Extraer tribunal real de la fila para logging
                    celdas = fila.query_selector_all("td")
                    tribunal_real = ""
                    for celda in celdas:
                        txt = celda.inner_text().strip()
                        if "juzgado" in txt.lower() or "tribunal" in txt.lower():
                            tribunal_real = txt
                            break

                    # Extraer caratulado para logging
                    caratulado = ""
                    for celda in celdas:
                        txt = celda.inner_text().strip()
                        if "/" in txt and len(txt) > 5:
                            caratulado = txt
                            break

                    log.info("  Fallback Todos: Encontrada via caratulado %s"
                             " (tribunal real: %s)", caratulado, tribunal_real)

                    # Click en la lupa de esta fila para seleccionarla
                    lupa = (
                        fila.query_selector("a[title*='Detalle']") or
                        fila.query_selector("a.fa-search") or
                        fila.query_selector("i.fa-search")
                    )
                    if lupa:
                        lupa.scroll_into_view_if_needed()
                        lupa.click()
                        page.wait_for_timeout(2000)
                    return True

            except Exception as e:
                log.warning("  Fallback Todos: Error revisando fila: %s", e)

        # Si no se pudo cruzar por caratulado, tomar la primera fila
        log.info("  Fallback Todos: No se pudo cruzar por caratulado,"
                 " usando primera fila")
        return True

    except Exception as e:
        log.warning("  Fallback Todos fallo: %s", e)
        return False


# =========================================================================
# NAVEGACION RAPIDA: reusar dropdowns ya seleccionados (Optimizacion 1)
# =========================================================================

def _seleccionar_corte_tribunal(page, corte, tribunal):
    """
    Selecciona competencia Civil + corte + tribunal + libro C en los dropdowns.
    Usado una vez por grupo de causas del mismo tribunal.
    """
    page.select_option("#competencia", value="3")
    time.sleep(0.8)

    if not seleccionar_por_texto(page, "conCorte", corte, timeout_seg=10):
        log.warning("  Corte no encontrada: %s", corte)
        return False
    time.sleep(1.0)

    if not seleccionar_por_texto(page, "conTribunal", tribunal, timeout_seg=15):
        log.warning("  Tribunal no encontrado: %s", tribunal)
        return False
    time.sleep(1.5)

    page.select_option("#conTipoCausa", value="C")
    time.sleep(0.5)
    return True


def _buscar_solo_rol(page, rol, anio):
    """
    Busca causa por ROL+anio asumiendo que los dropdowns ya estan seleccionados.
    Mucho mas rapido que buscar_causa() completa (~3s vs ~8s).
    """
    try:
        page.fill("#conRolCausa", "")
        page.fill("#conEraCausa", "")
        page.fill("#conRolCausa", rol)
        page.fill("#conEraCausa", anio)
        time.sleep(0.3)

        if page.input_value("#conRolCausa") != rol or page.input_value("#conEraCausa") != anio:
            page.fill("#conRolCausa", rol)
            page.fill("#conEraCausa", anio)
            time.sleep(0.5)

        page.click("#btnConConsulta")
        page.wait_for_load_state("domcontentloaded", timeout=20000)
        time.sleep(2.5)
        cerrar_popups(page)

        try:
            if "No se han encontrado resultados" in page.inner_text("body"):
                return False
        except Exception:
            pass

        filas = page.query_selector_all("table#veDetalle tbody tr")
        if not filas:
            filas = [
                f for f in page.query_selector_all("table tbody tr")
                if f.inner_text().strip()
                and "No se han encontrado" not in f.inner_text()
                and "VALOR RECUSACI" not in f.inner_text()
            ]
        return bool(filas)

    except Exception as e:
        log.warning("  _buscar_solo_rol fallo: %s", e)
        return False


# =========================================================================
# HTML AUDITORIA (Mejora 3 — guardar outerHTML para causas PENDIENTE_ACTA)
# =========================================================================

_HTML_AUDIT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "audit_html")


def _guardar_html_auditoria(page, causa, audit_dir, wtag=""):
    """Guarda outerHTML de #historiaCiv + encabezado del modal para revision offline."""
    try:
        tabla_html = page.evaluate('''() => {
            const tabla = document.getElementById("historiaCiv");
            return tabla ? tabla.outerHTML : "<p>Tabla no encontrada</p>";
        }''')

        header_html = page.evaluate('''() => {
            const modal = document.getElementById("modalDetalleCivil");
            if (!modal) return "";
            const panelBody = modal.querySelector(".panel-body");
            return panelBody ? panelBody.outerHTML : "";
        }''')

        rol = str(causa.get("ROL", "unknown")).replace("-", "_")
        anio = str(causa.get("AÑO", ""))
        tribunal_short = str(causa.get("TRIBUNAL", "unknown"))[:30].replace(" ", "_")
        filename = f"PENDIENTE_ACTA_C-{rol}-{anio}_{tribunal_short}.html"
        filepath = os.path.join(audit_dir, filename)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"<!-- ROL: C-{causa.get('ROL', '')}-{causa.get('AÑO', '')} -->\n")
            f.write(f"<!-- Tribunal: {causa.get('TRIBUNAL', '')} -->\n")
            f.write(f"<!-- Corte: {causa.get('CORTE', '')} -->\n")
            f.write(f"<!-- DTE: {causa.get('demandante', '')} -->\n")
            f.write(f"<!-- DDO: {causa.get('demandado', '')} -->\n")
            f.write(f"<!-- Cuaderno: Apremio -->\n")
            f.write(f"<!-- Fecha extraccion: {datetime.now().strftime('%Y-%m-%d %H:%M')} -->\n\n")
            f.write("<h3>Encabezado Causa</h3>\n")
            f.write(header_html + "\n\n")
            f.write("<h3>Tabla Historial Apremio (completa, sin filtro fecha)</h3>\n")
            f.write(tabla_html)

        log.info("  %sHTML guardado: %s", wtag, filename)
    except Exception as e:
        log.warning("  %sError guardando HTML auditoria: %s", wtag, e)


# =========================================================================
# PROCESAMIENTO OJV DE UNA CAUSA
# =========================================================================

def _procesar_causa_ojv(page, context, causa, solo_filtro1=False,
                        skip_nav=False, worker_id=0, audit_dir=None,
                        skip_pdf=False):
    """
    Navega a la causa en OJV, lee tabla, aplica filtros.
    Modifica el dict causa in-place.

    skip_nav=True: dropdowns ya estan seleccionados, solo buscar por ROL.
    worker_id: ID del worker para logging (0 = secuencial).
    audit_dir: ruta a carpeta para guardar HTMLs de auditoria (None = no guardar).

    Retorna True si la interaccion OJV fue exitosa.
    """
    _cerrar_modal_detalle(page)  # Limpiar modal residual de causa anterior

    rol = str(causa.get("ROL", ""))
    anio = str(causa.get("AÑO", ""))
    corte = str(causa.get("CORTE", ""))
    tribunal = str(causa.get("TRIBUNAL", ""))
    estado = causa.get("estado", "PENDIENTE_FILTRO1")

    wtag = f"W{worker_id} " if worker_id else ""

    # --- Cambio 7: Eliminacion manual pre-OJV ---
    for rol_manual, anio_manual, razon_manual in CAUSAS_ELIMINAR_MANUAL:
        if rol == rol_manual and anio == anio_manual:
            causa["estado"] = "ELIMINADA"
            causa["log_decision"] = f"Eliminacion manual: {razon_manual}"
            log.info("  %s%s-%s: Eliminacion manual: %s", wtag, rol, anio, razon_manual)
            return True

    hoy = date.today()
    fecha_remate = None
    fr_str = causa.get("fecha_remate", "")
    if fr_str:
        try:
            fecha_remate = date.fromisoformat(str(fr_str))
        except ValueError:
            pass

    if not fecha_remate:
        causa["log_decision"] = "ERROR: No se pudo parsear fecha_remate"
        return False

    dias = (hoy - fecha_remate).days

    if estado == "PENDIENTE_FILTRO1" and dias < 2:
        log.info("  %s%s: Muy pronto (%d dias), saltando", wtag, rol, dias)
        return True

    # --- Buscar causa en OJV ---
    log.info("  %sBuscando %s-%s en OJV...", wtag, rol, anio)

    found = False
    if skip_nav:
        # Optimizacion 1: dropdowns ya seleccionados, solo ROL
        found = _buscar_solo_rol(page, rol, anio)
    else:
        # Navegacion completa con reintentos
        for intento in range(1, 4):
            try:
                found = buscar_causa(page, rol, anio, corte, tribunal)
                if found:
                    break
            except Exception as e:
                log.warning("  %sIntento %d/3 buscar_causa fallo: %s",
                            wtag, intento, e)
                if intento < 3:
                    page.wait_for_timeout(2000)

    # Fallback con "Todos" los tribunales
    if not found:
        log.info("  %s%s: No encontrada en '%s'. Reintentando con Todos...",
                 wtag, rol, tribunal)
        try:
            limpiar_formulario(page)
        except Exception:
            pass
        found = _buscar_causa_fallback_todos(page, rol, anio, corte, causa)

    if not found:
        causa["log_decision"] = "ERROR OJV: Causa no encontrada ni con Todos los tribunales"
        causa["notas"] = "ERROR_OJV"
        log.info("  %s%s: No encontrada en OJV (ni con fallback Todos)", wtag, rol)
        return False

    try:
        detalle = abrir_detalle(page, rol, anio)
    except Exception as e:
        log.warning("  %sError abriendo detalle: %s", wtag, e)
        _cerrar_modal_detalle(page)  # Limpiar modal que pudo quedar abierto
        causa["log_decision"] = f"ERROR OJV: abrir_detalle fallo - {e}"
        causa["notas"] = "ERROR_OJV"
        return False

    if not detalle:
        causa["log_decision"] = "ERROR OJV: No se pudo abrir detalle de causa"
        causa["notas"] = "ERROR_OJV"
        return False

    causa["fecha_ultimo_check"] = hoy.isoformat()

    # --- Cambio 2: Detectar causa "Archivada" en OJV ---
    try:
        estado_span = page.query_selector(
            "#modalDetalleCivil table tbody tr:nth-child(2) td:first-child span.topTool"
        )
        if estado_span:
            estado_texto = estado_span.inner_text().strip().lower()
            if "archivada" in estado_texto:
                log.info("  %s%s: Estado OJV = ARCHIVADA -> PENDIENTE_REVISION_MANUAL",
                         wtag, rol)
                causa["estado"] = "PENDIENTE_REVISION_MANUAL"
                causa["detalle_auditoria"] = (
                    "Causa archivada en OJV. Posible causa acumulada. "
                    "Requiere revision manual."
                )
                _cerrar_modal_detalle(page)
                return True
    except Exception as e:
        log.warning("  %s%s: Error verificando estado procesal: %s", wtag, rol, e)

    # =====================================================================
    # FILTRO 1: Lectura de tabla #historiaCiv (sin PDFs)
    # =====================================================================
    if estado == "PENDIENTE_FILTRO1" and dias >= 2:
        try:
            tabla_historial, total_filas, fecha_limite = _leer_tabla_historial(
                page, fecha_remate
            )

            if not tabla_historial:
                if total_filas > 0:
                    # Hay filas pero todas son anteriores al filtro
                    msg = (
                        f"Tabla historial: 0 filas recientes ({total_filas} totales, "
                        f"filtro >= {fecha_limite.isoformat()}). "
                        "Sin actividad reciente en Apremio."
                    )
                    log.info("  %s%s: %s", wtag, rol, msg)
                    causa["detalle_auditoria"] = msg

                    # --- Cambio 3: Sin movimiento 14+ dias -> PENDIENTE_REVISION_MANUAL ---
                    # Leer fecha del movimiento mas reciente de la tabla completa
                    try:
                        primera_fecha_str = page.evaluate('''() => {
                            const filas = document.querySelectorAll("#historiaCiv table tbody tr");
                            if (filas.length === 0) return null;
                            const celdas = filas[0].querySelectorAll("td");
                            return celdas.length > 6 ? celdas[6].innerText.trim() : null;
                        }''')
                        if primera_fecha_str:
                            primera_fecha = _parsear_fecha_dd_mm_yyyy(primera_fecha_str)
                            if primera_fecha:
                                dias_sin_mov = (hoy - primera_fecha).days
                                if dias_sin_mov > 14:
                                    log.info("  %s%s: Sin movimiento en %d dias -> PENDIENTE_REVISION_MANUAL",
                                             wtag, rol, dias_sin_mov)
                                    causa["estado"] = "PENDIENTE_REVISION_MANUAL"
                                    causa["detalle_auditoria"] = (
                                        f"Sin movimiento en Apremio por {dias_sin_mov} dias. "
                                        f"Ultimo mov: {primera_fecha_str}. Posible causa acumulada."
                                    )
                                    causa["log_decision"] = f"Sin movimiento {dias_sin_mov}d -> revision manual"
                                    _cerrar_modal_detalle(page)
                                    return True
                    except Exception as e:
                        log.warning("  %s%s: Error leyendo fecha ultimo movimiento: %s",
                                    wtag, rol, e)

                else:
                    log.warning("  %s%s: Tabla historial vacia (0 filas totales)",
                                wtag, rol)
                causa["log_decision"] = "WARNING: Tabla #historiaCiv vacia"
            else:
                result_f1 = _aplicar_filtro1(
                    tabla_historial, causa, modo_auditoria=False
                )
                resultado_f1 = result_f1[0]
                detalle_f1 = result_f1[1]
                folio_acta = result_f1[2] if len(result_f1) > 2 else None

                log.info("  [FILTRO1%s] %s (%s) -> %s: %s",
                         f"-W{worker_id}" if worker_id else "",
                         rol, tribunal, resultado_f1, detalle_f1)

                if resultado_f1 == "ELIMINAR":
                    causa["estado"] = "ELIMINADA"
                    causa["log_decision"] = detalle_f1
                elif resultado_f1 == "PENDIENTE_REVISION_MANUAL":
                    causa["estado"] = "PENDIENTE_REVISION_MANUAL"
                    causa["log_decision"] = detalle_f1
                elif resultado_f1 == "NECESITA_PDF_ACTA":
                    # Acta detectada en tabla -> intentar F2 hibrido con PDF
                    causa["estado"] = "PENDIENTE_ACTA"
                    causa["log_decision"] = detalle_f1
                    causa["_folio_acta"] = folio_acta
                    estado = "PENDIENTE_ACTA"
                else:
                    causa["estado"] = resultado_f1
                    causa["log_decision"] = detalle_f1
                    estado = resultado_f1

                # Guardar senales informativas en detalle_auditoria
                info_items = _recopilar_info_tabla(tabla_historial)
                if info_items:
                    causa["detalle_auditoria"] = "; ".join(info_items)

                # Mejora 3: guardar outerHTML si quedo como PENDIENTE_ACTA
                if causa.get("estado") == "PENDIENTE_ACTA" and audit_dir:
                    _guardar_html_auditoria(page, causa, audit_dir, wtag)
        finally:
            _cerrar_modal_detalle(page)  # SIEMPRE cerrar despues de leer tabla

    # Si --solo-filtro1, no ejecutar F2/F3
    if solo_filtro1:
        _cerrar_modal(page)
        return True

    # =====================================================================
    # FILTRO 2 HIBRIDO: Descarga PDF del acta si _folio_acta disponible
    # =====================================================================
    folio_acta = causa.pop("_folio_acta", None)
    if folio_acta and estado == "PENDIENTE_ACTA" and not solo_filtro1 and skip_pdf:
        # --skip-pdf: mapear directo a PENDIENTE_LIQUIDACION sin descargar
        causa["estado"] = "PENDIENTE_LIQUIDACION"
        causa["log_decision"] = (
            f"F2: Acta detectada (folio {folio_acta}) pero --skip-pdf activo"
            f" -> PENDIENTE_LIQUIDACION"
        )
        estado = "PENDIENTE_LIQUIDACION"
        log.info("  %s%s: --skip-pdf, acta folio %s -> PENDIENTE_LIQUIDACION",
                 wtag, rol, folio_acta)
        folio_acta = None  # Evitar que entre al bloque de descarga

    if folio_acta and estado == "PENDIENTE_ACTA" and not solo_filtro1:
        log.info("  %s%s: F2 hibrido - descargando PDF acta folio %s...",
                 wtag, rol, folio_acta)
        os.makedirs(ACTAS_DIR, exist_ok=True)
        pdf_path = _descargar_pdf_acta_por_folio(
            page, context, folio_acta, causa, ACTAS_DIR
        )

        if pdf_path:
            analisis = _analizar_pdf_acta(pdf_path)
            deuda = _deuda_a_int(causa.get("MONTO_DEUDA_CLP", "0"))
            log.info("  %s%s: PDF analizado: cargo=%s, monto=%s, texto=%s",
                     wtag, rol, analisis["cargo_al_credito"],
                     analisis["monto_adjudicacion"], analisis.get("texto_monto"))

            if analisis["cargo_al_credito"]:
                causa["estado"] = "ELIMINADA"
                causa["log_decision"] = (
                    f"Filtro 2: Adjudicado con cargo al credito"
                    f" (folio {folio_acta}). Sin excedente posible."
                )
                log.info("  %s%s: Cargo al credito -> ELIMINAR", wtag, rol)
            elif analisis["monto_adjudicacion"] is not None:
                monto = analisis["monto_adjudicacion"]
                causa["monto_acta_remate"] = str(monto)
                if deuda > 0 and monto < deuda:
                    causa["estado"] = "ELIMINADA"
                    causa["log_decision"] = (
                        f"Filtro 2: Monto adjudicacion ${monto:,}"
                        f" < deuda ${deuda:,} (folio {folio_acta}). Sin excedente."
                    )
                    log.info("  %s%s: Monto %s < deuda %s -> ELIMINAR",
                             wtag, rol, f"${monto:,}", f"${deuda:,}")
                else:
                    causa["estado"] = "PENDIENTE_LIQUIDACION"
                    causa["log_decision"] = (
                        f"Filtro 2: Monto adjudicacion ${monto:,}"
                        f" >= deuda ${deuda:,} (folio {folio_acta}). Posible excedente."
                    )
                    estado = "PENDIENTE_LIQUIDACION"
                    log.info("  %s%s: Monto %s >= deuda %s -> PENDIENTE_LIQUIDACION",
                             wtag, rol, f"${monto:,}", f"${deuda:,}")
            else:
                causa["estado"] = "PENDIENTE_LIQUIDACION"
                causa["log_decision"] = (
                    f"Filtro 2: Acta encontrada pero no se pudo extraer monto"
                    f" (folio {folio_acta}). Revisar PDF manualmente."
                )
                estado = "PENDIENTE_LIQUIDACION"
                log.info("  %s%s: Acta sin monto extraible -> PENDIENTE_LIQUIDACION",
                         wtag, rol)
        else:
            causa["log_decision"] = (
                f"Filtro 2: Acta detectada en tabla pero PDF no descargable"
                f" (folio {folio_acta})."
            )
            log.info("  %s%s: PDF acta no descargable -> PENDIENTE_ACTA", wtag, rol)

    # =====================================================================
    # FILTRO 2 LEGACY y FILTRO 3: Requieren cuaderno Apremio + descarga PDFs
    # =====================================================================
    if estado in ("PENDIENTE_ACTA", "PENDIENTE_LIQUIDACION") and not solo_filtro1:
        if dias > 30 and estado == "PENDIENTE_ACTA":
            causa["estado"] = "PENDIENTE_LIQUIDACION"
            causa["log_decision"] = (
                "Causa antigua (>30 dias), saltando directo a buscar liquidacion"
            )
            estado = "PENDIENTE_LIQUIDACION"

        # Hasta 3 intentos. El WAF del PJUD a veces devuelve la pagina parcial
        # (selector existe pero tbody vacio); re-seleccionar el cuaderno fuerza
        # un nuevo fetch del lado server.
        filas_elements = []
        for intento in range(1, 4):
            try:
                seleccionar_cuaderno(page, "Apremio")
            except Exception as e:
                log.warning("  %sError seleccionando cuaderno Apremio (intento %d/3): %s", wtag, intento, e)
            # Backoff progresivo: 1.5s, 3s, 6s
            if intento == 1:
                page.wait_for_timeout(1500)
            elif intento == 2:
                page.wait_for_timeout(3000)
            else:
                page.wait_for_timeout(6000)
            try:
                filas_elements = filas_del_modal(page)
            except Exception as e:
                log.warning("  %sError obteniendo filas (intento %d/3): %s", wtag, intento, e)
                filas_elements = []
            if filas_elements:
                if intento > 1:
                    log.info("  %s%s: tabla recuperada en intento %d", wtag, rol, intento)
                break

        if not filas_elements:
            causa["log_decision"] = "ERROR OJV: Tabla de tramitacion vacia (cuaderno) tras 3 intentos"
            _cerrar_modal(page)
            return False

        filas_data = _parsear_filas_tabla(page, filas_elements)
        log.info("  %s%s: %d filas en tabla de tramitacion", wtag, rol, len(filas_data))

        # Solo usar legacy F2 si el hibrido no resolvio (sin _folio_acta)
        if estado == "PENDIENTE_ACTA" and dias >= 3 and not folio_acta:
            new_estado, log_dec = filtro2_acta_remate(
                page, context, causa, filas_data
            )
            if new_estado is None:
                causa["estado"] = "ELIMINADA"
                causa["log_decision"] = log_dec
            else:
                causa["estado"] = new_estado
                causa["log_decision"] = log_dec
                estado = new_estado

        if estado == "PENDIENTE_LIQUIDACION" and dias >= 14:
            new_estado, log_dec = filtro3_liquidacion(
                page, context, causa, filas_data
            )
            causa["estado"] = new_estado
            causa["log_decision"] = log_dec

    _cerrar_modal(page)
    return True


# =========================================================================
# REAUDIT: reprocesar causas PENDIENTE_ACTA con tabla completa
# =========================================================================

def _procesar_causa_reaudit(page, context, causa, worker_id=0, audit_dir=None,
                            skip_pdf=False):
    """
    Modo --reaudit: abre causa en OJV, lee tabla completa (sin filtro fecha),
    guarda outerHTML, re-ejecuta Filtro 1 sobre TODAS las filas.
    """
    _cerrar_modal_detalle(page)

    rol = str(causa.get("ROL", ""))
    anio = str(causa.get("AÑO", ""))
    corte = str(causa.get("CORTE", ""))
    tribunal = str(causa.get("TRIBUNAL", ""))
    wtag = f"W{worker_id} " if worker_id else ""

    hoy = date.today()

    # --- Eliminacion manual pre-OJV ---
    for rol_manual, anio_manual, razon_manual in CAUSAS_ELIMINAR_MANUAL:
        if rol == rol_manual and anio == anio_manual:
            causa["estado"] = "ELIMINADA"
            causa["log_decision"] = f"Eliminacion manual: {razon_manual}"
            log.info("  %s%s-%s: Eliminacion manual (reaudit): %s", wtag, rol, anio, razon_manual)
            return True

    log.info("  %s[REAUDIT] Buscando %s-%s en OJV...", wtag, rol, anio)

    # Buscar causa
    found = False
    for intento in range(1, 4):
        try:
            found = buscar_causa(page, rol, anio, corte, tribunal)
            if found:
                break
        except Exception as e:
            log.warning("  %sIntento %d/3 buscar_causa fallo: %s", wtag, intento, e)
            if intento < 3:
                page.wait_for_timeout(2000)

    if not found:
        try:
            limpiar_formulario(page)
        except Exception:
            pass
        found = _buscar_causa_fallback_todos(page, rol, anio, corte, causa)

    if not found:
        causa["log_decision"] = "REAUDIT ERROR: Causa no encontrada en OJV"
        causa["notas"] = "ERROR_OJV"
        log.info("  %s%s: No encontrada en OJV (reaudit)", wtag, rol)
        return False

    try:
        detalle = abrir_detalle(page, rol, anio)
    except Exception as e:
        log.warning("  %sError abriendo detalle (reaudit): %s", wtag, e)
        _cerrar_modal_detalle(page)
        causa["log_decision"] = f"REAUDIT ERROR: abrir_detalle fallo - {e}"
        causa["notas"] = "ERROR_OJV"
        return False

    if not detalle:
        causa["log_decision"] = "REAUDIT ERROR: No se pudo abrir detalle"
        causa["notas"] = "ERROR_OJV"
        return False

    causa["fecha_ultimo_check"] = hoy.isoformat()

    # Detectar causa Archivada
    try:
        estado_span = page.query_selector(
            "#modalDetalleCivil table tbody tr:nth-child(2) td:first-child span.topTool"
        )
        if estado_span:
            estado_texto = estado_span.inner_text().strip().lower()
            if "archivada" in estado_texto:
                log.info("  %s%s: Estado OJV = ARCHIVADA -> PENDIENTE_REVISION_MANUAL (reaudit)",
                         wtag, rol)
                causa["estado"] = "PENDIENTE_REVISION_MANUAL"
                causa["detalle_auditoria"] = (
                    "Causa archivada en OJV. Posible causa acumulada. "
                    "Requiere revision manual."
                )
                _cerrar_modal_detalle(page)
                return True
    except Exception:
        pass

    try:
        # Leer tabla completa (sin filtro fecha)
        tabla_completa = _leer_tabla_historial_completa(page)

        # Guardar outerHTML para auditoria
        if audit_dir:
            _guardar_html_auditoria(page, causa, audit_dir, wtag)

        if not tabla_completa:
            causa["log_decision"] = "REAUDIT: Tabla historial vacia"
            log.info("  %s%s: Tabla historial vacia (reaudit)", wtag, rol)
        else:
            # Evaluar filtros sobre TODAS las filas (sin filtro fecha)
            estado_eval, folio_eval, fecha_eval, detalle_eval = _evaluar_filtros_tabla(
                tabla_completa
            )
            # Recopilar senales informativas
            info_items = _recopilar_info_tabla(tabla_completa)
            info_str = "; ".join(info_items) if info_items else ""

            if estado_eval is None:
                # Sin senales -> mantener como PENDIENTE_ACTA
                causa["log_decision"] = f"REAUDIT: Sin senales nuevas"
                if info_str:
                    causa["detalle_auditoria"] = info_str
                log.info("  %s[REAUDIT] %s -> sin senales", wtag, rol)
            else:
                log.info("  %s[REAUDIT] %s -> %s: %s", wtag, rol, estado_eval, detalle_eval)

                if estado_eval == "ELIMINAR":
                    causa["estado"] = "ELIMINADA"
                    causa["log_decision"] = f"REAUDIT: {detalle_eval}"
                elif estado_eval == "PENDIENTE_LIQUIDACION":
                    causa["estado"] = "PENDIENTE_LIQUIDACION"
                    causa["log_decision"] = f"REAUDIT: {detalle_eval}"
                elif estado_eval == "NECESITA_PDF_ACTA":
                    if skip_pdf:
                        causa["estado"] = "PENDIENTE_LIQUIDACION"
                        causa["log_decision"] = f"REAUDIT: {detalle_eval} (--skip-pdf -> PENDIENTE_LIQUIDACION)"
                    else:
                        # Intentar descarga y analisis del PDF del acta
                        log.info("  %s%s: Acta detectada folio %s. Descargando PDF...",
                                 wtag, rol, folio_eval)
                        os.makedirs(ACTAS_DIR, exist_ok=True)
                        pdf_path = _descargar_pdf_acta_por_folio(
                            page, context, folio_eval, causa, ACTAS_DIR
                        )
                        if pdf_path:
                            analisis = _analizar_pdf_acta(pdf_path)
                            deuda = _deuda_a_int(causa.get("MONTO_DEUDA_CLP", "0"))
                            log.info("  %s%s: PDF analizado: cargo=%s, monto=%s, texto=%s",
                                     wtag, rol, analisis["cargo_al_credito"],
                                     analisis["monto_adjudicacion"],
                                     analisis.get("texto_monto"))

                            if analisis["cargo_al_credito"]:
                                causa["estado"] = "ELIMINADA"
                                causa["log_decision"] = (
                                    f"REAUDIT F2: Cargo al credito (folio {folio_eval}). "
                                    "Sin excedente."
                                )
                            else:
                                monto_adj = analisis["monto_adjudicacion"] or 0
                                if monto_adj > 0:
                                    causa["monto_acta_remate"] = str(monto_adj)
                                if monto_adj > 0 and deuda > 0 and monto_adj < deuda:
                                    causa["estado"] = "ELIMINADA"
                                    causa["log_decision"] = (
                                        f"REAUDIT F2: Adjudicacion ${monto_adj:,}"
                                        f" < deuda ${deuda:,} (folio {folio_eval})."
                                    )
                                elif monto_adj > 0 and deuda > 0:
                                    causa["estado"] = "PENDIENTE_LIQUIDACION"
                                    causa["log_decision"] = (
                                        f"REAUDIT F2: Adjudicacion ${monto_adj:,}"
                                        f" >= deuda ${deuda:,} (folio {folio_eval})."
                                        " Posible excedente."
                                    )
                                else:
                                    causa["estado"] = "PENDIENTE_LIQUIDACION"
                                    causa["log_decision"] = (
                                        f"REAUDIT F2: Acta descargada, monto_adj=${monto_adj:,},"
                                        f" deuda=${deuda:,} (folio {folio_eval})."
                                        " Comparacion incompleta."
                                    )
                        else:
                            causa["estado"] = "PENDIENTE_LIQUIDACION"
                            causa["log_decision"] = (
                                f"REAUDIT F2: Acta en tabla pero PDF no descargable"
                                f" (folio {folio_eval})."
                            )
                elif estado_eval == "PENDIENTE_REVISION_MANUAL":
                    causa["estado"] = "PENDIENTE_REVISION_MANUAL"
                    causa["log_decision"] = f"REAUDIT: {detalle_eval}"

                if info_str:
                    causa["detalle_auditoria"] = info_str
    finally:
        _cerrar_modal_detalle(page)

    return True


def _procesar_chunk_reaudit(page, context, causas_con_idx, worker_id, audit_dir=None,
                            skip_pdf=False):
    """Procesa un chunk de causas en modo reaudit."""
    t0 = time.time()
    resultados = {}

    for i, (idx, causa) in enumerate(causas_con_idx, 1):
        rol = str(causa.get("ROL", ""))
        anio = str(causa.get("AÑO", ""))
        wtag = f"W{worker_id}" if worker_id else ""

        log.info("  [%s] REAUDIT %d/%d: %s-%s", wtag, i, len(causas_con_idx), rol, anio)

        try:
            _procesar_causa_reaudit(
                page, context, causa,
                worker_id=worker_id,
                audit_dir=audit_dir,
                skip_pdf=skip_pdf,
            )
        except Exception as e:
            log.error("[%s] Error reaudit causa %s: %s", wtag, rol, e)
            causa["log_decision"] = f"REAUDIT ERROR: {e}"
            causa["notas"] = "ERROR_OJV"

        resultados[idx] = causa
        time.sleep(2.0)

    elapsed = time.time() - t0
    log.info("  [W%d] Reaudit completado: %d causas en %d min %d seg",
             worker_id, len(causas_con_idx),
             int(elapsed) // 60, int(elapsed) % 60)
    return resultados


# =========================================================================
# PROCESAMIENTO POR CHUNKS (Optimizacion 1 + 2)
# =========================================================================

def _procesar_chunk(page, context, causas_con_idx, worker_id,
                    solo_filtro1, audit_dir=None, skip_pdf=False):
    """
    Procesa un chunk de causas con agrupacion por corte+tribunal.
    Cada worker ejecuta esta funcion en su propia page.

    causas_con_idx: lista de (df_index, causa_dict)
    audit_dir: ruta a carpeta para HTMLs de auditoria (None = no guardar)
    Retorna: dict {df_index: causa_dict_modificado}
    """
    t0_worker = time.time()
    resultados = {}
    ultimo_corte = None
    ultimo_tribunal = None
    nav_ok = False  # True si los dropdowns estan correctamente seleccionados

    for i, (idx, causa) in enumerate(causas_con_idx, 1):
        corte = str(causa.get("CORTE", ""))
        tribunal = str(causa.get("TRIBUNAL", ""))
        rol = str(causa.get("ROL", ""))
        anio = str(causa.get("AÑO", ""))
        wtag = f"W{worker_id}" if worker_id else ""

        mismo_tribunal = (corte == ultimo_corte and tribunal == ultimo_tribunal
                          and nav_ok)

        if not mismo_tribunal:
            # Navegacion completa: seleccionar corte + tribunal
            try:
                limpiar_formulario(page)
            except Exception:
                pass

            log.info("  [%s] Grupo: %s / %s", wtag, corte, tribunal)

            nav_ok = _seleccionar_corte_tribunal(page, corte, tribunal)
            ultimo_corte = corte
            ultimo_tribunal = tribunal

            if not nav_ok:
                log.warning("  [%s] No se pudo seleccionar corte/tribunal, "
                            "usando buscar_causa completa", wtag)

        log.info("  [%s] Causa %d/%d: %s-%s [%s]%s",
                 wtag, i, len(causas_con_idx), rol, anio,
                 causa.get("estado", "?"),
                 " (mismo tribunal)" if mismo_tribunal else "")

        try:
            _procesar_causa_ojv(
                page, context, causa,
                solo_filtro1=solo_filtro1,
                skip_nav=(mismo_tribunal and nav_ok),
                worker_id=worker_id,
                audit_dir=audit_dir,
                skip_pdf=skip_pdf,
            )
        except Exception as e:
            log.error("[%s] Error procesando causa %s: %s", wtag, rol, e)
            causa["log_decision"] = f"ERROR: excepcion no capturada - {e}"
            causa["notas"] = "ERROR_OJV"
            # Si fallo, resetear nav para forzar navegacion completa
            nav_ok = False

        # Si _cerrar_modal_detalle hizo reload, resetear nav
        # (el modal ya fue cerrado pero los dropdowns se perdieron)
        try:
            modal_check = page.query_selector("#conCorte")
            if modal_check and not page.query_selector("#modalDetalleCivil.in"):
                # Verificar si los dropdowns siguen seleccionados
                corte_actual = page.evaluate('document.getElementById("conCorte").value')
                if not corte_actual:
                    nav_ok = False
        except Exception:
            nav_ok = False

        resultados[idx] = causa

        # Pausa entre causas: 1s mismo tribunal, 2s diferente
        if mismo_tribunal:
            time.sleep(1.0)
        else:
            time.sleep(2.0)

    elapsed = time.time() - t0_worker
    log.info("  [W%d] Chunk completado: %d causas en %d min %d seg",
             worker_id, len(causas_con_idx),
             int(elapsed) // 60, int(elapsed) % 60)

    return resultados


# =========================================================================
# PARALELIZACION: subprocess (cada worker = proceso Python independiente)
# =========================================================================

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_TEMP_WORKERS_DIR = os.path.join(_BASE_DIR, "temp_workers")


def _serializar_causas(causas_con_idx):
    """Convierte lista de (idx, causa_dict) a JSON-serializable."""
    resultado = []
    for idx, causa in causas_con_idx:
        c = {"_idx": idx}
        for k, v in causa.items():
            if hasattr(v, "isoformat"):
                c[k] = v.isoformat()
            elif isinstance(v, float) and (v != v):  # NaN
                c[k] = ""
            else:
                c[k] = v
        resultado.append(c)
    return resultado


def _deserializar_causas(lista):
    """Convierte lista de dicts JSON a lista de (idx, causa_dict)."""
    resultado = []
    for c in lista:
        idx = c.pop("_idx")
        resultado.append((idx, c))
    return resultado


# Asignacion fija de Cortes por Worker
# Balanceado para ~270 causas tipicas (proyeccion estable)
WORKER_CORTES = {
    1: ["C.A. de La Serena"],
    2: ["C.A. de Valparaíso", "C.A. de Punta Arenas", "C.A. de Talca"],
    3: ["C.A. de Concepción", "C.A. de Coyhaique", "C.A. de Chillán"],
    4: ["C.A. de Antofagasta", "C.A. de Rancagua"],
    5: ["C.A. de Iquique", "C.A. de Temuco", "C.A. de Copiapó", "C.A. de Puerto Montt"],
}


def _ejecutar_paralelo(causas_para_procesar, n_workers, solo_filtro1, primera_run,
                       reaudit=False, skip_pdf=False):
    """
    Lanza N subprocesos Python, cada uno con su propio Playwright browser.
    Distribuye causas entre workers por Corte de Apelaciones.
    """
    os.makedirs(_TEMP_WORKERS_DIR, exist_ok=True)

    # Limitar al maximo configurado
    n_workers = min(n_workers, len(WORKER_CORTES))

    # Agrupar causas por worker segun su corte
    worker_causas = {w: [] for w in range(1, n_workers + 1)}

    # Construir mapa inverso corte -> worker_id
    corte_to_worker = {}
    for w_id, cortes_list in WORKER_CORTES.items():
        if w_id > n_workers:
            break
        for corte in cortes_list:
            corte_to_worker[corte] = w_id

    for idx, causa in causas_para_procesar:
        corte = str(causa.get("CORTE", ""))
        w_id = corte_to_worker.get(corte)
        if w_id and w_id <= n_workers:
            worker_causas[w_id].append((idx, causa))
        else:
            # Causas con corte no mapeada o None -> worker 4 (o ultimo disponible)
            target = min(n_workers, 4)
            worker_causas[target].append((idx, causa))

    # No lanzar workers vacios
    workers_activos = {w: c for w, c in worker_causas.items() if len(c) > 0}
    log.info("Lanzando %d workers activos de %d configurados", len(workers_activos), n_workers)

    for w_id in sorted(workers_activos):
        log.info("  Worker %d: %d causas", w_id, len(workers_activos[w_id]))

    procesos = []
    for worker_id, chunk in sorted(workers_activos.items()):
        chunk_file = os.path.join(_TEMP_WORKERS_DIR, f"chunk_{worker_id}.json")
        result_file = os.path.join(_TEMP_WORKERS_DIR, f"result_{worker_id}.json")

        # Serializar chunk a JSON
        with open(chunk_file, "w", encoding="utf-8") as f:
            json.dump(_serializar_causas(chunk), f, ensure_ascii=False)

        # Construir comando
        cmd = [
            sys.executable, os.path.join(_BASE_DIR, "filtrador_saldos.py"),
            "--worker-mode",
            "--chunk-file", chunk_file,
            "--result-file", result_file,
            "--worker-id", str(worker_id),
        ]
        if primera_run:
            cmd.append("--primera-run")
        if solo_filtro1:
            cmd.append("--solo-filtro1")
        if reaudit:
            cmd.append("--reaudit")
        if skip_pdf:
            cmd.append("--skip-pdf")
        cmd.extend(["--audit-dir", _HTML_AUDIT_DIR])

        # Cada worker escribe a su propio archivo, que el orquestador anexa
        # al log central tras terminar. Asi queda trazabilidad de errores
        # de workers en el archivo de log principal.
        worker_log_path = os.path.join(
            LOGS_LIQUI_DIR,
            f"worker_{worker_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        )
        worker_log_handle = open(worker_log_path, 'w', encoding='utf-8', errors='replace')
        log.info("Lanzando Worker %d (%d causas)...", worker_id, len(chunk))
        # PYTHONIOENCODING=utf-8 fuerza al subprocess a imprimir en utf-8.
        # Sin esto, Windows usa cp1252 para sys.stdout cuando el stdout del
        # subprocess se redirige a archivo, lo que rompe los print() con
        # caracteres como '✓' o '✗' presentes en ojv_remates.py.
        worker_env = os.environ.copy()
        worker_env['PYTHONIOENCODING'] = 'utf-8'
        proc = subprocess.Popen(
            cmd,
            cwd=_BASE_DIR,
            stdout=worker_log_handle,
            stderr=subprocess.STDOUT,
            env=worker_env,
        )
        procesos.append((proc, worker_id, result_file, chunk_file, worker_log_handle, worker_log_path))

    # Esperar a que todos terminen
    for proc, wid, _, _, wlog_handle, wlog_path in procesos:
        proc.wait()
        wlog_handle.close()
        log.info("Worker %d terminado (exit code: %d)", wid, proc.returncode)
        # Anexar el log del worker al log central
        try:
            with open(wlog_path, encoding='utf-8', errors='replace') as f:
                for linea in f:
                    linea = linea.rstrip()
                    if linea:
                        # Encoding ASCII para Windows cp1252
                        safe = linea.encode("ascii", errors="replace").decode("ascii")
                        log.info("  [W%d] %s", wid, safe[:300])
        except Exception as e:
            log.warning("No se pudo anexar log de worker %d: %s", wid, e)

    # Leer resultados de cada worker
    all_resultados = {}
    for _, wid, result_file, chunk_file, _, _ in procesos:
        if os.path.exists(result_file):
            with open(result_file, "r", encoding="utf-8") as f:
                resultados_json = json.load(f)
            causas_resultado = _deserializar_causas(resultados_json)
            for idx, causa in causas_resultado:
                all_resultados[idx] = causa
            log.info("Worker %d: %d resultados leidos", wid, len(causas_resultado))
        else:
            log.error("Worker %d: ERROR - no se genero archivo de resultados", wid)

        # Limpiar temporales
        for tmp in [result_file, chunk_file]:
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except OSError:
                    pass

    # Limpiar directorio temporal si esta vacio
    try:
        os.rmdir(_TEMP_WORKERS_DIR)
    except OSError:
        pass

    return all_resultados


def _run_worker_mode(args):
    """
    Modo worker: proceso hijo que procesa un chunk de causas.
    NO hace merge, NO toca el Excel. Solo OJV + filtros.
    Escribe resultados a JSON y sale.
    """
    worker_id = args.worker_id
    solo_filtro1 = args.solo_filtro1

    audit_dir = getattr(args, "audit_dir", None)

    log.info("[W%d] Iniciando worker (PID %d)...", worker_id, os.getpid())

    # Leer chunk
    with open(args.chunk_file, "r", encoding="utf-8") as f:
        causas_json = json.load(f)
    causas_con_idx = _deserializar_causas(causas_json)

    log.info("[W%d] %d causas a procesar", worker_id, len(causas_con_idx))

    # Crear carpeta audit_html si se paso la ruta
    if audit_dir:
        os.makedirs(audit_dir, exist_ok=True)

    # Abrir SU PROPIO Playwright browser (perfil por worker para no colisionar)
    with sync_playwright() as p:
        _profile_dir = os.path.join(REMATES_DIR, f".chrome-profile-w{worker_id}")
        context = p.chromium.launch_persistent_context(
            _profile_dir,
            headless=False,
            slow_mo=100,
            channel="chrome",
            args=["--disable-blink-features=AutomationControlled"],
            accept_downloads=True,
        )
        page = context.pages[0] if context.pages else context.new_page()
        Stealth().apply_stealth_sync(page)
        page.set_default_timeout(15000)

        if not navegar_a_consulta(page):
            log.error("[W%d] No se pudo navegar a OJV. Abortando worker.", worker_id)
            context.close()
            # Escribir resultados vacios para que el orquestador no falle
            with open(args.result_file, "w", encoding="utf-8") as f:
                json.dump(_serializar_causas(causas_con_idx), f, ensure_ascii=False)
            return

        reaudit = getattr(args, "reaudit", False)
        skip_pdf = getattr(args, "skip_pdf", False)

        if reaudit:
            resultados = _procesar_chunk_reaudit(
                page, context, causas_con_idx,
                worker_id=worker_id,
                audit_dir=audit_dir,
                skip_pdf=skip_pdf,
            )
        else:
            resultados = _procesar_chunk(
                page, context, causas_con_idx,
                worker_id=worker_id,
                solo_filtro1=solo_filtro1,
                audit_dir=audit_dir,
                skip_pdf=skip_pdf,
            )

        context.close()

    # Serializar resultados a JSON
    resultados_lista = [(idx, causa) for idx, causa in resultados.items()]
    with open(args.result_file, "w", encoding="utf-8") as f:
        json.dump(_serializar_causas(resultados_lista), f, ensure_ascii=False)

    log.info("[W%d] Worker finalizado. Resultados escritos en %s",
             worker_id, args.result_file)


# =========================================================================
# MAIN
# =========================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Filtrador de Saldos - Tracking Post-Remate"
    )
    parser.add_argument(
        "--primera-run", action="store_true",
        help="Modo auditoria: no elimina causas, solo llena log_decision",
    )
    parser.add_argument(
        "--solo-merge", action="store_true",
        help="Solo importar reportes sin consultar OJV",
    )
    parser.add_argument(
        "--solo-filtro1", action="store_true",
        help="Merge + solo Filtro 1 (no ejecuta Filtro 2/3)",
    )
    parser.add_argument(
        "--reaudit", action="store_true",
        help="Reprocesar causas PENDIENTE_ACTA: leer tabla completa, guardar HTML, re-aplicar Filtro 1",
    )
    parser.add_argument(
        "--skip-pdf", action="store_true",
        help="No descargar PDFs de actas, marcar NECESITA_PDF_ACTA como PENDIENTE_LIQUIDACION",
    )
    parser.add_argument(
        "--recheck-cargo", action="store_true",
        help="Re-analiza PDFs de actas ya descargados para verificar cargo al credito con regex actualizado",
    )
    parser.add_argument(
        "--solo-filtro3", action="store_true",
        help="Solo ejecutar Filtro 3 (descarga liquidaciones) en causas elegibles",
    )
    parser.add_argument(
        "--recheck-liq", action="store_true",
        help="Re-analiza PDFs de liquidacion ya descargados con regex actualizado",
    )
    parser.add_argument(
        "--workers", type=int, default=5,
        help="Numero de procesos paralelos (default: 5, max: 5)",
    )
    parser.add_argument(
        "--skip-refresh-deuda", action="store_true",
        help="Salta el Paso 0.5 (refresh de MONTO_DEUDA_CLP via M3 sobre PDFs locales)",
    )
    # Args internos para modo worker (no usar manualmente)
    parser.add_argument("--worker-mode", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--chunk-file", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--result-file", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--worker-id", type=int, default=0, help=argparse.SUPPRESS)
    parser.add_argument("--audit-dir", type=str, default=None, help=argparse.SUPPRESS)
    args = parser.parse_args()

    # === Modo worker: proceso hijo, no orquestador ===
    if args.worker_mode:
        _run_worker_mode(args)
        return

    # === Modo orquestador (normal) ===
    primera_run = args.primera_run
    solo_filtro1 = args.solo_filtro1
    n_workers = min(max(args.workers, 1), 5)

    reaudit = args.reaudit
    skip_pdf = args.skip_pdf

    log.info("=== FILTRADOR DE SALDOS ===")
    if reaudit:
        log.info("*** MODO REAUDIT: Reprocesando causas PENDIENTE_ACTA ***")
    if skip_pdf:
        log.info("*** MODO --skip-pdf: No se descargaran PDFs de actas ***")
    if primera_run:
        log.info("*** MODO AUDITORIA (--primera-run): No se eliminaran causas ***")
    if solo_filtro1:
        log.info("*** MODO SOLO FILTRO 1: No se ejecutaran Filtros 2/3 ***")
    if args.solo_filtro3:
        log.info("*** MODO SOLO FILTRO 3: Descarga de liquidaciones ***")
    if n_workers > 1:
        log.info("*** Workers: %d procesos paralelos ***", n_workers)

    # === PASO 0: Merge reportes ===
    df = paso0_merge_reportes()

    if args.solo_merge:
        _guardar_excel_formateado(df)
        log.info("Excel madre guardado (solo merge).")
        return

    # === PASO 0.5: Refresh deuda via M3 (modos compatibles solamente) ===
    # Excluir modos que no necesitan/deben recomputar deuda:
    #   --solo-filtro1, --solo-filtro3, --recheck-cargo, --recheck-liq
    # Incluir: run normal, --reaudit, --skip-pdf (lee PDFs locales, no descarga).
    _modos_excluyen_05 = (
        args.solo_filtro1 or args.solo_filtro3
        or args.recheck_cargo or args.recheck_liq
    )
    if args.skip_refresh_deuda:
        log.info("[PASO 0.5] Saltado por flag --skip-refresh-deuda")
    elif _modos_excluyen_05:
        log.info("[PASO 0.5] Saltado (modo --solo-* o --recheck-* incompatible)")
    else:
        df = _paso_refresh_deuda_m3(df)

    # === Modo REAUDIT: reprocesar PENDIENTE_ACTA ===
    if reaudit:
        # Incluir PENDIENTE_ACTA + PENDIENTE_LIQUIDACION sin monto_acta
        # (acta detectada en tabla pero PDF nunca descargado)
        mask_acta = df["estado"] == "PENDIENTE_ACTA"
        mask_liq_sin_monto = (
            (df["estado"] == "PENDIENTE_LIQUIDACION")
            & (df["monto_acta_remate"].isna() | (df["monto_acta_remate"].astype(str).str.strip() == ""))
        )
        mask_reaudit = mask_acta | mask_liq_sin_monto
        indices_reaudit = df[mask_reaudit].index.tolist()

        if not indices_reaudit:
            log.info("No hay causas para reaudit")
            _guardar_excel_formateado(df)
            return

        cnt_acta = mask_acta.sum()
        cnt_liq = mask_liq_sin_monto.sum()
        log.info("[REAUDIT] Procesando %d causas (%d PENDIENTE_ACTA + %d PENDIENTE_LIQUIDACION sin monto_acta)",
                 len(indices_reaudit), cnt_acta, cnt_liq)

        causas_reaudit = []
        for idx in indices_reaudit:
            causa = df.loc[idx].to_dict()
            causas_reaudit.append((idx, causa))

        causas_reaudit.sort(
            key=lambda x: (str(x[1].get("CORTE", "")),
                           str(x[1].get("TRIBUNAL", "")))
        )

        os.makedirs(_HTML_AUDIT_DIR, exist_ok=True)
        # Limpiar HTMLs anteriores
        for _f in os.listdir(_HTML_AUDIT_DIR):
            if _f.endswith(".html"):
                try:
                    os.remove(os.path.join(_HTML_AUDIT_DIR, _f))
                except OSError:
                    pass

        t0_reaudit = time.time()
        effective_workers = min(n_workers, len(causas_reaudit))

        if effective_workers == 1:
            with sync_playwright() as p:
                _profile_dir = os.path.join(REMATES_DIR, ".chrome-profile")
                context = p.chromium.launch_persistent_context(
                    _profile_dir,
                    headless=False,
                    slow_mo=100,
                    channel="chrome",
                    args=["--disable-blink-features=AutomationControlled"],
                    accept_downloads=True,
                )
                page = context.pages[0] if context.pages else context.new_page()
                Stealth().apply_stealth_sync(page)
                page.set_default_timeout(15000)

                if not navegar_a_consulta(page):
                    log.error("No se pudo navegar a OJV. Abortando reaudit.")
                    context.close()
                    _guardar_excel_formateado(df)
                    return

                resultados = _procesar_chunk_reaudit(
                    page, context, causas_reaudit,
                    worker_id=0, audit_dir=_HTML_AUDIT_DIR,
                    skip_pdf=skip_pdf,
                )
                for idx, causa in resultados.items():
                    for col in causa:
                        if col in df.columns:
                            df.at[idx, col] = causa[col]

                context.close()
        else:
            all_resultados = _ejecutar_paralelo(
                causas_reaudit, effective_workers,
                solo_filtro1=False, primera_run=primera_run,
                reaudit=True, skip_pdf=skip_pdf,
            )
            for idx, causa in all_resultados.items():
                for col in causa:
                    if col in df.columns:
                        df.at[idx, col] = causa[col]

        elapsed_reaudit = time.time() - t0_reaudit

        # Contadores reaudit
        cnt_reaudit_elim = (df.loc[indices_reaudit, "estado"] == "ELIMINADA").sum()
        cnt_reaudit_liq = (df.loc[indices_reaudit, "estado"] == "PENDIENTE_LIQUIDACION").sum()
        cnt_reaudit_rev = (df.loc[indices_reaudit, "estado"] == "PENDIENTE_REVISION_MANUAL").sum()
        cnt_reaudit_acta = (df.loc[indices_reaudit, "estado"] == "PENDIENTE_ACTA").sum()

        # Contar HTMLs
        n_htmls = 0
        try:
            n_htmls = len([f for f in os.listdir(_HTML_AUDIT_DIR) if f.endswith(".html")])
        except OSError:
            pass

        # Eliminar causas si no es modo auditoria
        if not primera_run:
            mascara_elim = df["estado"] == "ELIMINADA"
            n_elim = int(mascara_elim.sum())
            if n_elim:
                _registrar_en_historial_eliminadas(df, mascara_elim)
                df = df[~mascara_elim].reset_index(drop=True)
                log.info("Causas eliminadas del Excel: %d", n_elim)

        _guardar_excel_formateado(df)

        log.info("")
        log.info("=" * 52)
        log.info("  RESUMEN REAUDIT")
        log.info("=" * 52)
        log.info("  Causas procesadas:            %d", len(indices_reaudit))
        log.info("  Tiempo total: %d min %d seg",
                 int(elapsed_reaudit) // 60, int(elapsed_reaudit) % 60)
        log.info("  Eliminadas:                   %d", cnt_reaudit_elim)
        log.info("  Pendientes liquidacion:       %d", cnt_reaudit_liq)
        log.info("  Pendientes revision manual:   %d", cnt_reaudit_rev)
        log.info("  Pendientes acta (sin cambio): %d", cnt_reaudit_acta)
        log.info("  Auditorias HTML guardadas:    %d", n_htmls)
        log.info("=" * 52)
        log.info("Excel madre guardado: %s", EXCEL_MADRE)
        return

    # === Modo RECHECK-CARGO: re-analizar PDFs locales con regex actualizado ===
    if args.recheck_cargo:
        from config import ACTAS_DIR
        log.info("[RECHECK-CARGO] Re-verificando cargo al credito en PDFs existentes...")
        os.makedirs(ACTAS_DIR, exist_ok=True)
        cambios = 0
        revisadas = 0
        for idx, row in df.iterrows():
            if row["estado"] == "PENDIENTE_LIQUIDACION" and pd.notna(row.get("monto_acta_remate")) and str(row.get("monto_acta_remate", "")).strip():
                rol = str(row["ROL"]).strip()
                año = str(row["AÑO"]).strip()
                filename = f"acta_{rol}_{año}.pdf"
                filepath = os.path.join(ACTAS_DIR, filename)
                if os.path.exists(filepath):
                    revisadas += 1
                    analisis = _analizar_pdf_acta(filepath)
                    # Actualizar monto si cambió (tope de sanidad puede anularlo)
                    nuevo_monto = analisis["monto_adjudicacion"]
                    monto_anterior = row.get("monto_acta_remate")
                    try:
                        monto_anterior_int = int(float(monto_anterior)) if pd.notna(monto_anterior) and str(monto_anterior).strip() else None
                    except (ValueError, TypeError):
                        monto_anterior_int = None
                    if nuevo_monto != monto_anterior_int:
                        df.at[idx, "monto_acta_remate"] = str(nuevo_monto) if nuevo_monto else ""
                        log.info("  C-%s-%s: Monto actualizado %s -> %s", rol, año, monto_anterior, nuevo_monto)

                    if analisis["cargo_al_credito"]:
                        log.info("  C-%s-%s: CARGO AL CREDITO detectado -> ELIMINAR", rol, año)
                        df.at[idx, "estado"] = "ELIMINAR"
                        df.at[idx, "detalle_auditoria"] = "RECHECK: Cargo al credito confirmado con regex actualizado"
                        cambios += 1
                    else:
                        log.info("  C-%s-%s: cargo=False confirmado", rol, año)
                else:
                    log.info("  C-%s-%s: PDF no encontrado (%s)", rol, año, filename)

        # Guardado progresivo: salvar cambios aunque haya crash posterior
        if cambios > 0:
            _guardar_excel_formateado(df)
            log.info("[RECHECK-CARGO] Excel guardado progresivamente con %d cambios", cambios)

        log.info("[RECHECK-CARGO] %d causas revisadas, %d cambiadas a ELIMINAR", revisadas, cambios)

        if not primera_run and cambios > 0:
            mascara_elim = df["estado"] == "ELIMINADA"
            n_elim = int(mascara_elim.sum())
            if n_elim:
                _registrar_en_historial_eliminadas(df, mascara_elim)
                df = df[~mascara_elim].reset_index(drop=True)
                log.info("Causas eliminadas del Excel: %d", n_elim)

        _guardar_excel_formateado(df)
        log.info("Excel madre guardado: %s", EXCEL_MADRE)
        return

    # === Modo RECHECK-LIQ: re-analizar PDFs de liquidacion locales ===
    if args.recheck_liq:
        log.info("[RECHECK-LIQ] Re-analizando PDFs de liquidacion con regex actualizado...")
        cambios = 0
        revisadas = 0

        for idx, row in df.iterrows():
            ruta_liq = str(row.get("ruta_liquidacion", "")).strip()
            if not ruta_liq or not os.path.exists(ruta_liq):
                continue
            estado = str(row.get("estado", ""))
            if estado not in ("PENDIENTE_REVISION_MANUAL", "PENDIENTE_LIQUIDACION",
                              "LIQUIDACION_ENCONTRADA"):
                continue

            rol = str(row["ROL"]).strip()
            anio = str(row["AÑO"]).strip()
            revisadas += 1
            causa_dict = row.to_dict()

            analisis = _analizar_pdf_liquidacion(ruta_liq, causa_dict)

            nuevo_estado = None
            nuevo_log = None

            if analisis["saldo_encontrado"] and analisis.get("monto_saldo") and analisis["monto_saldo"] > 0:
                nuevo_estado = "EXCEDENTE_CONFIRMADO"
                monto = analisis["monto_saldo"]
                df.at[idx, "monto_liquidacion_saldo"] = str(monto)
                nuevo_log = f"RECHECK-LIQ: Saldo ${monto:,} (patron: '{analisis['patron_usado']}')"
                log.info("  C-%s-%s: Saldo $%s -> EXCEDENTE_CONFIRMADO", rol, anio, f"{monto:,}")
            elif analisis.get("saldo_contra") or analisis.get("sin_excedente"):
                nuevo_estado = "ELIMINAR"
                nuevo_log = f"RECHECK-LIQ: {analisis['patron_usado']}"
                log.info("  C-%s-%s: %s -> ELIMINAR", rol, anio, analisis["patron_usado"])
            elif analisis.get("tipo_liquidacion") == "B" and analisis.get("credito_adeudado"):
                credito = analisis["credito_adeudado"]
                df.at[idx, "monto_credito_liquidado"] = str(credito)
                monto_acta = _deuda_a_int(row.get("monto_acta_remate", ""))
                if monto_acta > 0:
                    delta = monto_acta - credito
                    if delta > 0:
                        nuevo_estado = "EXCEDENTE_CONFIRMADO"
                        df.at[idx, "monto_liquidacion_saldo"] = str(delta)
                        nuevo_log = (
                            f"RECHECK-LIQ: Tipo B. Acta ${monto_acta:,} - Credito ${credito:,} "
                            f"= Excedente ${delta:,}"
                        )
                        log.info("  C-%s-%s: Tipo B excedente $%s", rol, anio, f"{delta:,}")
                    else:
                        nuevo_estado = "ELIMINAR"
                        nuevo_log = (
                            f"RECHECK-LIQ: Tipo B. Acta ${monto_acta:,} - Credito ${credito:,} "
                            f"= ${delta:,} (sin excedente)"
                        )
                        log.info("  C-%s-%s: Tipo B sin excedente", rol, anio)
                else:
                    nuevo_estado = "PENDIENTE_REVISION_MANUAL"
                    nuevo_log = (
                        f"RECHECK-LIQ: Tipo B. Credito adeudado ${credito:,}. "
                        f"Sin monto acta para comparar."
                    )
                    log.info("  C-%s-%s: Tipo B sin acta -> REVISION_MANUAL", rol, anio)
            else:
                log.info("  C-%s-%s: Sin match de regex", rol, anio)
                continue

            if nuevo_estado:
                df.at[idx, "estado"] = nuevo_estado
                df.at[idx, "log_decision"] = nuevo_log
                cambios += 1

        log.info("[RECHECK-LIQ] %d revisadas, %d cambios", revisadas, cambios)

        if not primera_run and cambios > 0:
            mascara_elim = df["estado"] == "ELIMINADA"
            n_elim = int(mascara_elim.sum())
            if n_elim:
                _registrar_en_historial_eliminadas(df, mascara_elim)
                df = df[~mascara_elim].reset_index(drop=True)
                log.info("Causas eliminadas del Excel: %d", n_elim)

        _guardar_excel_formateado(df)
        _generar_excel_liquidaciones(df)
        log.info("Excel madre guardado: %s", EXCEL_MADRE)
        return

    # === Modo SOLO-FILTRO3: procesar causas con keyword liquidacion ===
    if args.solo_filtro3:
        causas_f3 = _seleccionar_causas_f3(df)
        if not causas_f3:
            log.info("[F3] No hay causas elegibles para Filtro 3")
            _guardar_excel_formateado(df)
            return

        log.info("[F3] Procesando %d causas PENDIENTE_LIQUIDACION", len(causas_f3))
        os.makedirs(LIQUIDACIONES_DIR, exist_ok=True)
        os.makedirs(LIQUIDACIONES_RAW_DIR, exist_ok=True)

        cnt_excedente = 0
        cnt_eliminar = 0
        cnt_revision = 0
        cnt_pendiente = 0

        with sync_playwright() as p:
            _profile_dir = os.path.join(REMATES_DIR, ".chrome-profile")
            context = p.chromium.launch_persistent_context(
                _profile_dir,
                headless=False,
                slow_mo=100,
                channel="chrome",
                args=["--disable-blink-features=AutomationControlled"],
                accept_downloads=True,
            )
            page = context.pages[0] if context.pages else context.new_page()
            Stealth().apply_stealth_sync(page)
            page.set_default_timeout(15000)

            if not navegar_a_consulta(page):
                log.error("[F3] No se pudo navegar a OJV. Abortando.")
                context.close()
                _guardar_excel_formateado(df)
                return

            ultimo_tribunal = None

            for idx, causa in causas_f3:
                rol = str(causa.get("ROL", ""))
                anio = str(causa.get("AÑO", ""))
                tribunal = str(causa.get("TRIBUNAL", ""))
                corte = str(causa.get("CORTE", ""))

                log.info("[F3] C-%s-%s (%s)...", rol, anio, tribunal)

                # Navegar a la causa en OJV
                try:
                    tribunal_actual = f"{corte}|{tribunal}"
                    if tribunal_actual != ultimo_tribunal:
                        limpiar_formulario(page)
                        page.wait_for_timeout(500)

                    found = buscar_causa(page, rol, anio, corte, tribunal)
                    if not found:
                        log.warning("  [F3] C-%s-%s: No encontrada en OJV", rol, anio)
                        cnt_pendiente += 1
                        continue

                    if not abrir_detalle(page, rol, anio):
                        log.warning("  [F3] C-%s-%s: No se pudo abrir detalle", rol, anio)
                        cnt_pendiente += 1
                        continue

                    ultimo_tribunal = tribunal_actual

                    # Seleccionar cuaderno Apremio
                    try:
                        seleccionar_cuaderno(page, "Apremio")
                    except Exception:
                        log.warning("  [F3] C-%s-%s: Error seleccionando cuaderno Apremio", rol, anio)

                    page.wait_for_timeout(1500)

                    # Procesar F3
                    result = _procesar_causa_f3(page, context, causa)

                    # Actualizar DataFrame
                    df.at[idx, "estado"] = result["estado"]
                    df.at[idx, "log_decision"] = result["log_decision"]
                    if result.get("ruta_liquidacion"):
                        df.at[idx, "ruta_liquidacion"] = result["ruta_liquidacion"]
                    if result.get("monto_liquidacion_saldo"):
                        df.at[idx, "monto_liquidacion_saldo"] = result["monto_liquidacion_saldo"]
                    if result.get("monto_credito_liquidado"):
                        df.at[idx, "monto_credito_liquidado"] = result["monto_credito_liquidado"]

                    if result["estado"] == "EXCEDENTE_CONFIRMADO":
                        cnt_excedente += 1
                    elif result["estado"] == "ELIMINAR":
                        cnt_eliminar += 1
                    elif result["estado"] == "PENDIENTE_REVISION_MANUAL":
                        cnt_revision += 1
                    else:
                        cnt_pendiente += 1

                    # Cerrar modal
                    _cerrar_modal_detalle(page)

                except Exception as e:
                    log.warning("  [F3] C-%s-%s: Error general: %s", rol, anio, e)
                    cnt_pendiente += 1
                    try:
                        _cerrar_modal_detalle(page)
                    except Exception:
                        pass

            context.close()

        # Guardar resultados
        if not primera_run:
            mascara_elim = df["estado"] == "ELIMINADA"
            n_elim_total = int(mascara_elim.sum())
            if n_elim_total:
                _registrar_en_historial_eliminadas(df, mascara_elim)
                df = df[~mascara_elim].reset_index(drop=True)
                log.info("[F3] Causas eliminadas del Excel: %d", n_elim_total)

        _guardar_excel_formateado(df)
        _generar_excel_liquidaciones(df)

        log.info("")
        log.info("=" * 52)
        log.info("  RESUMEN FILTRO 3")
        log.info("=" * 52)
        log.info("  Causas procesadas:          %d", len(causas_f3))
        log.info("  Excedentes confirmados:     %d", cnt_excedente)
        log.info("  Eliminadas (sin excedente): %d", cnt_eliminar)
        log.info("  Revision manual:            %d", cnt_revision)
        log.info("  Pendientes (sin cambio):    %d", cnt_pendiente)
        log.info("=" * 52)
        log.info("Excel madre guardado: %s", EXCEL_MADRE)
        return

    # === Determinar causas a procesar ===
    if solo_filtro1:
        estados_objetivo = _ESTADOS_FILTRO1
    else:
        estados_objetivo = _ESTADOS_PROCESAR

    mask = df["estado"].isin(estados_objetivo)
    indices = df[mask].index.tolist()

    # Filtrar no elegibles por timing (< 2 dias)
    hoy = date.today()
    cnt_no_elegible = 0
    indices_elegibles = []
    for idx in indices:
        fr_str = str(df.at[idx, "fecha_remate"])
        try:
            fecha_r = date.fromisoformat(fr_str)
            dias = (hoy - fecha_r).days
            if dias < 2:
                cnt_no_elegible += 1
                df.at[idx, "detalle_auditoria"] = (
                    f"No elegible: {dias} dia(s) desde remate ({fecha_r.isoformat()}). "
                    "Minimo 2 dias."
                )
                continue
        except (ValueError, TypeError):
            pass
        indices_elegibles.append(idx)

    # Contadores
    cnt_f1_error = 0
    elapsed_total = 0
    effective_workers = n_workers

    if not indices_elegibles:
        log.info("No hay causas elegibles para procesar en OJV")
        if cnt_no_elegible:
            log.info("  (%d causas con < 2 dias, no elegibles aun)", cnt_no_elegible)
    else:
        log.info("Procesando %d causas elegibles (dias >= 2)...",
                 len(indices_elegibles))
        if cnt_no_elegible:
            log.info("  (%d causas con < 2 dias, no elegibles aun)", cnt_no_elegible)

        # =================================================================
        # Optimizacion 1: Ordenar por (corte, tribunal)
        # =================================================================
        causas_para_procesar = []
        for idx in indices_elegibles:
            causa = df.loc[idx].to_dict()
            causas_para_procesar.append((idx, causa))

        causas_para_procesar.sort(
            key=lambda x: (str(x[1].get("CORTE", "")),
                           str(x[1].get("TRIBUNAL", "")))
        )

        # Log grupos
        from itertools import groupby
        for (corte, trib), grupo in groupby(
            causas_para_procesar,
            key=lambda x: (str(x[1].get("CORTE", "")),
                           str(x[1].get("TRIBUNAL", "")))
        ):
            cnt = sum(1 for _ in grupo)
            log.info("  Grupo: %s / %s (%d causas)", corte, trib, cnt)

        # Limpiar HTMLs de auditoria anteriores
        os.makedirs(_HTML_AUDIT_DIR, exist_ok=True)
        for _f in os.listdir(_HTML_AUDIT_DIR):
            if _f.endswith(".html"):
                try:
                    os.remove(os.path.join(_HTML_AUDIT_DIR, _f))
                except OSError:
                    pass

        t0_total = time.time()

        # =================================================================
        # Optimizacion 2: Workers (secuencial o paralelo con subprocess)
        # =================================================================
        effective_workers = min(n_workers, len(causas_para_procesar))

        if effective_workers == 1:
            # --- Modo secuencial (sin subprocesos) ---
            with sync_playwright() as p:
                _profile_dir = os.path.join(REMATES_DIR, ".chrome-profile")
                context = p.chromium.launch_persistent_context(
                    _profile_dir,
                    headless=False,
                    slow_mo=100,
                    channel="chrome",
                    args=["--disable-blink-features=AutomationControlled"],
                    accept_downloads=True,
                )
                page = context.pages[0] if context.pages else context.new_page()
                Stealth().apply_stealth_sync(page)
                page.set_default_timeout(15000)

                if not navegar_a_consulta(page):
                    log.error("No se pudo navegar a OJV. Abortando.")
                    context.close()
                    _guardar_excel_formateado(df)
                    return

                resultados = _procesar_chunk(
                    page, context, causas_para_procesar,
                    worker_id=0, solo_filtro1=solo_filtro1,
                    audit_dir=_HTML_AUDIT_DIR,
                    skip_pdf=skip_pdf,
                )

                for idx, causa in resultados.items():
                    for col in causa:
                        if col in df.columns:
                            df.at[idx, col] = causa[col]

                context.close()

        else:
            # --- Modo paralelo (subprocess — cada worker = proceso Python) ---
            all_resultados = _ejecutar_paralelo(
                causas_para_procesar, effective_workers,
                solo_filtro1=solo_filtro1,
                primera_run=primera_run,
                skip_pdf=skip_pdf,
            )

            for idx, causa in all_resultados.items():
                for col in causa:
                    if col in df.columns:
                        df.at[idx, col] = causa[col]

        elapsed_total = time.time() - t0_total

    # === FILTRO 3: Liquidaciones (en flujo normal, despues de F1/F2) ===
    if not solo_filtro1:
        causas_f3 = _seleccionar_causas_f3(df)
        if causas_f3:
            log.info("[F3] Procesando %d causas PENDIENTE_LIQUIDACION", len(causas_f3))
            os.makedirs(LIQUIDACIONES_DIR, exist_ok=True)
            os.makedirs(LIQUIDACIONES_RAW_DIR, exist_ok=True)

            cnt_f3_excedente = 0
            cnt_f3_eliminar = 0
            cnt_f3_revision = 0
            cnt_f3_pendiente = 0

            with sync_playwright() as p:
                _profile_dir = os.path.join(REMATES_DIR, ".chrome-profile")
                context = p.chromium.launch_persistent_context(
                    _profile_dir,
                    headless=False,
                    slow_mo=100,
                    channel="chrome",
                    args=["--disable-blink-features=AutomationControlled"],
                    accept_downloads=True,
                )
                page = context.pages[0] if context.pages else context.new_page()
                Stealth().apply_stealth_sync(page)
                page.set_default_timeout(15000)

                if not navegar_a_consulta(page):
                    log.error("[F3] No se pudo navegar a OJV. Saltando F3.")
                else:
                    ultimo_tribunal = None

                    for idx, causa in causas_f3:
                        rol = str(causa.get("ROL", ""))
                        anio = str(causa.get("AÑO", ""))
                        tribunal = str(causa.get("TRIBUNAL", ""))
                        corte = str(causa.get("CORTE", ""))

                        log.info("[F3] C-%s-%s (%s)...", rol, anio, tribunal)

                        try:
                            tribunal_actual = f"{corte}|{tribunal}"
                            if tribunal_actual != ultimo_tribunal:
                                limpiar_formulario(page)
                                page.wait_for_timeout(500)

                            found = buscar_causa(page, rol, anio, corte, tribunal)
                            if not found:
                                log.warning("  [F3] C-%s-%s: No encontrada en OJV", rol, anio)
                                cnt_f3_pendiente += 1
                                continue

                            if not abrir_detalle(page, rol, anio):
                                log.warning("  [F3] C-%s-%s: No se pudo abrir detalle", rol, anio)
                                cnt_f3_pendiente += 1
                                continue

                            ultimo_tribunal = tribunal_actual

                            try:
                                seleccionar_cuaderno(page, "Apremio")
                            except Exception:
                                log.warning("  [F3] C-%s-%s: Error seleccionando cuaderno Apremio", rol, anio)

                            page.wait_for_timeout(1500)

                            result = _procesar_causa_f3(page, context, causa)

                            df.at[idx, "estado"] = result["estado"]
                            df.at[idx, "log_decision"] = result["log_decision"]
                            if result.get("ruta_liquidacion"):
                                df.at[idx, "ruta_liquidacion"] = result["ruta_liquidacion"]
                            if result.get("monto_liquidacion_saldo"):
                                df.at[idx, "monto_liquidacion_saldo"] = result["monto_liquidacion_saldo"]
                            if result.get("monto_credito_liquidado"):
                                df.at[idx, "monto_credito_liquidado"] = result["monto_credito_liquidado"]

                            if result["estado"] == "EXCEDENTE_CONFIRMADO":
                                cnt_f3_excedente += 1
                            elif result["estado"] == "ELIMINAR":
                                cnt_f3_eliminar += 1
                            elif result["estado"] == "PENDIENTE_REVISION_MANUAL":
                                cnt_f3_revision += 1
                            else:
                                cnt_f3_pendiente += 1

                            _cerrar_modal_detalle(page)

                        except Exception as e:
                            log.warning("  [F3] C-%s-%s: Error general: %s", rol, anio, e)
                            cnt_f3_pendiente += 1
                            try:
                                _cerrar_modal_detalle(page)
                            except Exception:
                                pass

                context.close()

            log.info("[F3] Resumen: %d excedentes, %d eliminadas, %d revision, %d pendientes",
                     cnt_f3_excedente, cnt_f3_eliminar, cnt_f3_revision, cnt_f3_pendiente)
        else:
            log.info("[F3] No hay causas elegibles para Filtro 3")

    # === Contadores globales para resumen ===
    cnt_elim_susp = 0
    cnt_elim_postores = 0
    cnt_elim_reprog = 0
    cnt_elim_excedente = 0

    eliminadas_mask = df["estado"] == "ELIMINADA"
    for _, row in df[eliminadas_mask].iterrows():
        log_dec = str(row.get("log_decision", "")).lower()
        if "suspension" in log_dec or "suspende" in log_dec:
            cnt_elim_susp += 1
        elif "postores" in log_dec:
            cnt_elim_postores += 1
        elif "reprogramad" in log_dec or "nuevo d" in log_dec:
            cnt_elim_reprog += 1
        elif "excedente" in log_dec or "monto acta" in log_dec:
            cnt_elim_excedente += 1

    cnt_liquidacion = (df["estado"] == "LIQUIDACION_ENCONTRADA").sum()
    cnt_pendiente_acta = (df["estado"] == "PENDIENTE_ACTA").sum()
    cnt_pendiente_liq = (df["estado"] == "PENDIENTE_LIQUIDACION").sum()
    cnt_pendiente_f1 = (df["estado"] == "PENDIENTE_FILTRO1").sum()
    cnt_revision_manual = (df["estado"] == "PENDIENTE_REVISION_MANUAL").sum()

    # Contar errores
    for idx in indices_elegibles:
        if idx < len(df):
            log_dec = str(df.at[idx, "log_decision"])
            if "ERROR" in log_dec:
                cnt_f1_error += 1

    # === Eliminar causas (o no, en modo auditoria) ===
    if not primera_run:
        mascara_elim = df["estado"] == "ELIMINADA"
        n_elim = int(mascara_elim.sum())
        if n_elim:
            _registrar_en_historial_eliminadas(df, mascara_elim)
            df = df[~mascara_elim].reset_index(drop=True)
            log.info("Causas eliminadas del Excel: %d", n_elim)
    else:
        log.info(
            "Modo auditoria: causas marcadas ELIMINADA se mantienen para revision"
        )

    # === Guardar Excel madre formateado ===
    _guardar_excel_formateado(df)
    _generar_excel_liquidaciones(df)

    # === Resumen en consola ===
    log.info("")
    log.info("=" * 52)
    log.info("  RESUMEN DE EJECUCION")
    log.info("=" * 52)

    if indices_elegibles:
        if n_workers > 1:
            log.info("  Workers: %d", effective_workers)
        log.info("  Tiempo total: %d min %d seg",
                 int(elapsed_total) // 60, int(elapsed_total) % 60)
        log.info("")

    log.info("  Resumen Filtro 1:")
    log.info("    Eliminadas (suspension):    %d", cnt_elim_susp)
    log.info("    Eliminadas (sin postores):  %d", cnt_elim_postores)
    log.info("    Eliminadas (reprogramado):  %d", cnt_elim_reprog)
    log.info("    Pendientes acta:            %d", cnt_pendiente_acta)
    log.info("    Pendientes liquidacion:     %d", cnt_pendiente_liq)
    log.info("    Errores OJV:                %d", cnt_f1_error)
    log.info("    No elegibles (< 2 dias):    %d", cnt_no_elegible)

    if not solo_filtro1:
        log.info("  Eliminadas (sin excedente):   %d", cnt_elim_excedente)
        log.info("  Con liquidacion encontrada:   %d", cnt_liquidacion)

    log.info("")
    log.info("  Pendientes filtro 1:          %d", cnt_pendiente_f1)
    log.info("  Pendientes revision manual:   %d", cnt_revision_manual)
    log.info("  Total causas en Excel madre:  %d", len(df))

    # Contar HTMLs de auditoria
    try:
        n_htmls = len([f for f in os.listdir(_HTML_AUDIT_DIR) if f.endswith(".html")])
        if n_htmls:
            log.info("  HTMLs auditoria guardados:    %d (en %s)", n_htmls, _HTML_AUDIT_DIR)
    except OSError:
        pass

    log.info("=" * 52)

    if primera_run:
        log.info("")
        log.info("Modo auditoria activo: NO se eliminaron causas.")
        log.info("Revisa columna \"Detalle Auditoria\" en el Excel.")

    log.info("Excel madre guardado: %s", EXCEL_MADRE)


if __name__ == "__main__":
    main()
