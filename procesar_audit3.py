"""
procesar_audit3.py
==================
Aplica las decisiones del TERCER_INFORME (Audit3) al Excel madre
'Causas_posible_saldo.xlsx'.

USO:
    python procesar_audit3.py [--dry-run]

QUE HACE:
1. Backup automatico del Excel con timestamp -> 'Causas_posible_saldo_backup_YYYYMMDD_HHMMSS.xlsx'
2. Asegura que existen las columnas 'observacion_abogado' y 'fecha_revision_abogado'
   en la hoja '_datos_internos' (las crea si faltan).
3. Para cada una de las 11 causas del Audit3:
   - 9 NO    -> estado = ELIMINAR, log_decision se pre-pendea con "Audit3 [fecha]: NO - <obs>"
   - 2 SI    -> mantiene estado, log se pre-pendea con "Audit3 [fecha]: SI - <obs>"
   En ambos casos: pobla 'observacion_abogado' y 'fecha_revision_abogado'.

IDEMPOTENCIA:
- Si ya existe una entrada Audit3 en el log para esa causa, NO la duplica
  (compara prefijo "Audit3 [" en log_decision).
- Si las columnas ya existen, no las recrea.

REQUIERE:
- openpyxl (ya instalado segun dependencias del proyecto)
- Excel CERRADO al ejecutar (script reintenta 3 veces si esta abierto).

NO MODIFICA: pestañas derivadas (Causas con Saldo, Por Antiguedad, Revision Manual).
Esas se regeneran en el proximo run del filtrador (la fuente de verdad es _datos_internos).
"""

import sys
import shutil
import time
import argparse
from datetime import date, datetime
from pathlib import Path

# ----------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------
EXCEL_PATH = Path(r"D:\Remates\Causas con liq\Causas_posible_saldo.xlsx")
SHEET = "_datos_internos"
HOY = date.today().isoformat()
PREFIJO_LOG = f"Audit3 [{HOY}]"

# Datos extraidos del TERCER_INFORME devuelto por el abogado
# Formato: (rol, decision, observacion)
# decision: "SI" o "NO"
DECISIONES = [
    ("C-7262-2023", "NO", "FOLIO  117 CUADERNO DE APREMIO REMATADA CON CARGO AL CREDITO….. NO HABRÀ EXCEDENTE"),
    ("C-2395-2024", "NO", "DEMAMDA ES POR GASTOS COMUNES PERO TAMBIEN HAY HIPOTECA DEL BANCO DE CHILE, NO QUEDARÀ SALDO"),
    ("C-928-2025",  "NO", "LIQUIDACION 15 DE ABRIL CON SALDO NEGATIVO"),
    ("C-1081-2018", "NO", "MUCHOS ATADOS LEGALES        NUEVA FECHA REMATE JUNIO 2026"),
    ("C-2440-2025", "NO", "FOLIO 28 APREMIO REMATE CON CARGO AL CREDITO"),
    ("C-1349-2023", "NO", "POSIBLE AVENIMIENTO"),
    ("C-3302-2024", "NO", "REMATADA POR 11 MILLONES Y LA DEUDA SON CASI 50 MILLONES….CERO EXCEDENTE"),
    ("C-1643-2025", "NO", "REMATADA CON CARGO AL CREDITO"),
    ("C-82-2020",   "NO", "DEMANDADA PAGÒ LA DEUDA"),
    ("C-1485-2025", "SI", "LIQUIDACION DEUDA APROX 39 MILLONES T REMATADA POR 46 MILLONES, ESPERAR LIQUIDACION POSTERIOR AL REMATE POSIBLE EXCEDENTE DE APROX 5 MILLONES"),
    ("C-4306-2024", "SI", "ESPERAR LIQUIDACION POSIBLE SALDO DE APROX 30 MILLONES"),
]

COLS_NUEVAS = ["observacion_abogado", "fecha_revision_abogado"]


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def log(msg):
    """Print solo ASCII (Windows cp1252)."""
    safe = msg.encode("ascii", errors="replace").decode("ascii")
    print(safe)


def backup(path: Path) -> Path:
    """Copia el Excel con timestamp."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = path.with_name(f"{path.stem}_backup_{ts}{path.suffix}")
    shutil.copy(path, dst)
    return dst


def cargar_excel_con_retry(path: Path, intentos: int = 3, espera: float = 2.0):
    """Carga el Excel reintentando si esta abierto."""
    from openpyxl import load_workbook
    for i in range(intentos):
        try:
            return load_workbook(path)
        except PermissionError:
            log(f"  Excel abierto, reintento {i+1}/{intentos} en {espera}s...")
            time.sleep(espera)
    raise PermissionError(f"No se pudo abrir {path}. Cierre Excel y reintente.")


def guardar_excel_con_retry(wb, path: Path, intentos: int = 3, espera: float = 2.0):
    """Guarda el Excel reintentando si esta abierto."""
    for i in range(intentos):
        try:
            wb.save(path)
            return
        except PermissionError:
            log(f"  Excel abierto al guardar, reintento {i+1}/{intentos} en {espera}s...")
            time.sleep(espera)
    raise PermissionError(f"No se pudo guardar {path}. Cierre Excel y reintente.")


def construir_clave(rol_raw, ano):
    """Normaliza ROL+AÑO a 'C-XXXX-YYYY'."""
    if rol_raw is None or ano is None:
        return None
    rol_str = str(rol_raw).strip()
    if rol_str.upper().startswith("C-"):
        num = rol_str[2:].split("-")[0]
    else:
        num = rol_str
    return f"C-{num}-{int(ano)}"


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main(dry_run: bool = False):
    log("=" * 70)
    log(f"procesar_audit3.py  (fecha: {HOY})")
    log(f"Modo: {'DRY-RUN (no guarda)' if dry_run else 'EJECUCION REAL'}")
    log("=" * 70)

    if not EXCEL_PATH.exists():
        log(f"ERROR: no existe {EXCEL_PATH}")
        sys.exit(1)

    # 1) Backup
    if not dry_run:
        bak = backup(EXCEL_PATH)
        log(f"\n[1] Backup creado: {bak.name}")
    else:
        log("\n[1] Backup omitido (dry-run)")

    # 2) Cargar
    log(f"\n[2] Cargando {EXCEL_PATH.name}...")
    wb = cargar_excel_con_retry(EXCEL_PATH)
    if SHEET not in wb.sheetnames:
        log(f"ERROR: no existe la hoja '{SHEET}' en {EXCEL_PATH}")
        sys.exit(1)
    ws = wb[SHEET]

    # 3) Indexar columnas
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    requeridas = ["ROL", "AAO" if "AAO" in col_idx else "AÑO", "estado", "log_decision"]
    # Buscar AÑO independientemente del encoding
    ano_col = None
    for h in headers:
        if h and str(h).strip().upper() in ("AÑO", "ANO", "AAO"):
            ano_col = h
            break
    if not ano_col:
        log("ERROR: no encuentro columna AÑO en headers")
        sys.exit(1)

    for c in ["ROL", "estado", "log_decision"]:
        if c not in col_idx:
            log(f"ERROR: falta columna '{c}'")
            sys.exit(1)

    # 4) Crear columnas nuevas si faltan
    log("\n[3] Verificando columnas nuevas...")
    for col in COLS_NUEVAS:
        if col not in col_idx:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col)
            col_idx[col] = new_col
            log(f"  Columna agregada: '{col}' (col {new_col})")
        else:
            log(f"  Columna ya existia: '{col}'")

    # 5) Mapear filas
    log("\n[4] Indexando filas por ROL-AÑO...")
    mapa = {}
    for r in range(2, ws.max_row + 1):
        rol = ws.cell(row=r, column=col_idx["ROL"]).value
        ano = ws.cell(row=r, column=col_idx[ano_col]).value
        clave = construir_clave(rol, ano)
        if clave:
            mapa[clave] = r
    log(f"  {len(mapa)} filas indexadas")

    # 6) Aplicar cambios
    log("\n[5] Aplicando decisiones del Audit3:")
    log("-" * 70)
    aplicadas = 0
    saltadas = 0
    no_encontradas = 0
    for rol, decision, obs in DECISIONES:
        fila = mapa.get(rol)
        if not fila:
            log(f"  [SKIP] {rol}: NO ENCONTRADA en _datos_internos")
            no_encontradas += 1
            continue

        log_actual = ws.cell(row=fila, column=col_idx["log_decision"]).value or ""

        # Idempotencia: si ya tiene una entrada Audit3 con esta misma fecha, saltar
        if PREFIJO_LOG in str(log_actual):
            log(f"  [SKIP] {rol}: ya procesada en {PREFIJO_LOG}")
            saltadas += 1
            continue

        estado_antes = ws.cell(row=fila, column=col_idx["estado"]).value

        if decision == "NO":
            nuevo_estado = "ELIMINAR"
        else:  # SI
            nuevo_estado = estado_antes  # mantener

        nueva_linea_log = f"{PREFIJO_LOG}: {decision} - {obs}"
        nuevo_log = f"{nueva_linea_log}\n{log_actual}".strip()

        ws.cell(row=fila, column=col_idx["estado"]).value = nuevo_estado
        ws.cell(row=fila, column=col_idx["log_decision"]).value = nuevo_log
        ws.cell(row=fila, column=col_idx["observacion_abogado"]).value = obs
        ws.cell(row=fila, column=col_idx["fecha_revision_abogado"]).value = HOY

        flecha = "->"
        log(f"  [{decision}] {rol:<14}  fila {fila:<4}  {estado_antes} {flecha} {nuevo_estado}")
        aplicadas += 1

    # 7) Guardar (o no)
    log("-" * 70)
    log(f"\n[6] Resumen:")
    log(f"  Aplicadas:      {aplicadas}")
    log(f"  Saltadas (ya):  {saltadas}")
    log(f"  No encontradas: {no_encontradas}")

    if dry_run:
        log("\n>> DRY-RUN: cambios NO guardados.")
    else:
        log("\n[7] Guardando Excel...")
        guardar_excel_con_retry(wb, EXCEL_PATH)
        log(f"  Guardado: {EXCEL_PATH}")

    log("\n" + "=" * 70)
    log("LISTO.")
    log("=" * 70)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="Simula sin guardar")
    args = ap.parse_args()
    main(dry_run=args.dry_run)
