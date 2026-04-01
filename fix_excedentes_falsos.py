"""
Fix: Corregir excedentes falsos de C-93-2024 y C-2609-2025.

C-93-2024: Saldo real es $4,809,783 (no $210M). Tipo A con "demandado".
C-2609-2025: Credito real $417M > acta $180M. Sin excedente -> ELIMINAR.

Uso: python fix_excedentes_falsos.py
"""

import openpyxl
from config import EXCEL_MADRE

print(f"Leyendo {EXCEL_MADRE}...")
wb = openpyxl.load_workbook(EXCEL_MADRE)

ws = wb["_datos_internos"]
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

col = {h: i + 1 for i, h in enumerate(headers) if h}

cambios = 0
for row in range(2, ws.max_row + 1):
    rol = str(ws.cell(row=row, column=col["ROL"]).value or "").strip()
    anio = str(ws.cell(row=row, column=col["AÑO"]).value or "").strip()

    if rol == "93" and anio == "2024":
        ws.cell(row=row, column=col["monto_liquidacion_saldo"]).value = "4809783"
        if "monto_credito_liquidado" in col:
            ws.cell(row=row, column=col["monto_credito_liquidado"]).value = ""
        ws.cell(row=row, column=col["estado"]).value = "EXCEDENTE_CONFIRMADO"
        ws.cell(row=row, column=col["log_decision"]).value = (
            "FIX: Saldo corregido. SALDO A FAVOR DEL DEMANDADO $4,809,783 (folio 196)"
        )
        print("  FIX: C-93-2024 saldo corregido $210M -> $4.8M")
        cambios += 1

    elif rol == "2609" and anio == "2025":
        ws.cell(row=row, column=col["estado"]).value = "ELIMINAR"
        ws.cell(row=row, column=col["monto_liquidacion_saldo"]).value = ""
        if "monto_credito_liquidado" in col:
            ws.cell(row=row, column=col["monto_credito_liquidado"]).value = "417475692"
        ws.cell(row=row, column=col["log_decision"]).value = (
            "FIX: Credito adeudado real $417,475,692 > acta $180,000,000. Sin excedente -> ELIMINAR"
        )
        print("  FIX: C-2609-2025 excedente falso. Credito $417M > acta $180M -> ELIMINAR")
        cambios += 1

print(f"\nTotal cambios: {cambios}")

if cambios > 0:
    for intento in range(3):
        try:
            wb.save(EXCEL_MADRE)
            print(f"Excel guardado: {EXCEL_MADRE}")
            break
        except PermissionError:
            if intento < 2:
                print("  PermissionError. Cierra el Excel y presiona Enter...")
                input()
            else:
                backup = EXCEL_MADRE.replace(".xlsx", "_backup.xlsx")
                wb.save(backup)
                print(f"  No se pudo guardar. Backup en: {backup}")

wb.close()
