"""Script one-shot complementario a fix_deuda_cero.py: recalcula la columna
_delta para causas que tienen monto_acta_remate y MONTO_DEUDA_CLP > 0 pero
_delta quedo en None (porque cuando F2 corrio originalmente la deuda era $0).

Replica EXACTAMENTE la formula del filtrador (no la importa para no acoplar):
    _delta = _deuda_a_int(monto_acta_remate) - _deuda_a_int(MONTO_DEUDA_CLP)
solo si ambos > 0, else None.

Uso:
    python fix_recalcular_delta.py            # dry-run (default)
    python fix_recalcular_delta.py --apply    # aplica cambios al Excel

Reglas:
- No modifica filtrador_saldos.py.
- Backup timestampeado del Excel antes de cualquier escritura.
- Idempotente: marca FixRecalcularDelta en log_decision -> salta.
- No cambia el estado, aunque delta sea negativo. F3 lo reevalua naturalmente.
"""
import argparse
import logging
import os
import re
import shutil
import sys
import time
from datetime import datetime

from openpyxl import load_workbook

from config import EXCEL_MADRE

SHEET = "_datos_internos"
ESTADOS_TARGET = {"PENDIENTE_LIQUIDACION", "PENDIENTE_ACTA"}
COL_ROL = "ROL"
COL_ANO = "AÑO"
COL_DEUDA = "MONTO_DEUDA_CLP"
COL_ACTA = "monto_acta_remate"
COL_DELTA = "_delta"
COL_ESTADO = "estado"
COL_LOG = "log_decision"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
)
log = logging.getLogger("fix_recalcular_delta")


# --- Replica exacta de filtrador_saldos._parsear_monto y _deuda_a_int -------
def _parsear_monto(texto_monto):
    """Parse monto chileno a int. Ej: '125.757.208' -> 125757208."""
    if not texto_monto:
        return 0
    limpio = re.sub(r"[\s$]", "", str(texto_monto))
    if "," in limpio:
        entero = limpio.split(",")[0].replace(".", "")
    else:
        entero = limpio.replace(".", "")
    try:
        return int(entero)
    except ValueError:
        return 0


def _deuda_a_int(valor):
    """Convierte MONTO_DEUDA_CLP / monto_acta_remate (str/float/int) a int."""
    if not valor or str(valor).strip() in ("", "nan", "None"):
        return 0
    try:
        return int(float(str(valor)))
    except (ValueError, TypeError):
        return _parsear_monto(str(valor))
# ----------------------------------------------------------------------------


def _es_delta_vacio(valor):
    """True si la celda _delta esta vacia (no calculado todavia)."""
    if valor is None:
        return True
    s = str(valor).strip().lower()
    return s in ("", "nan", "none")


def _backup_excel(ruta):
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    base, ext = os.path.splitext(ruta)
    backup = f"{base}_backup_fix_recalcular_delta_{stamp}{ext}"
    shutil.copy2(ruta, backup)
    return backup


def _guardar_con_retry(wb, ruta, max_intentos=3):
    for intento in range(max_intentos):
        try:
            wb.save(ruta)
            return True
        except PermissionError:
            if intento < max_intentos - 1:
                log.error("ERROR: No se puede guardar %s. Esta abierto en Excel?", ruta)
                log.error("  Cierra el archivo y presiona Enter (%d/%d)...",
                          intento + 1, max_intentos)
                try:
                    input(">>> Presiona Enter para reintentar: ")
                except EOFError:
                    time.sleep(5)
            else:
                log.error("ERROR CRITICO: No se pudo guardar despues de %d intentos.", max_intentos)
                fallback = ruta.replace(
                    ".xlsx",
                    f"_backup_{datetime.now().strftime('%H%M%S')}.xlsx",
                )
                try:
                    wb.save(fallback)
                    log.info("  Backup guardado en: %s", fallback)
                except Exception:
                    pass
    return False


def _mapa_columnas(ws):
    mapa = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            mapa[str(cell.value)] = idx
    return mapa


def _ya_aplicado(log_actual):
    if not log_actual:
        return False
    return "FixRecalcularDelta" in str(log_actual)


def _fmt_clp_signo(monto):
    """Formato '+$1.234.567' / '-$1.234.567' / '$0'."""
    abs_fmt = f"${abs(monto):,}".replace(",", ".")
    if monto > 0:
        return f"+{abs_fmt}"
    if monto < 0:
        return f"-{abs_fmt}"
    return abs_fmt


def main():
    parser = argparse.ArgumentParser(
        description="Recalcula _delta para causas con monto_acta y deuda>0 pero _delta vacio.")
    parser.add_argument("--apply", action="store_true",
                        help="Aplica cambios al Excel. Sin este flag corre en dry-run.")
    parser.add_argument("--limit", type=int, default=0,
                        help="Procesar solo las primeras N candidatas (0 = sin limite).")
    args = parser.parse_args()

    if not os.path.exists(EXCEL_MADRE):
        log.error("Excel madre no encontrado: %s", EXCEL_MADRE)
        sys.exit(2)

    modo = "APPLY" if args.apply else "DRY-RUN"
    log.info("=== Fix Recalcular Delta (%s) ===", modo)
    log.info("Excel: %s", EXCEL_MADRE)

    log.info("Cargando workbook (read-only para listar)...")
    wb_ro = load_workbook(EXCEL_MADRE, read_only=True, data_only=True)
    if SHEET not in wb_ro.sheetnames:
        log.error("Hoja %s no existe. Hojas: %s", SHEET, wb_ro.sheetnames)
        sys.exit(2)
    ws_ro = wb_ro[SHEET]
    mapa = _mapa_columnas(ws_ro)
    for col in (COL_ROL, COL_ANO, COL_DEUDA, COL_ACTA, COL_DELTA, COL_ESTADO, COL_LOG):
        if col not in mapa:
            log.error("Columna requerida no encontrada: %s. Disponibles: %s",
                      col, list(mapa.keys()))
            sys.exit(2)

    candidatas = []
    for row in ws_ro.iter_rows(min_row=2, values_only=True):
        try:
            rol = str(row[mapa[COL_ROL] - 1] or "").strip()
            ano = str(row[mapa[COL_ANO] - 1] or "").strip()
            estado = str(row[mapa[COL_ESTADO] - 1] or "").strip()
            deuda_raw = row[mapa[COL_DEUDA] - 1]
            acta_raw = row[mapa[COL_ACTA] - 1]
            delta_raw = row[mapa[COL_DELTA] - 1]
            log_dec = row[mapa[COL_LOG] - 1]
        except (IndexError, KeyError):
            continue
        if not rol or not ano:
            continue
        if estado not in ESTADOS_TARGET:
            continue
        if not _es_delta_vacio(delta_raw):
            continue
        deuda_int = _deuda_a_int(deuda_raw)
        acta_int = _deuda_a_int(acta_raw)
        if deuda_int <= 0 or acta_int <= 0:
            # Mismo guard que el filtrador: ambos > 0 o queda None
            continue
        if _ya_aplicado(log_dec):
            log.info("[IDEMPOTENT] C-%s-%s ya tiene marca FixRecalcularDelta, salteo", rol, ano)
            continue
        delta = acta_int - deuda_int
        candidatas.append((rol, ano, acta_int, deuda_int, delta))
    wb_ro.close()

    log.info("Candidatas encontradas: %d", len(candidatas))
    if args.limit > 0:
        candidatas = candidatas[:args.limit]
        log.info("Limitando a las primeras %d", len(candidatas))

    if not candidatas:
        log.info("Nada que hacer. Salgo.")
        return

    for rol, ano, acta, deuda, delta in candidatas:
        log.info("[%s] C-%s-%s: delta = %s  (acta %s - deuda %s)",
                 "DRY" if not args.apply else "APPLY",
                 rol, ano,
                 _fmt_clp_signo(delta),
                 _fmt_clp_signo(acta).lstrip("+"),
                 _fmt_clp_signo(deuda).lstrip("+"))

    if not args.apply:
        log.info("")
        log.info("Dry-run completo. Para aplicar, re-correr con --apply.")
        _imprimir_resumen(candidatas, aplicadas=0)
        return

    log.info("Backup del Excel antes de escribir...")
    backup_path = _backup_excel(EXCEL_MADRE)
    log.info("  Backup: %s", backup_path)

    log.info("Cargando workbook completo (read-write)...")
    wb = load_workbook(EXCEL_MADRE)
    ws = wb[SHEET]
    mapa_rw = _mapa_columnas(ws)
    rol_idx = mapa_rw[COL_ROL]
    ano_idx = mapa_rw[COL_ANO]
    deuda_idx = mapa_rw[COL_DEUDA]
    acta_idx = mapa_rw[COL_ACTA]
    delta_idx = mapa_rw[COL_DELTA]
    estado_idx = mapa_rw[COL_ESTADO]
    log_idx = mapa_rw[COL_LOG]

    objetivo = {(r, a): (acta, deuda, delta) for (r, a, acta, deuda, delta) in candidatas}
    fecha_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    aplicadas_real = 0

    for fila in ws.iter_rows(min_row=2):
        rol_c = str(fila[rol_idx - 1].value or "").strip()
        ano_c = str(fila[ano_idx - 1].value or "").strip()
        if (rol_c, ano_c) not in objetivo:
            continue
        # Re-verificar las precondiciones por si el run cambio entre dry-run y apply
        estado_c = str(fila[estado_idx - 1].value or "").strip()
        if estado_c not in ESTADOS_TARGET:
            log.warning("[SKIP-RACE] C-%s-%s estado cambio a %s", rol_c, ano_c, estado_c)
            continue
        if not _es_delta_vacio(fila[delta_idx - 1].value):
            log.warning("[SKIP-RACE] C-%s-%s ya tiene _delta=%s",
                        rol_c, ano_c, fila[delta_idx - 1].value)
            continue
        deuda_actual = _deuda_a_int(fila[deuda_idx - 1].value)
        acta_actual = _deuda_a_int(fila[acta_idx - 1].value)
        if deuda_actual <= 0 or acta_actual <= 0:
            log.warning("[SKIP-RACE] C-%s-%s precondiciones cambiaron (acta=%d, deuda=%d)",
                        rol_c, ano_c, acta_actual, deuda_actual)
            continue
        log_actual = fila[log_idx - 1].value or ""
        if _ya_aplicado(log_actual):
            log.warning("[SKIP-RACE] C-%s-%s ya marcada FixRecalcularDelta", rol_c, ano_c)
            continue
        # Recalcular con valores actuales del Excel (no del dry-run anterior)
        delta_real = acta_actual - deuda_actual
        fila[delta_idx - 1].value = int(delta_real)
        msg = (
            f"FixRecalcularDelta {fecha_str}: delta calculado = "
            f"${delta_real:,} (acta ${acta_actual:,} - deuda ${deuda_actual:,})"
        ).replace(",", ".")
        nuevo_log = f"{log_actual} | {msg}".strip(" |") if log_actual else msg
        fila[log_idx - 1].value = nuevo_log
        aplicadas_real += 1

    log.info("Filas modificadas en memoria: %d", aplicadas_real)
    log.info("Guardando Excel con retry...")
    if not _guardar_con_retry(wb, EXCEL_MADRE):
        log.error("No se pudo guardar. Restaurar desde backup si es necesario: %s", backup_path)
        sys.exit(3)
    log.info("Excel guardado OK.")
    _imprimir_resumen(candidatas, aplicadas=aplicadas_real)


def _imprimir_resumen(candidatas, aplicadas):
    print("")
    print("=== Fix Recalcular Delta - resumen ===")
    print(f"Candidatas: {len(candidatas)}")
    print(f"Aplicadas:  {aplicadas}")
    if candidatas:
        print("")
        print("Deltas calculados:")
        for rol, ano, acta, deuda, delta in candidatas:
            print(f"  C-{rol}-{ano}: {_fmt_clp_signo(delta)}  "
                  f"(acta ${acta:,}, deuda ${deuda:,})".replace(",", "."))


if __name__ == "__main__":
    main()
