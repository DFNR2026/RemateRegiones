"""Script one-shot: refresca MONTO_DEUDA_CLP para causas PENDIENTE_LIQUIDACION
con deuda=$0 cuyo PDF de mandamiento ahora si existe en Descargas/.

Bug: al importar del reporte M5, M2 todavia no habia bajado el PDF, M3 no
extrajo monto y la causa entro al Excel madre con deuda=$0. Las semanas
siguientes el filtrador no refresca el monto -> queda zombi en $0.

Uso:
    python fix_deuda_cero.py            # dry-run (default)
    python fix_deuda_cero.py --apply    # aplica cambios al Excel

Reglas:
- No modifica filtrador_saldos.py ni modulo3_extractor.py (solo importa de M3).
- Backup timestampeado del Excel antes de cualquier escritura.
- Idempotente: si una causa ya tiene MONTO_DEUDA_CLP > 0 no se procesa.
"""
import argparse
import logging
import os
import shutil
import sys
import time
from datetime import datetime

from openpyxl import load_workbook

from config import EXCEL_MADRE, BASE_DIR
from modulo3_extractor import (
    _extraer_de_mandamiento,
    _extraer_texto_pdf,
    obtener_uf_hoy,
)

DESCARGAS_DIR = os.path.join(BASE_DIR, "Descargas")
SHEET = "_datos_internos"
ESTADO_TARGET = "PENDIENTE_LIQUIDACION"
COL_ROL = "ROL"
COL_ANO = "AÑO"
COL_DEUDA = "MONTO_DEUDA_CLP"
COL_ESTADO = "estado"
COL_LOG = "log_decision"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
)
log = logging.getLogger("fix_deuda_cero")


def _es_cero(valor):
    """True si la celda representa deuda cero/vacia/None."""
    if valor is None:
        return True
    if isinstance(valor, (int, float)):
        try:
            return float(valor) == 0.0
        except Exception:
            return False
    s = str(valor).strip().lower()
    if s in ("", "0", "0.0", "nan", "none"):
        return True
    return False


def _backup_excel(ruta):
    """Crea copia con sufijo _backup_fix_deuda_cero_AAAAMMDD_HHMM.xlsx."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    base, ext = os.path.splitext(ruta)
    backup = f"{base}_backup_fix_deuda_cero_{stamp}{ext}"
    shutil.copy2(ruta, backup)
    return backup


def _guardar_con_retry(wb, ruta, max_intentos=3):
    """Replica del patron de filtrador_saldos._guardar_excel_con_retry."""
    for intento in range(max_intentos):
        try:
            wb.save(ruta)
            return True
        except PermissionError:
            if intento < max_intentos - 1:
                log.error("ERROR: No se puede guardar %s. Esta abierto en Excel?", ruta)
                log.error("  Cierra el archivo y presiona Enter (%d/%d)...",
                          intento + 1, max_intentos)
                try:
                    input(">>> Presiona Enter para reintentar: ")
                except EOFError:
                    time.sleep(5)
            else:
                log.error("ERROR CRITICO: No se pudo guardar despues de %d intentos.", max_intentos)
                fallback = ruta.replace(
                    ".xlsx",
                    f"_backup_{datetime.now().strftime('%H%M%S')}.xlsx",
                )
                try:
                    wb.save(fallback)
                    log.info("  Backup guardado en: %s", fallback)
                except Exception:
                    pass
    return False


def _mapa_columnas(ws):
    """Retorna dict {nombre_col: indice_1based} leyendo la fila 1."""
    mapa = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            mapa[str(cell.value)] = idx
    return mapa


def _calcular_monto_clp(valor, moneda):
    """Convierte (valor, moneda) -> int CLP. Si UF -> multiplica por UF actual."""
    if moneda == "UF":
        uf = obtener_uf_hoy()
        return int(round(float(valor) * float(uf)))
    return int(float(valor))


def _ya_aplicado(log_actual):
    """True si ya hay marca FixDeudaCero (idempotencia)."""
    if not log_actual:
        return False
    return "FixDeudaCero" in str(log_actual)


def main():
    parser = argparse.ArgumentParser(description="Refresca MONTO_DEUDA_CLP para causas PENDIENTE_LIQUIDACION con deuda=$0.")
    parser.add_argument("--apply", action="store_true",
                        help="Aplica cambios al Excel. Sin este flag corre en dry-run.")
    parser.add_argument("--limit", type=int, default=0,
                        help="Procesar solo las primeras N candidatas (0 = sin limite).")
    args = parser.parse_args()

    if not os.path.exists(EXCEL_MADRE):
        log.error("Excel madre no encontrado: %s", EXCEL_MADRE)
        sys.exit(2)

    modo = "APPLY" if args.apply else "DRY-RUN"
    log.info("=== Fix Deuda Cero (%s) ===", modo)
    log.info("Excel: %s", EXCEL_MADRE)
    log.info("PDFs:  %s", DESCARGAS_DIR)

    log.info("Cargando workbook (read-only para listar)...")
    wb_ro = load_workbook(EXCEL_MADRE, read_only=True, data_only=True)
    if SHEET not in wb_ro.sheetnames:
        log.error("Hoja %s no existe en %s. Hojas: %s", SHEET, EXCEL_MADRE, wb_ro.sheetnames)
        sys.exit(2)
    ws_ro = wb_ro[SHEET]
    mapa = _mapa_columnas(ws_ro)
    for col in (COL_ROL, COL_ANO, COL_DEUDA, COL_ESTADO, COL_LOG):
        if col not in mapa:
            log.error("Columna requerida no encontrada: %s. Disponibles: %s", col, list(mapa.keys()))
            sys.exit(2)

    candidatas = []
    for row in ws_ro.iter_rows(min_row=2, values_only=True):
        try:
            rol = str(row[mapa[COL_ROL] - 1] or "").strip()
            ano = str(row[mapa[COL_ANO] - 1] or "").strip()
            estado = str(row[mapa[COL_ESTADO] - 1] or "").strip()
            deuda = row[mapa[COL_DEUDA] - 1]
            log_dec = row[mapa[COL_LOG] - 1]
        except (IndexError, KeyError):
            continue
        if not rol or not ano:
            continue
        if estado != ESTADO_TARGET:
            continue
        if not _es_cero(deuda):
            continue
        if _ya_aplicado(log_dec):
            log.info("[IDEMPOTENT] %s-%s ya tiene marca FixDeudaCero, salteo", rol, ano)
            continue
        candidatas.append((rol, ano))
    wb_ro.close()

    log.info("Candidatas encontradas: %d", len(candidatas))
    if args.limit > 0:
        candidatas = candidatas[:args.limit]
        log.info("Limitando a las primeras %d", len(candidatas))

    if not candidatas:
        log.info("Nada que hacer. Salgo.")
        return

    actualizadas = []
    sin_pdf = []
    fallo_m3 = []

    for rol, ano in candidatas:
        nombre_pdf = f"C-{rol}-{ano}_MANDAMIENTO.pdf"
        ruta_pdf = os.path.join(DESCARGAS_DIR, nombre_pdf)
        if not os.path.exists(ruta_pdf):
            log.info("[SKIP] C-%s-%s: PDF no encontrado en %s", rol, ano, ruta_pdf)
            sin_pdf.append((rol, ano))
            continue
        try:
            texto = _extraer_texto_pdf(ruta_pdf)
            orig, valor, moneda = _extraer_de_mandamiento(texto)
        except Exception as e:
            log.warning("[SKIP] C-%s-%s: error extrayendo PDF: %s", rol, ano, e)
            fallo_m3.append((rol, ano))
            continue
        if not orig or valor is None:
            log.info("[SKIP] C-%s-%s: M3 no pudo extraer monto", rol, ano)
            fallo_m3.append((rol, ano))
            continue
        try:
            monto_clp = _calcular_monto_clp(valor, moneda)
        except Exception as e:
            log.warning("[SKIP] C-%s-%s: no pudo convertir %s %s a CLP: %s",
                        rol, ano, valor, moneda, e)
            fallo_m3.append((rol, ano))
            continue
        if monto_clp <= 0:
            log.info("[SKIP] C-%s-%s: M3 devolvio monto <= 0 (%d)", rol, ano, monto_clp)
            fallo_m3.append((rol, ano))
            continue
        log.info("[%s] C-%s-%s: $0 -> $%s  (%s, fuente: %s)",
                 "DRY" if not args.apply else "APPLY",
                 rol, ano, f"{monto_clp:,}".replace(",", "."),
                 moneda, nombre_pdf)
        actualizadas.append((rol, ano, monto_clp, nombre_pdf))

    if not actualizadas:
        log.info("Sin actualizaciones para aplicar.")
        _imprimir_resumen(len(candidatas), sin_pdf, fallo_m3, [])
        return

    if not args.apply:
        log.info("")
        log.info("Dry-run completo. Para aplicar, re-correr con --apply.")
        _imprimir_resumen(len(candidatas), sin_pdf, fallo_m3, actualizadas)
        return

    log.info("Backup del Excel antes de escribir...")
    backup_path = _backup_excel(EXCEL_MADRE)
    log.info("  Backup: %s", backup_path)

    log.info("Cargando workbook completo (read-write)...")
    wb = load_workbook(EXCEL_MADRE)
    ws = wb[SHEET]
    mapa_rw = _mapa_columnas(ws)
    rol_idx = mapa_rw[COL_ROL]
    ano_idx = mapa_rw[COL_ANO]
    deuda_idx = mapa_rw[COL_DEUDA]
    estado_idx = mapa_rw[COL_ESTADO]
    log_idx = mapa_rw[COL_LOG]

    objetivo = {(r, a): (m, p) for (r, a, m, p) in actualizadas}
    fecha_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    aplicadas_real = 0

    for fila in ws.iter_rows(min_row=2):
        rol_c = str(fila[rol_idx - 1].value or "").strip()
        ano_c = str(fila[ano_idx - 1].value or "").strip()
        estado_c = str(fila[estado_idx - 1].value or "").strip()
        if (rol_c, ano_c) not in objetivo:
            continue
        if estado_c != ESTADO_TARGET:
            log.warning("[SKIP-RACE] C-%s-%s cambio de estado a %s entre dry-run y apply",
                        rol_c, ano_c, estado_c)
            continue
        deuda_actual = fila[deuda_idx - 1].value
        if not _es_cero(deuda_actual):
            log.warning("[SKIP-RACE] C-%s-%s ya tiene deuda=%s, salteo",
                        rol_c, ano_c, deuda_actual)
            continue
        log_actual = fila[log_idx - 1].value or ""
        if _ya_aplicado(log_actual):
            log.warning("[SKIP-RACE] C-%s-%s ya marcada FixDeudaCero", rol_c, ano_c)
            continue
        monto, nombre_pdf = objetivo[(rol_c, ano_c)]
        fila[deuda_idx - 1].value = monto
        msg = (
            f"FixDeudaCero {fecha_str}: deuda actualizada de $0 a "
            f"${monto:,} (M3 sobre {nombre_pdf})"
        ).replace(",", ".")
        nuevo_log = f"{log_actual} | {msg}".strip(" |") if log_actual else msg
        fila[log_idx - 1].value = nuevo_log
        aplicadas_real += 1

    log.info("Filas modificadas en memoria: %d", aplicadas_real)
    log.info("Guardando Excel con retry...")
    if not _guardar_con_retry(wb, EXCEL_MADRE):
        log.error("No se pudo guardar. Restaurar desde backup si es necesario: %s", backup_path)
        sys.exit(3)
    log.info("Excel guardado OK.")
    _imprimir_resumen(len(candidatas), sin_pdf, fallo_m3, actualizadas)


def _imprimir_resumen(n_candidatas, sin_pdf, fallo_m3, actualizadas):
    print("")
    print("=== Fix Deuda Cero - resumen ===")
    print(f"Candidatas encontradas: {n_candidatas}")
    print(f"PDFs no encontrados:    {len(sin_pdf)}")
    print(f"M3 fallo:               {len(fallo_m3)}")
    print(f"Actualizadas:           {len(actualizadas)}")
    if actualizadas:
        print("")
        print("Causas actualizadas:")
        for rol, ano, monto, _ in actualizadas:
            monto_fmt = f"${monto:,}".replace(",", ".")
            print(f"  C-{rol}-{ano}: $0 -> {monto_fmt}")
    if sin_pdf:
        print("")
        print("PDFs faltantes:")
        for rol, ano in sin_pdf:
            print(f"  C-{rol}-{ano}")
    if fallo_m3:
        print("")
        print("Causas con fallo M3:")
        for rol, ano in fallo_m3:
            print(f"  C-{rol}-{ano}")


if __name__ == "__main__":
    main()
