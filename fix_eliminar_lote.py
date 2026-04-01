"""
fix_eliminar_lote.py — Eliminar 14 causas (8 revision abogados + 6 verificacion manual)

Uso: python fix_eliminar_lote.py
"""

import time
import openpyxl
from config import EXCEL_MADRE

# 8 causas determinadas por revision de abogados (sin futuro comercial):
eliminar_abogados = [
    (3323, 2024, "Revision abogados: causa sin futuro comercial"),
    (7205, 2024, "Revision abogados: causa sin futuro comercial"),
    (1810, 2025, "Revision abogados: causa sin futuro comercial"),
    (5950, 2020, "Revision abogados: causa sin futuro comercial"),
    (2601, 2023, "Revision abogados: causa sin futuro comercial"),
    (1833, 2025, "Revision abogados: causa sin futuro comercial"),
    (4885, 2023, "Revision abogados: causa sin futuro comercial"),
    (2431, 2024, "Revision abogados: causa sin futuro comercial"),
]

# 6 causas verificadas manualmente (cargo al credito / sin excedente / suspension):
eliminar_manual = [
    (7366, 2024, "CCC en acta (folio 77). Banco Itau-Chile adjudico por $321,674,000 con cargo a sus creditos"),
    (93, 2024, "CCC en acta (folio 75). FactorOne S.A. adjudico por $215,000,000 con cargo a su credito. Saldo $4.8M no perseguible (tribunal notifico al ejecutado)"),
    (3535, 2025, "CCC en acta (folio 21, tramite Actuacion). Banco Santander adjudico por $100,417,547 con cargo a su credito"),
    (4836, 2022, "CCC en acta (folio 233, tramite Actuacion). Lote C $30,730,735 + Lote D $36,863,881 ambos con cargo al credito"),
    (867, 2024, "Suspension de remate (folio 220). Sin acta de remate"),
    (1138, 2024, "Adjudicacion por tercero (Empresa Astore) $110,366,666 pero credito adeudado $166,698,818. Sin excedente (acta < deuda)"),
]

# C-867: limpiar monto_credito_liquidado (liquidacion para remate que no ocurrio)
_LIMPIAR_CREDITO = {(867, 2024)}

TODAS = eliminar_abogados + eliminar_manual


def main():
    print(f"Leyendo Excel madre: {EXCEL_MADRE}")
    wb = openpyxl.load_workbook(EXCEL_MADRE)
    ws = wb["_datos_internos"]

    # Encontrar indices de columnas por header (fila 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx

    col_rol = headers.get("ROL") or headers.get("Rol")
    col_anio = headers.get("AÑO") or headers.get("Año")
    col_estado = headers.get("estado") or headers.get("Estado")
    col_log = headers.get("log_decision") or headers.get("Decision")
    col_saldo = headers.get("monto_liquidacion_saldo")
    col_credito = headers.get("monto_credito_liquidado")

    if not all([col_rol, col_anio, col_estado, col_log]):
        print("ERROR: No se encontraron columnas requeridas")
        print(f"  Headers encontrados: {list(headers.keys())}")
        return

    cnt_eliminadas = 0
    cnt_skip = 0

    for rol, anio, razon in TODAS:
        found = False
        for row_idx in range(2, ws.max_row + 1):
            cell_rol = ws.cell(row=row_idx, column=col_rol).value
            cell_anio = ws.cell(row=row_idx, column=col_anio).value

            try:
                r = int(float(str(cell_rol)))
            except (ValueError, TypeError):
                continue
            try:
                a = int(float(str(cell_anio)))
            except (ValueError, TypeError):
                continue

            if r == rol and a == anio:
                found = True
                estado_actual = str(ws.cell(row=row_idx, column=col_estado).value or "").strip()

                if estado_actual == "ELIMINAR":
                    print(f"SKIP: C-{rol}-{anio} ya es ELIMINAR")
                    cnt_skip += 1
                else:
                    ws.cell(row=row_idx, column=col_estado).value = "ELIMINAR"
                    ws.cell(row=row_idx, column=col_log).value = (
                        f"ELIMINAR: {razon}. [Anterior: {estado_actual}]"
                    )

                    # Limpiar monto_liquidacion_saldo si tenia valor
                    if col_saldo:
                        val_saldo = ws.cell(row=row_idx, column=col_saldo).value
                        if val_saldo and str(val_saldo).strip():
                            print(f"  Limpiando monto_liquidacion_saldo: {val_saldo}")
                            ws.cell(row=row_idx, column=col_saldo).value = ""

                    # Limpiar monto_credito_liquidado solo para causas especificas
                    if col_credito and (rol, anio) in _LIMPIAR_CREDITO:
                        val_credito = ws.cell(row=row_idx, column=col_credito).value
                        if val_credito and str(val_credito).strip():
                            print(f"  Limpiando monto_credito_liquidado: {val_credito}")
                            ws.cell(row=row_idx, column=col_credito).value = ""

                    print(f"ELIMINAR: C-{rol}-{anio} ({estado_actual} -> ELIMINAR): {razon}")
                    cnt_eliminadas += 1
                break

        if not found:
            print(f"WARNING: C-{rol}-{anio} no encontrada en Excel")

    # Guardar con retry
    for intento in range(1, 4):
        try:
            wb.save(EXCEL_MADRE)
            print(f"\nExcel guardado: {EXCEL_MADRE}")
            break
        except PermissionError:
            if intento < 3:
                print(f"PermissionError (intento {intento}/3). Cierra el Excel y espera...")
                time.sleep(3)
            else:
                print("ERROR: No se pudo guardar. Cierra el Excel e intenta de nuevo.")

    wb.close()

    print(f"\nResumen: {cnt_eliminadas} causas eliminadas, {cnt_skip} ya estaban eliminadas")


if __name__ == "__main__":
    main()
